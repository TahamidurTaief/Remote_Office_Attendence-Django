import io
import csv
from decimal import Decimal
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


def generate_payslip_pdf(calculation):
    """
    Generates a high-quality ReportLab PDF for an EmployeePayrollCalculation.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    story = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        textColor=colors.HexColor('#1E293B'),
        alignment=1
    )
    sub_title_style = ParagraphStyle(
        'SubTitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#64748B'),
        alignment=1
    )
    section_heading = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading3'],
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#1E293B')
    )
    cell_bold = ParagraphStyle(
        'CellBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#1E293B')
    )
    cell_normal = ParagraphStyle(
        'CellNormal',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#334155')
    )
    cell_right = ParagraphStyle(
        'CellRight',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#334155'),
        alignment=2
    )
    cell_right_bold = ParagraphStyle(
        'CellRightBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#1E293B'),
        alignment=2
    )

    # 1. Header
    story.append(Paragraph("SIGNTECH COMMUNICATIONS", title_style))
    story.append(Paragraph("Smart Attendance & Workforce Management", sub_title_style))
    period_str = f"{calculation.payroll_run.period_start.strftime('%B %d, %Y')} to {calculation.payroll_run.period_end.strftime('%B %d, %Y')}"
    story.append(Paragraph(f"PAYSLIP FOR THE PERIOD: <b>{period_str}</b>", sub_title_style))
    story.append(Spacer(1, 15))

    # 2. Employee Details Block
    emp = calculation.employee
    dept_name = emp.department.name if getattr(emp, 'department', None) else 'General'
    desig_name = emp.designation.name if getattr(emp, 'designation', None) else 'Staff'
    branch_name = emp.branch.name if getattr(emp, 'branch', None) else 'Head Office'
    join_date = emp.joined_date.strftime('%d/%m/%Y') if emp.joined_date else 'N/A'

    emp_info_data = [
        [
            Paragraph("<b>Employee ID:</b>", cell_bold), Paragraph(emp.employee_number, cell_normal),
            Paragraph("<b>Department:</b>", cell_bold), Paragraph(dept_name, cell_normal)
        ],
        [
            Paragraph("<b>Employee Name:</b>", cell_bold), Paragraph(emp.get_full_name(), cell_normal),
            Paragraph("<b>Designation:</b>", cell_bold), Paragraph(desig_name, cell_normal)
        ],
        [
            Paragraph("<b>Branch:</b>", cell_bold), Paragraph(branch_name, cell_normal),
            Paragraph("<b>Joining Date:</b>", cell_bold), Paragraph(join_date, cell_normal)
        ],
        [
            Paragraph("<b>Payment Mode:</b>", cell_bold), Paragraph(calculation.get_payment_mode_display(), cell_normal),
            Paragraph("<b>Bank Account:</b>", cell_bold), Paragraph(emp.bank_account or 'N/A', cell_normal)
        ],
    ]
    emp_table = Table(emp_info_data, colWidths=[100, 160, 100, 160])
    emp_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(emp_table)
    story.append(Spacer(1, 10))

    # 3. Attendance Snapshot Summary
    att_summary_data = [
        [
            Paragraph("<b>Present Days:</b>", cell_bold), Paragraph(str(calculation.source_total_present_days), cell_normal),
            Paragraph("<b>Paid Leave Days:</b>", cell_bold), Paragraph(str(calculation.source_total_approved_leave_days), cell_normal),
            Paragraph("<b>Unpaid Absent Days:</b>", cell_bold), Paragraph(str(calculation.unpaid_absent_days), cell_normal),
            Paragraph("<b>OT Hours:</b>", cell_bold), Paragraph(f"{calculation.ot_hours} hrs", cell_normal)
        ]
    ]
    att_table = Table(att_summary_data, colWidths=[80, 50, 85, 45, 110, 45, 60, 45])
    att_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#EFF6FF')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#BFDBFE')),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#DBEAFE')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(att_table)
    story.append(Spacer(1, 12))

    # 4. Earnings vs Deductions Breakdown Table
    snapshot = calculation.structure_snapshot or {}
    components = snapshot.get('components', [])

    earnings_list = []
    deductions_list = []

    for comp in components:
        comp_type = comp.get('type')
        name = comp.get('name', comp.get('code', ''))
        amount = Decimal(str(comp.get('amount', '0.00')))
        if comp_type == 'earning':
            earnings_list.append((name, amount))
        elif comp_type == 'deduction':
            deductions_list.append((name, amount))

    # Add overtime if not already in components
    if calculation.ot_amount > Decimal('0.00'):
        if not any('overtime' in name.lower() or 'ot' in name.lower() for name, _ in earnings_list):
            earnings_list.append((f"Overtime ({calculation.ot_hours} hrs)", calculation.ot_amount))

    # Add absence deduction if not already in components
    if calculation.absence_deduction > Decimal('0.00'):
        if not any('absence' in name.lower() for name, _ in deductions_list):
            deductions_list.append((f"Absence Deduction ({calculation.unpaid_absent_days} days)", calculation.absence_deduction))

    if calculation.other_deduction > Decimal('0.00'):
        if not any('other' in name.lower() for name, _ in deductions_list):
            deductions_list.append(("Other Deduction", calculation.other_deduction))

    # Build side-by-side rows
    max_rows = max(len(earnings_list), len(deductions_list), 1)
    table_rows = [
        [
            Paragraph("<b>EARNINGS</b>", cell_bold),
            Paragraph("<b>AMOUNT (BDT)</b>", cell_right_bold),
            Paragraph("<b>DEDUCTIONS</b>", cell_bold),
            Paragraph("<b>AMOUNT (BDT)</b>", cell_right_bold)
        ]
    ]

    for i in range(max_rows):
        e_name, e_amt = earnings_list[i] if i < len(earnings_list) else ("", "")
        d_name, d_amt = deductions_list[i] if i < len(deductions_list) else ("", "")

        table_rows.append([
            Paragraph(e_name, cell_normal),
            Paragraph(f"{e_amt:,.2f}" if e_amt != "" else "", cell_right),
            Paragraph(d_name, cell_normal),
            Paragraph(f"{d_amt:,.2f}" if d_amt != "" else "", cell_right)
        ])

    # Totals Row
    table_rows.append([
        Paragraph("<b>Total Earnings</b>", cell_bold),
        Paragraph(f"<b>{calculation.total_earnings:,.2f}</b>", cell_right_bold),
        Paragraph("<b>Total Deductions</b>", cell_bold),
        Paragraph(f"<b>{calculation.total_deductions:,.2f}</b>", cell_right_bold)
    ])

    breakdown_table = Table(table_rows, colWidths=[170, 90, 170, 90])
    breakdown_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (2, 0), (3, 0), colors.HexColor('#E2E8F0')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#94A3B8')),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(breakdown_table)
    story.append(Spacer(1, 12))

    # 5. Net Payable Banner
    net_rows = [
        [
            Paragraph("<b>NET PAYABLE AMOUNT:</b>", ParagraphStyle('NetLabel', fontName='Helvetica-Bold', fontSize=10, textColor=colors.HexColor('#0F172A'))),
            Paragraph(f"<b>BDT {calculation.net_payable:,.2f}</b>", ParagraphStyle('NetVal', fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor('#0F172A'), alignment=2))
        ]
    ]
    if calculation.payment_mode == 'split':
        net_rows.append([
            Paragraph("<b>Disbursement Breakdown:</b>", cell_bold),
            Paragraph(f"Bank: BDT {calculation.bank_payable:,.2f} | Cash: BDT {calculation.cash_payable:,.2f}", cell_right_bold)
        ])
    elif calculation.payment_mode == 'bank':
        net_rows.append([
            Paragraph("<b>Disbursement:</b>", cell_bold),
            Paragraph(f"100% Bank Transfer (BDT {calculation.bank_payable:,.2f})", cell_right)
        ])
    else:
        net_rows.append([
            Paragraph("<b>Disbursement:</b>", cell_bold),
            Paragraph(f"100% Cash / Cheque (BDT {calculation.cash_payable:,.2f})", cell_right)
        ])

    net_table = Table(net_rows, colWidths=[260, 260])
    net_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#DCFCE7')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#86EFAC')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(net_table)
    story.append(Spacer(1, 40))

    # 6. Signatures
    sig_data = [
        [
            Paragraph("____________________________<br/><b>Prepared By</b>", ParagraphStyle('Sig1', fontName='Helvetica', fontSize=8, alignment=0)),
            Paragraph("____________________________<br/><b>Checked By (HR)</b>", ParagraphStyle('Sig2', fontName='Helvetica', fontSize=8, alignment=1)),
            Paragraph("____________________________<br/><b>Employee Signature</b>", ParagraphStyle('Sig3', fontName='Helvetica', fontSize=8, alignment=2)),
        ]
    ]
    sig_table = Table(sig_data, colWidths=[170, 180, 170])
    sig_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(sig_table)

    # Build document
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def export_payroll_register_excel(payroll_run, calculations):
    wb = openpyxl.Workbook()
    ws = wb.active
    period_label = payroll_run.period_start.strftime('%B %Y')
    ws.title = f"Payroll {payroll_run.period_start.strftime('%b %Y')}"

    # Styles
    title_font = Font(name='Calibri', size=14, bold=True, color='1E293B')
    subtitle_font = Font(name='Calibri', size=10, italic=True, color='64748B')
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    total_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    total_font = Font(name='Calibri', size=10, bold=True, color='0F172A')
    regular_font = Font(name='Calibri', size=9)
    bold_font = Font(name='Calibri', size=9, bold=True)

    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title rows
    ws.merge_cells('A1:S1')
    ws['A1'] = f"SIGNTECH COMMUNICATIONS - MONTHLY PAYROLL REGISTER ({period_label.upper()})"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:S2')
    ws['A2'] = f"Status: {payroll_run.get_status_display()} | Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = [
        'Sl', 'Employee ID', 'Name', 'Department', 'Designation', 'Gross Salary',
        'Basic', 'Allowances', 'OT Hours', 'OT Amount', 'Adjustments (+)', 'Total Earnings',
        'PF Deduction', 'Absence Deduction', 'Adjustments (-)', 'Total Deductions',
        'Net Payable', 'Payment Mode', 'Bank Payable', 'Cash Payable'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    row_num = 5
    tot_gross = Decimal('0.00')
    tot_earnings = Decimal('0.00')
    tot_ot_hours = Decimal('0.00')
    tot_ot_amt = Decimal('0.00')
    tot_deductions = Decimal('0.00')
    tot_pf = Decimal('0.00')
    tot_absence = Decimal('0.00')
    tot_net = Decimal('0.00')
    tot_bank = Decimal('0.00')
    tot_cash = Decimal('0.00')

    for idx, calc in enumerate(calculations, 1):
        emp = calc.employee
        dept_name = emp.department.name if getattr(emp, 'department', None) else 'General'
        desig_name = emp.designation.name if getattr(emp, 'designation', None) else 'Staff'

        # Parse components from structure_snapshot
        snapshot = calc.structure_snapshot or {}
        components = snapshot.get('components', [])
        basic_amt = Decimal('0.00')
        allowances_amt = Decimal('0.00')
        adj_plus = Decimal('0.00')
        adj_minus = Decimal('0.00')
        pf_amt = Decimal('0.00')

        for c in components:
            c_type = c.get('type')
            code = c.get('code', '')
            amt = Decimal(str(c.get('amount', '0.00')))
            if c.get('is_adjustment'):
                if c_type == 'earning':
                    adj_plus += amt
                else:
                    adj_minus += amt
            elif code == 'BASIC':
                basic_amt += amt
            elif c.get('is_pf') or code == 'PF':
                pf_amt += amt
            elif c_type == 'earning':
                allowances_amt += amt

        ws.cell(row=row_num, column=1, value=idx).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=2, value=emp.employee_number).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=3, value=emp.get_full_name())
        ws.cell(row=row_num, column=4, value=dept_name)
        ws.cell(row=row_num, column=5, value=desig_name)
        ws.cell(row=row_num, column=6, value=float(calc.gross_salary))
        ws.cell(row=row_num, column=7, value=float(basic_amt))
        ws.cell(row=row_num, column=8, value=float(allowances_amt))
        ws.cell(row=row_num, column=9, value=float(calc.ot_hours))
        ws.cell(row=row_num, column=10, value=float(calc.ot_amount))
        ws.cell(row=row_num, column=11, value=float(adj_plus))
        ws.cell(row=row_num, column=12, value=float(calc.total_earnings))
        ws.cell(row=row_num, column=13, value=float(pf_amt))
        ws.cell(row=row_num, column=14, value=float(calc.absence_deduction))
        ws.cell(row=row_num, column=15, value=float(adj_minus))
        ws.cell(row=row_num, column=16, value=float(calc.total_deductions))
        ws.cell(row=row_num, column=17, value=float(calc.net_payable))
        ws.cell(row=row_num, column=18, value=calc.get_payment_mode_display()).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=19, value=float(calc.bank_payable))
        ws.cell(row=row_num, column=20, value=float(calc.cash_payable))

        for c_idx in range(1, 21):
            cell = ws.cell(row=row_num, column=c_idx)
            cell.font = regular_font
            cell.border = border
            if c_idx in [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 19, 20]:
                cell.number_format = '#,##0.00'

        tot_gross += calc.gross_salary
        tot_earnings += calc.total_earnings
        tot_ot_hours += calc.ot_hours
        tot_ot_amt += calc.ot_amount
        tot_deductions += calc.total_deductions
        tot_pf += pf_amt
        tot_absence += calc.absence_deduction
        tot_net += calc.net_payable
        tot_bank += calc.bank_payable
        tot_cash += calc.cash_payable

        row_num += 1

    # Totals Row
    ws.cell(row=row_num, column=1, value="TOTAL")
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
    ws.cell(row=row_num, column=6, value=float(tot_gross))
    ws.cell(row=row_num, column=9, value=float(tot_ot_hours))
    ws.cell(row=row_num, column=10, value=float(tot_ot_amt))
    ws.cell(row=row_num, column=12, value=float(tot_earnings))
    ws.cell(row=row_num, column=13, value=float(tot_pf))
    ws.cell(row=row_num, column=14, value=float(tot_absence))
    ws.cell(row=row_num, column=16, value=float(tot_deductions))
    ws.cell(row=row_num, column=17, value=float(tot_net))
    ws.cell(row=row_num, column=19, value=float(tot_bank))
    ws.cell(row=row_num, column=20, value=float(tot_cash))

    for c_idx in range(1, 21):
        cell = ws.cell(row=row_num, column=c_idx)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        if c_idx in [6, 9, 10, 12, 13, 14, 16, 17, 19, 20]:
            cell.number_format = '#,##0.00'

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_payroll_register_csv(payroll_run, calculations):
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    writer.writerow([f"SIGNTECH COMMUNICATIONS - PAYROLL REGISTER ({payroll_run.period_start.strftime('%B %Y')})"])
    writer.writerow([f"Status: {payroll_run.get_status_display()}"])
    writer.writerow([])

    headers = [
        'Sl', 'Employee ID', 'Name', 'Department', 'Designation', 'Gross Salary',
        'Basic', 'Allowances', 'OT Hours', 'OT Amount', 'Adjustments (+)', 'Total Earnings',
        'PF Deduction', 'Absence Deduction', 'Adjustments (-)', 'Total Deductions',
        'Net Payable', 'Payment Mode', 'Bank Payable', 'Cash Payable'
    ]
    writer.writerow(headers)

    for idx, calc in enumerate(calculations, 1):
        emp = calc.employee
        dept_name = emp.department.name if getattr(emp, 'department', None) else 'General'
        desig_name = emp.designation.name if getattr(emp, 'designation', None) else 'Staff'

        snapshot = calc.structure_snapshot or {}
        components = snapshot.get('components', [])
        basic_amt = Decimal('0.00')
        allowances_amt = Decimal('0.00')
        adj_plus = Decimal('0.00')
        adj_minus = Decimal('0.00')
        pf_amt = Decimal('0.00')

        for c in components:
            c_type = c.get('type')
            code = c.get('code', '')
            amt = Decimal(str(c.get('amount', '0.00')))
            if c.get('is_adjustment'):
                if c_type == 'earning':
                    adj_plus += amt
                else:
                    adj_minus += amt
            elif code == 'BASIC':
                basic_amt += amt
            elif c.get('is_pf') or code == 'PF':
                pf_amt += amt
            elif c_type == 'earning':
                allowances_amt += amt

        writer.writerow([
            idx,
            emp.employee_number,
            emp.get_full_name(),
            dept_name,
            desig_name,
            f"{calc.gross_salary:.2f}",
            f"{basic_amt:.2f}",
            f"{allowances_amt:.2f}",
            f"{calc.ot_hours:.2f}",
            f"{calc.ot_amount:.2f}",
            f"{adj_plus:.2f}",
            f"{calc.total_earnings:.2f}",
            f"{pf_amt:.2f}",
            f"{calc.absence_deduction:.2f}",
            f"{adj_minus:.2f}",
            f"{calc.total_deductions:.2f}",
            f"{calc.net_payable:.2f}",
            calc.get_payment_mode_display(),
            f"{calc.bank_payable:.2f}",
            f"{calc.cash_payable:.2f}"
        ])

    return buffer.getvalue().encode('utf-8')


def export_payroll_register_pdf(payroll_run, calculations):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=25,
        bottomMargin=25
    )
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'LandscapeTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=13,
        leading=16,
        textColor=colors.HexColor('#1E293B'),
        alignment=1
    )
    sub_title = ParagraphStyle(
        'LandscapeSub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=11,
        textColor=colors.HexColor('#64748B'),
        alignment=1
    )
    cell_h = ParagraphStyle('CH', fontName='Helvetica-Bold', fontSize=6.5, leading=8, textColor=colors.white, alignment=1)
    cell_n = ParagraphStyle('CN', fontName='Helvetica', fontSize=6.5, leading=8, textColor=colors.HexColor('#1E293B'))
    cell_r = ParagraphStyle('CR', fontName='Helvetica', fontSize=6.5, leading=8, textColor=colors.HexColor('#1E293B'), alignment=2)
    cell_rb = ParagraphStyle('CRB', fontName='Helvetica-Bold', fontSize=6.5, leading=8, textColor=colors.HexColor('#0F172A'), alignment=2)

    story.append(Paragraph("SIGNTECH COMMUNICATIONS - MONTHLY PAYROLL REGISTER", title_style))
    story.append(Paragraph(f"Period: {payroll_run.period_start.strftime('%d/%m/%Y')} to {payroll_run.period_end.strftime('%d/%m/%Y')} | Status: {payroll_run.get_status_display()}", sub_title))
    story.append(Spacer(1, 10))

    headers = [
        Paragraph("<b>Sl</b>", cell_h),
        Paragraph("<b>ID</b>", cell_h),
        Paragraph("<b>Employee Name</b>", cell_h),
        Paragraph("<b>Department</b>", cell_h),
        Paragraph("<b>Gross</b>", cell_h),
        Paragraph("<b>Basic</b>", cell_h),
        Paragraph("<b>Allow.</b>", cell_h),
        Paragraph("<b>OT (Tk)</b>", cell_h),
        Paragraph("<b>Total Earn.</b>", cell_h),
        Paragraph("<b>PF Ded.</b>", cell_h),
        Paragraph("<b>Abs. Ded.</b>", cell_h),
        Paragraph("<b>Total Ded.</b>", cell_h),
        Paragraph("<b>Net Pay.</b>", cell_h),
        Paragraph("<b>Mode</b>", cell_h),
        Paragraph("<b>Bank Pay.</b>", cell_h),
        Paragraph("<b>Cash Pay.</b>", cell_h),
    ]

    table_data = [headers]
    tot_gross = tot_earnings = tot_ded = tot_net = tot_bank = tot_cash = Decimal('0.00')

    for idx, calc in enumerate(calculations, 1):
        emp = calc.employee
        dept_name = emp.department.name if getattr(emp, 'department', None) else 'General'

        snapshot = calc.structure_snapshot or {}
        components = snapshot.get('components', [])
        basic_amt = Decimal('0.00')
        allow_amt = Decimal('0.00')
        pf_amt = Decimal('0.00')

        for c in components:
            code = c.get('code', '')
            c_type = c.get('type')
            amt = Decimal(str(c.get('amount', '0.00')))
            if code == 'BASIC':
                basic_amt += amt
            elif c.get('is_pf') or code == 'PF':
                pf_amt += amt
            elif c_type == 'earning' and not c.get('is_adjustment'):
                allow_amt += amt

        tot_gross += calc.gross_salary
        tot_earnings += calc.total_earnings
        tot_ded += calc.total_deductions
        tot_net += calc.net_payable
        tot_bank += calc.bank_payable
        tot_cash += calc.cash_payable

        table_data.append([
            Paragraph(str(idx), cell_n),
            Paragraph(emp.employee_number, cell_n),
            Paragraph(emp.get_full_name()[:20], cell_n),
            Paragraph(dept_name[:15], cell_n),
            Paragraph(f"{calc.gross_salary:,.0f}", cell_r),
            Paragraph(f"{basic_amt:,.0f}", cell_r),
            Paragraph(f"{allow_amt:,.0f}", cell_r),
            Paragraph(f"{calc.ot_amount:,.0f}", cell_r),
            Paragraph(f"{calc.total_earnings:,.0f}", cell_rb),
            Paragraph(f"{pf_amt:,.0f}", cell_r),
            Paragraph(f"{calc.absence_deduction:,.0f}", cell_r),
            Paragraph(f"{calc.total_deductions:,.0f}", cell_rb),
            Paragraph(f"{calc.net_payable:,.0f}", cell_rb),
            Paragraph(calc.payment_mode.upper()[:4], cell_n),
            Paragraph(f"{calc.bank_payable:,.0f}", cell_r),
            Paragraph(f"{calc.cash_payable:,.0f}", cell_r),
        ])

    # Totals Row
    table_data.append([
        Paragraph("<b>TOT</b>", cell_n),
        Paragraph("", cell_n),
        Paragraph("<b>Total</b>", cell_n),
        Paragraph("", cell_n),
        Paragraph(f"{tot_gross:,.0f}", cell_rb),
        Paragraph("", cell_r),
        Paragraph("", cell_r),
        Paragraph("", cell_r),
        Paragraph(f"{tot_earnings:,.0f}", cell_rb),
        Paragraph("", cell_r),
        Paragraph("", cell_r),
        Paragraph(f"{tot_ded:,.0f}", cell_rb),
        Paragraph(f"{tot_net:,.0f}", cell_rb),
        Paragraph("", cell_n),
        Paragraph(f"{tot_bank:,.0f}", cell_rb),
        Paragraph(f"{tot_cash:,.0f}", cell_rb),
    ])

    col_widths = [18, 45, 95, 65, 48, 42, 42, 42, 52, 42, 42, 52, 55, 30, 52, 52]
    reg_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    reg_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#94A3B8')),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(reg_table)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def export_bank_report_excel(payroll_run, calculations):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Transfer"

    title_font = Font(name='Calibri', size=14, bold=True, color='1E293B')
    subtitle_font = Font(name='Calibri', size=10, italic=True, color='64748B')
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    total_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    total_font = Font(name='Calibri', size=10, bold=True, color='0F172A')
    regular_font = Font(name='Calibri', size=9)

    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:G1')
    ws['A1'] = f"BANK SALARY TRANSFER SHEET - {payroll_run.period_start.strftime('%B %Y').upper()}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:G2')
    ws['A2'] = f"Period: {payroll_run.period_start} to {payroll_run.period_end} | Status: {payroll_run.get_status_display()}"
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ['Sl', 'Employee ID', 'Employee Name', 'Department', 'Bank Name', 'Account Number', 'Net Payable (BDT)']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row_num = 5
    tot_bank = Decimal('0.00')

    for idx, calc in enumerate(calculations, 1):
        emp = calc.employee
        dept_name = emp.department.name if getattr(emp, 'department', None) else 'General'

        ws.cell(row=row_num, column=1, value=idx).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=2, value=emp.employee_number).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=3, value=emp.get_full_name())
        ws.cell(row=row_num, column=4, value=dept_name)
        ws.cell(row=row_num, column=5, value=emp.bank_name or 'City Bank Ltd.')
        ws.cell(row=row_num, column=6, value=emp.bank_account or 'N/A').alignment = Alignment(horizontal='center')
        cell_amt = ws.cell(row=row_num, column=7, value=float(calc.bank_payable))
        cell_amt.number_format = '#,##0.00'

        for c_idx in range(1, 8):
            cell = ws.cell(row=row_num, column=c_idx)
            cell.font = regular_font
            cell.border = border

        tot_bank += calc.bank_payable
        row_num += 1

    # Total row
    ws.cell(row=row_num, column=1, value="TOTAL")
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
    cell_tot = ws.cell(row=row_num, column=7, value=float(tot_bank))
    cell_tot.number_format = '#,##0.00'

    for c_idx in range(1, 8):
        cell = ws.cell(row=row_num, column=c_idx)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_bank_report_csv(payroll_run, calculations):
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    writer.writerow([f"BANK SALARY TRANSFER SHEET - {payroll_run.period_start.strftime('%B %Y')}"])
    writer.writerow([])
    writer.writerow(['Sl', 'Employee ID', 'Employee Name', 'Department', 'Bank Name', 'Account Number', 'Amount (BDT)'])

    for idx, calc in enumerate(calculations, 1):
        emp = calc.employee
        dept_name = emp.department.name if getattr(emp, 'department', None) else 'General'
        writer.writerow([
            idx,
            emp.employee_number,
            emp.get_full_name(),
            dept_name,
            emp.bank_name or 'City Bank Ltd.',
            emp.bank_account or 'N/A',
            f"{calc.bank_payable:.2f}"
        ])

    return buffer.getvalue().encode('utf-8')


def export_bank_report_pdf(payroll_run, calculations):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'BankTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#1E293B'),
        alignment=1
    )
    sub_title = ParagraphStyle(
        'BankSub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#64748B'),
        alignment=1
    )
    cell_h = ParagraphStyle('BCH', fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.white, alignment=1)
    cell_n = ParagraphStyle('BCN', fontName='Helvetica', fontSize=8, leading=10, textColor=colors.HexColor('#1E293B'))
    cell_r = ParagraphStyle('BCR', fontName='Helvetica', fontSize=8, leading=10, textColor=colors.HexColor('#1E293B'), alignment=2)
    cell_rb = ParagraphStyle('BCRB', fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.HexColor('#0F172A'), alignment=2)

    story.append(Paragraph("SIGNTECH COMMUNICATIONS - BANK SALARY TRANSFER SHEET", title_style))
    story.append(Paragraph(f"Period: {payroll_run.period_start.strftime('%B %Y')} | Status: {payroll_run.get_status_display()}", sub_title))
    story.append(Spacer(1, 15))

    headers = [
        Paragraph("<b>Sl</b>", cell_h),
        Paragraph("<b>ID</b>", cell_h),
        Paragraph("<b>Employee Name</b>", cell_h),
        Paragraph("<b>Bank Name</b>", cell_h),
        Paragraph("<b>Account Number</b>", cell_h),
        Paragraph("<b>Amount (BDT)</b>", cell_h),
    ]
    table_data = [headers]
    tot_bank = Decimal('0.00')

    for idx, calc in enumerate(calculations, 1):
        emp = calc.employee
        tot_bank += calc.bank_payable
        table_data.append([
            Paragraph(str(idx), cell_n),
            Paragraph(emp.employee_number, cell_n),
            Paragraph(emp.get_full_name(), cell_n),
            Paragraph(emp.bank_name or 'City Bank Ltd.', cell_n),
            Paragraph(emp.bank_account or 'N/A', cell_n),
            Paragraph(f"{calc.bank_payable:,.2f}", cell_r),
        ])

    table_data.append([
        Paragraph("<b>TOTAL</b>", cell_n),
        Paragraph("", cell_n),
        Paragraph("", cell_n),
        Paragraph("", cell_n),
        Paragraph("", cell_n),
        Paragraph(f"<b>{tot_bank:,.2f}</b>", cell_rb),
    ])

    col_widths = [30, 65, 145, 110, 95, 75]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#94A3B8')),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(table)
    story.append(Spacer(1, 40))

    # Authorizations
    sig_data = [
        [
            Paragraph("____________________________<br/><b>Prepared By</b>", ParagraphStyle('BS1', fontName='Helvetica', fontSize=8, alignment=0)),
            Paragraph("____________________________<br/><b>Manager (Accounts)</b>", ParagraphStyle('BS2', fontName='Helvetica', fontSize=8, alignment=1)),
            Paragraph("____________________________<br/><b>Managing Director</b>", ParagraphStyle('BS3', fontName='Helvetica', fontSize=8, alignment=2)),
        ]
    ]
    sig_table = Table(sig_data, colWidths=[170, 180, 170])
    story.append(sig_table)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def export_cash_report_excel(payroll_run, calculations):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Disbursal"

    title_font = Font(name='Calibri', size=14, bold=True, color='1E293B')
    subtitle_font = Font(name='Calibri', size=10, italic=True, color='64748B')
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    total_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    total_font = Font(name='Calibri', size=10, bold=True, color='0F172A')
    regular_font = Font(name='Calibri', size=9)

    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:F1')
    ws['A1'] = f"CASH / CHEQUE SALARY DISBURSEMENT SHEET - {payroll_run.period_start.strftime('%B %Y').upper()}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = f"Period: {payroll_run.period_start} to {payroll_run.period_end} | Status: {payroll_run.get_status_display()}"
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ['Sl', 'Employee ID', 'Employee Name', 'Department', 'Cash Amount (BDT)', 'Employee Signature']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row_num = 5
    tot_cash = Decimal('0.00')

    for idx, calc in enumerate(calculations, 1):
        emp = calc.employee
        dept_name = emp.department.name if getattr(emp, 'department', None) else 'General'

        ws.cell(row=row_num, column=1, value=idx).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=2, value=emp.employee_number).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=3, value=emp.get_full_name())
        ws.cell(row=row_num, column=4, value=dept_name)
        cell_amt = ws.cell(row=row_num, column=5, value=float(calc.cash_payable))
        cell_amt.number_format = '#,##0.00'
        ws.cell(row=row_num, column=6, value="")

        for c_idx in range(1, 7):
            cell = ws.cell(row=row_num, column=c_idx)
            cell.font = regular_font
            cell.border = border

        tot_cash += calc.cash_payable
        row_num += 1

    ws.cell(row=row_num, column=1, value="TOTAL")
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
    cell_tot = ws.cell(row=row_num, column=5, value=float(tot_cash))
    cell_tot.number_format = '#,##0.00'

    for c_idx in range(1, 7):
        cell = ws.cell(row=row_num, column=c_idx)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_cash_report_csv(payroll_run, calculations):
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    writer.writerow([f"CASH / CHEQUE SALARY DISBURSEMENT SHEET - {payroll_run.period_start.strftime('%B %Y')}"])
    writer.writerow([])
    writer.writerow(['Sl', 'Employee ID', 'Employee Name', 'Department', 'Amount (BDT)', 'Signature'])

    for idx, calc in enumerate(calculations, 1):
        emp = calc.employee
        dept_name = emp.department.name if getattr(emp, 'department', None) else 'General'
        writer.writerow([
            idx,
            emp.employee_number,
            emp.get_full_name(),
            dept_name,
            f"{calc.cash_payable:.2f}",
            ""
        ])

    return buffer.getvalue().encode('utf-8')


def export_cash_report_pdf(payroll_run, calculations):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CashTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#1E293B'),
        alignment=1
    )
    sub_title = ParagraphStyle(
        'CashSub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#64748B'),
        alignment=1
    )
    cell_h = ParagraphStyle('CCH', fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.white, alignment=1)
    cell_n = ParagraphStyle('CCN', fontName='Helvetica', fontSize=8, leading=10, textColor=colors.HexColor('#1E293B'))
    cell_r = ParagraphStyle('CCR', fontName='Helvetica', fontSize=8, leading=10, textColor=colors.HexColor('#1E293B'), alignment=2)
    cell_rb = ParagraphStyle('CCRB', fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.HexColor('#0F172A'), alignment=2)

    story.append(Paragraph("SIGNTECH COMMUNICATIONS - CASH / CHEQUE DISBURSEMENT SHEET", title_style))
    story.append(Paragraph(f"Period: {payroll_run.period_start.strftime('%B %Y')} | Status: {payroll_run.get_status_display()}", sub_title))
    story.append(Spacer(1, 15))

    headers = [
        Paragraph("<b>Sl</b>", cell_h),
        Paragraph("<b>ID</b>", cell_h),
        Paragraph("<b>Employee Name</b>", cell_h),
        Paragraph("<b>Department</b>", cell_h),
        Paragraph("<b>Amount (BDT)</b>", cell_h),
        Paragraph("<b>Receiver Signature</b>", cell_h),
    ]
    table_data = [headers]
    tot_cash = Decimal('0.00')

    for idx, calc in enumerate(calculations, 1):
        emp = calc.employee
        dept_name = emp.department.name if getattr(emp, 'department', None) else 'General'
        tot_cash += calc.cash_payable
        table_data.append([
            Paragraph(str(idx), cell_n),
            Paragraph(emp.employee_number, cell_n),
            Paragraph(emp.get_full_name(), cell_n),
            Paragraph(dept_name, cell_n),
            Paragraph(f"{calc.cash_payable:,.2f}", cell_r),
            Paragraph("________________", cell_n),
        ])

    table_data.append([
        Paragraph("<b>TOTAL</b>", cell_n),
        Paragraph("", cell_n),
        Paragraph("", cell_n),
        Paragraph("", cell_n),
        Paragraph(f"<b>{tot_cash:,.2f}</b>", cell_rb),
        Paragraph("", cell_n),
    ])

    col_widths = [30, 65, 150, 100, 80, 95]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#94A3B8')),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(table)
    story.append(Spacer(1, 40))

    sig_data = [
        [
            Paragraph("____________________________<br/><b>Disbursed By (Cashier)</b>", ParagraphStyle('CS1', fontName='Helvetica', fontSize=8, alignment=0)),
            Paragraph("____________________________<br/><b>Verified By (Accounts)</b>", ParagraphStyle('CS2', fontName='Helvetica', fontSize=8, alignment=1)),
            Paragraph("____________________________<br/><b>Approved By (HR/Admin)</b>", ParagraphStyle('CS3', fontName='Helvetica', fontSize=8, alignment=2)),
        ]
    ]
    sig_table = Table(sig_data, colWidths=[170, 180, 170])
    story.append(sig_table)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()
