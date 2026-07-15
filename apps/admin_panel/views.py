import csv
import calendar as cal_mod
from datetime import datetime, date as _date, timedelta
from collections import defaultdict
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from dateutil.relativedelta import relativedelta
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from apps.accounts.mixins import AdminRequiredMixin
from django.views.generic import TemplateView, ListView, View, FormView, DetailView, CreateView
from apps.attendance.models import Attendance
from apps.attendance.schedule_utils import (
    calculate_attendance_status,
    calculate_early_checkout,
    calculate_overtime,
)
from apps.employees.models import EmployeeProfile, EmployeeLocationSync
from apps.branches.models import Branch, OfficeSchedule
from apps.leave.models import LeaveType, LeaveBalance, LeaveRequest
from .forms import ManualAttendanceForm

def admin_required(view_func):
    from functools import wraps
    from django.http import HttpResponseForbidden
    from django.contrib.auth.decorators import login_required

    @wraps(view_func)
    @login_required
    def _wrapped(request, *args, **kwargs):
        if request.user.role != 'admin':
            return HttpResponseForbidden('Admins only.')
        return view_func(request, *args, **kwargs)
    return _wrapped

class AdminDashboardView(AdminRequiredMixin, TemplateView):
    template_name = 'admin_panel/admin_dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        date_str = self.request.GET.get('date')
        if date_str:
            try:
                today = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                today = timezone.localdate()
        else:
            today = timezone.localdate()

        branch_id = self.request.GET.get('branch')
        employee_id = self.request.GET.get('employee')
        
        todays_attendances = Attendance.objects.filter(date=today, is_expired=False).select_related('employee', 'employee__branch')
        
        if branch_id:
            todays_attendances = todays_attendances.filter(employee__branch_id=branch_id)
        if employee_id:
            todays_attendances = todays_attendances.filter(employee_id=employee_id)
        
        # For present count, count distinct employees who checked in today
        present_count = todays_attendances.filter(
            attendance_type='check_in'
        ).values('employee_id').distinct().count()
        late_count = todays_attendances.filter(attendance_type='check_in', status='late').values('employee_id').distinct().count()
        field_count = todays_attendances.filter(type='field').count()

        # Compute absent employees for today using get_absent_records
        absent_records = get_absent_records(
            date_from=today,
            date_to=today,
            employee_id=employee_id if employee_id else None,
            branch_id=branch_id if branch_id else None
        )
        on_leave_today = sum(1 for r in absent_records if r.status == 'on_leave')
        absent_count = sum(1 for r in absent_records if r.status == 'absent')
        not_checked_in = absent_records

        context.update({
            'branches': Branch.objects.all().order_by('name'),
            'employees': EmployeeProfile.objects.filter(is_active=True).order_by('full_name'),
            'selected_date': today.strftime('%Y-%m-%d'),
            'selected_branch': branch_id or '',
            'selected_employee': employee_id or '',
            'present_today': present_count,
            'absent_today':  absent_count,
            'on_leave_today': on_leave_today,
            'late_today':    late_count,
            'on_field':      field_count,
            'todays_attendances': todays_attendances.filter(attendance_type='check_in').order_by('-check_in_time'),
            'not_checked_in': not_checked_in,
            'today_str': today.strftime('%Y-%m-%d')
        })
        return context

class DashboardPartialView(AdminDashboardView):
    template_name = 'admin_panel/partials/dashboard_live.html'

class AdminAttendanceListView(AdminRequiredMixin, ListView):
    model = Attendance
    template_name = 'admin_panel/admin_attendance.html'
    context_object_name = 'attendances'
    paginate_by = 20
    
    def get_queryset(self):
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        emp_id = self.request.GET.get('employee')
        branch_id = self.request.GET.get('branch')
        att_type = self.request.GET.get('type')
        status = self.request.GET.get('status')
        
        if status == 'absent':
            return [r for r in get_absent_records(
                date_from=date_from,
                date_to=date_to,
                employee_id=emp_id,
                branch_id=branch_id
            ) if r.status == 'absent']
        elif status == 'on_leave':
            return [r for r in get_absent_records(
                date_from=date_from,
                date_to=date_to,
                employee_id=emp_id,
                branch_id=branch_id
            ) if r.status == 'on_leave']
            
        queryset = super().get_queryset().filter(is_expired=False).select_related('employee', 'employee__branch')
        
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        if emp_id:
            queryset = queryset.filter(employee_id=emp_id)
        if branch_id:
            queryset = queryset.filter(employee__branch_id=branch_id)
        if att_type:
            queryset = queryset.filter(type=att_type)
        if status:
            if status == 'present':
                queryset = queryset.filter(status__in=['on_time', 'late'])
            else:
                queryset = queryset.filter(status=status)
            
        return queryset.order_by('-date', '-check_in_time')
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        emp_id = self.request.GET.get('employee')
        branch_id = self.request.GET.get('branch')
        att_type = self.request.GET.get('type')
        status = self.request.GET.get('status')
        
        # Build base queryset for counting (excluding status filter)
        base_qs = Attendance.objects.filter(is_expired=False).select_related('employee', 'employee__branch')
        if date_from:
            base_qs = base_qs.filter(date__gte=date_from)
        if date_to:
            base_qs = base_qs.filter(date__lte=date_to)
        if emp_id:
            base_qs = base_qs.filter(employee_id=emp_id)
        if branch_id:
            base_qs = base_qs.filter(employee__branch_id=branch_id)
        if att_type:
            base_qs = base_qs.filter(type=att_type)
 
        # Count Present (distinct employees checked in today)
        total_present = base_qs.filter(
            attendance_type='check_in'
        ).values('employee_id', 'date').distinct().count()
 
        # Count Late (distinct employees late checked in today)
        total_late = base_qs.filter(
            attendance_type='check_in', status='late'
        ).values('employee_id', 'date').distinct().count()
 
        # Count Field (total field visits and check_ins of type field)
        total_field = base_qs.filter(type='field').count()
 
        # Count Absent / On Leave (using the helper)
        absent_records = get_absent_records(
            date_from=date_from,
            date_to=date_to,
            employee_id=emp_id,
            branch_id=branch_id
        )
        total_absent = sum(1 for r in absent_records if r.status == 'absent')
        total_on_leave = sum(1 for r in absent_records if r.status == 'on_leave')
 
        # Determine total_records based on active status filter
        if status == 'absent':
            total_records = total_absent
        elif status == 'on_leave':
            total_records = total_on_leave
        elif status == 'present':
            total_records = base_qs.filter(attendance_type='check_in').count()
        elif status == 'late':
            total_records = base_qs.filter(attendance_type='check_in', status='late').count()
        else:
            total_records = base_qs.count()
            
        context['total_records'] = total_records
        context['total_present'] = total_present
        context['total_absent'] = total_absent
        context['total_on_leave'] = total_on_leave
        context['total_late'] = total_late
        context['total_field'] = total_field
        
        context['employees'] = EmployeeProfile.objects.all()
        context['branches'] = Branch.objects.all()
        
        get_copy = self.request.GET.copy()
        if 'page' in get_copy:
            del get_copy['page']
        context['query_string'] = get_copy.urlencode()
        
        # Attach Leave Balance (remaining days, per leave type) as visible info
        year = timezone.localdate().year
        if date_from:
            try:
                year = datetime.strptime(date_from, '%Y-%m-%d').year
            except ValueError:
                pass
                
        page_attendances = context['attendances']
        employee_ids = {att.employee.id for att in page_attendances}
        
        leave_types = list(LeaveType.objects.all())
        balances_qs = LeaveBalance.objects.filter(employee_id__in=employee_ids, year=year)
        
        balances_by_emp = defaultdict(list)
        for bal in balances_qs:
            balances_by_emp[bal.employee_id].append({
                'type': bal.leave_type,
                'remaining': bal.remaining_days,
                'total': bal.total_days
            })
            
        for emp_id in employee_ids:
            emp_bals = balances_by_emp[emp_id]
            existing_types = {b['type'].id for b in emp_bals}
            for lt in leave_types:
                if lt.id not in existing_types:
                    emp_bals.append({
                        'type': lt,
                        'remaining': lt.default_days_per_year,
                        'total': lt.default_days_per_year
                    })
                    
        for att in page_attendances:
            att.leave_balances = balances_by_emp[att.employee.id]
            
        return context

class ExportAttendanceCSVView(AdminRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        emp_id = request.GET.get('employee')
        branch_id = request.GET.get('branch')
        att_type = request.GET.get('type')
        status = request.GET.get('status')
        
        if status == 'absent':
            records = get_absent_records(
                date_from=date_from,
                date_to=date_to,
                employee_id=emp_id,
                branch_id=branch_id
            )
        else:
            queryset = Attendance.objects.select_related(
                'employee', 'employee__branch'
            ).filter(is_expired=False)
            
            if date_from:
                queryset = queryset.filter(date__gte=date_from)
            if date_to:
                queryset = queryset.filter(date__lte=date_to)
            if emp_id:
                queryset = queryset.filter(employee_id=emp_id)
            if branch_id:
                queryset = queryset.filter(employee__branch_id=branch_id)
            if att_type:
                queryset = queryset.filter(type=att_type)
            if status:
                if status == 'present':
                    queryset = queryset.filter(status__in=['on_time', 'late'])
                else:
                    queryset = queryset.filter(status=status)
                
            records = queryset.order_by('-date', '-check_in_time')
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="attendance_export_{timezone.localdate()}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['SN', 'Date', 'Employee ID', 'Employee Name', 'Branch', 'Check In', 'Check Out', 'Hours', 'Type', 'Status', 'Note'])
        
        for idx, att in enumerate(records, 1):
            check_in = timezone.localtime(att.check_in_time).strftime('%H:%M:%S') if att.check_in_time else ''
            check_out = timezone.localtime(att.check_out_time).strftime('%H:%M:%S') if att.check_out_time else ''
            branch_name = att.employee.branch.name if att.employee.branch else 'Unassigned'
            writer.writerow([
                idx,
                att.date,
                att.employee.employee_id,
                att.employee.full_name,
                branch_name,
                check_in,
                check_out,
                att.total_hours or '',
                att.get_type_display(),
                att.get_status_display(),
                att.note
            ])
            
        return response

class ManualEntryView(AdminRequiredMixin, FormView):
    template_name = 'admin_panel/manual_entry.html'
    form_class = ManualAttendanceForm
    success_url = '/admin-panel/attendance/'
    
    def form_valid(self, form):
        att = form.save(commit=False)
        if att.check_in_time and att.check_out_time:
            if att.check_out_time > att.check_in_time:
                duration = att.check_out_time - att.check_in_time
                att.total_hours = round(duration.total_seconds() / 3600.0, 2)
                
        reason = form.cleaned_data.get('admin_override_reason')
        if reason:
            att.note = f"Admin Override: {reason}"
            
        att.save()
        messages.success(self.request, "Manual attendance record added successfully.")
        return super().form_valid(form)

class AttendanceDetailView(AdminRequiredMixin, DetailView):
    model = Attendance
    template_name = 'admin_panel/attendance_detail.html'
    context_object_name = 'attendance'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        attendance = self.object
        employee = attendance.employee
        locations = attendance.locations.all().order_by('timestamp')
        
        # Date range for today in BST timezone
        tz = timezone.get_current_timezone()
        today = attendance.date
        
        # Start and end datetimes for today
        start_dt = timezone.make_aware(datetime.combine(today, datetime.min.time()), tz)
        end_dt = timezone.make_aware(datetime.combine(today, datetime.max.time()), tz)
        
        time_from = self.request.GET.get('time_from')
        time_to = self.request.GET.get('time_to')
        
        if time_from:
            try:
                tf = datetime.strptime(time_from, '%H:%M').time()
                start_dt = timezone.make_aware(datetime.combine(today, tf), tz)
            except ValueError:
                pass
        if time_to:
            try:
                tt = datetime.strptime(time_to, '%H:%M').time()
                end_dt = timezone.make_aware(datetime.combine(today, tt), tz)
            except ValueError:
                pass
                
        syncs = EmployeeLocationSync.objects.filter(
            employee=employee,
            timestamp__range=(start_dt, end_dt)
        ).order_by('timestamp')
        
        context['location_syncs'] = syncs
        context['locations'] = locations
        context['first_location'] = locations.first()
        context['last_location'] = locations.last()
        context['time_from'] = time_from or ''
        context['time_to'] = time_to or ''
        return context


class AttendanceLocationsView(AdminRequiredMixin, View):
    def get(self, request, pk):
        attendance = get_object_or_404(Attendance, pk=pk)
        locations = attendance.locations.all().order_by('timestamp')
        return render(request, 'admin_panel/partials/location_timeline.html', {
            'locations': locations,
        })


class SyntheticAttendance:
    def __init__(self, employee, date, status='absent', note='Absent'):
        self.id = None
        self.employee = employee
        self.date = date
        self.check_in_time = None
        self.check_out_time = None
        self.total_hours = None
        self.photo = None
        self.attendance_type = ''  # empty to prevent showing check-in/out badges
        self.type = 'office'
        self.status = status
        self.is_early_checkout = False
        self.note = note

    def get_type_display(self):
        return 'Office'

    def get_status_display(self):
        if self.status == 'on_leave':
            return 'On Leave'
        return 'Absent'


def get_absent_records(date_from=None, date_to=None, employee_id=None, branch_id=None):
    """
    Computes absent employees for a date or date range.
    Returns a list of SyntheticAttendance objects representing dates on which active employees were absent.
    """
    today = timezone.localdate()
    
    if date_from:
        if isinstance(date_from, str):
            try:
                start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            except ValueError:
                start_date = today
        else:
            start_date = date_from
    else:
        start_date = today

    if date_to:
        if isinstance(date_to, str):
            try:
                end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            except ValueError:
                end_date = start_date
        else:
            end_date = date_to
    else:
        end_date = start_date

    # Ensure start_date <= end_date
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    # Get active employees
    employees = EmployeeProfile.objects.filter(is_active=True)
    if employee_id:
        employees = employees.filter(id=employee_id)
    if branch_id:
        employees = employees.filter(branch_id=branch_id)
    employees = employees.select_related('branch')

    # Get all check-in and field-visit attendance records in the date range
    attendances = Attendance.objects.filter(
        date__range=(start_date, end_date),
        is_expired=False
    )
    if employee_id:
        attendances = attendances.filter(employee_id=employee_id)
    if branch_id:
        attendances = attendances.filter(employee__branch_id=branch_id)

    # Group attendance by employee and date
    att_by_emp_date = defaultdict(list)
    for att in attendances:
        att_by_emp_date[(att.employee_id, att.date)].append(att)

    # Fetch approved leave requests in date range
    leave_requests = LeaveRequest.objects.filter(
        status='approved',
        start_date__lte=end_date,
        end_date__gte=start_date
    )
    if employee_id:
        leave_requests = leave_requests.filter(employee_id=employee_id)
    if branch_id:
        leave_requests = leave_requests.filter(employee__branch_id=branch_id)
    leave_requests = leave_requests.select_related('leave_type')

    # Map of (employee_id, date) -> LeaveRequest
    approved_leaves = {}
    for req in leave_requests:
        s_date = max(req.start_date, start_date)
        e_date = min(req.end_date, end_date)
        curr = s_date
        while curr <= e_date:
            approved_leaves[(req.employee_id, curr)] = req
            curr += timedelta(days=1)

    # Loop through each date and find absent employees
    synthetic_records = []
    curr_date = start_date
    schedule_cache = {}
    while curr_date <= end_date:
        for emp in employees:
            if emp.branch_id not in schedule_cache:
                schedule_cache[emp.branch_id] = _get_employee_schedule(emp)
            schedule = schedule_cache[emp.branch_id]

            # Skip non-working days for this employee's branch schedule
            if not _is_working_day(curr_date, schedule):
                continue

            # Don't mark future dates as absent
            if curr_date > today:
                continue

            day_atts = att_by_emp_date[(emp.id, curr_date)]
            has_check_in = any(a.attendance_type == 'check_in' for a in day_atts)
            has_field_visit = any(a.attendance_type == 'field_visit' for a in day_atts)
            
            # Check if there is no check_in and no field_visit
            if not has_check_in and not has_field_visit:
                req = approved_leaves.get((emp.id, curr_date))
                if req:
                    record = SyntheticAttendance(emp, curr_date, status='on_leave', note=f"On Leave - {req.leave_type.name}")
                else:
                    record = SyntheticAttendance(emp, curr_date, status='absent', note="Absent")
                synthetic_records.append(record)
        curr_date += timedelta(days=1)

    # Sort: employee name ascending, date descending
    synthetic_records.sort(key=lambda x: x.employee.full_name)
    synthetic_records.sort(key=lambda x: x.date, reverse=True)
    return synthetic_records


def get_unified_deductions(date_from=None, date_to=None, employee_id=None, branch_id=None, leave_type_id=None):
    from apps.attendance.models import AttendanceAbsentLog
    from apps.leave.models import LeaveRequest
    from apps.employees.models import EmployeeProfile
    from datetime import datetime, timedelta

    # 1. Fetch relevant active employees
    employees = EmployeeProfile.objects.filter(is_active=True)
    if employee_id:
        employees = employees.filter(id=employee_id)
    if branch_id:
        employees = employees.filter(branch_id=branch_id)
    employees = {emp.id: emp for emp in employees.select_related('branch')}
    
    valid_employee_ids = list(employees.keys())
    if not valid_employee_ids:
        return []

    # 2. Query AttendanceAbsentLog
    absent_logs = AttendanceAbsentLog.objects.filter(employee_id__in=valid_employee_ids).select_related('leave_type_deducted')
    
    if date_from and date_from != "":
        d_from = datetime.strptime(date_from, '%Y-%m-%d').date() if isinstance(date_from, str) else date_from
        absent_logs = absent_logs.filter(date__gte=d_from)
    else:
        d_from = None

    if date_to and date_to != "":
        d_to = datetime.strptime(date_to, '%Y-%m-%d').date() if isinstance(date_to, str) else date_to
        absent_logs = absent_logs.filter(date__lte=d_to)
    else:
        d_to = None

    if leave_type_id:
        absent_logs = absent_logs.filter(leave_type_deducted_id=leave_type_id)

    deductions = []
    
    # Add absent logs to list
    for log in absent_logs:
        emp = employees.get(log.employee_id)
        if emp:
            deductions.append({
                'employee': emp,
                'date': log.date,
                'leave_type_deducted': log.leave_type_deducted,
                'branch': emp.branch,
                'source_type': 'absent_log',
                'record_id': log.id,
                'leave_request': None,
            })

    # 3. Query approved LeaveRequest
    leave_requests = LeaveRequest.objects.filter(
        employee_id__in=valid_employee_ids,
        status='approved'
    ).select_related('leave_type')

    if d_from:
        leave_requests = leave_requests.filter(end_date__gte=d_from)
    if d_to:
        leave_requests = leave_requests.filter(start_date__lte=d_to)
    if leave_type_id:
        leave_requests = leave_requests.filter(leave_type_id=leave_type_id)

    schedule_cache = {}
    for req in leave_requests:
        emp = employees.get(req.employee_id)
        if not emp:
            continue

        if emp.branch_id not in schedule_cache:
            schedule_cache[emp.branch_id] = _get_employee_schedule(emp)
        schedule = schedule_cache[emp.branch_id]

        s_date = req.start_date
        if d_from:
            s_date = max(s_date, d_from)

        e_date = req.end_date
        if d_to:
            e_date = min(e_date, d_to)

        curr = s_date
        while curr <= e_date:
            if _is_working_day(curr, schedule):
                deductions.append({
                    'employee': emp,
                    'date': curr,
                    'leave_type_deducted': req.leave_type,
                    'branch': emp.branch,
                    'source_type': 'leave_request',
                    'record_id': req.id,
                    'leave_request': req,
                })
            curr += timedelta(days=1)

    # Sort unified list: date descending, employee name ascending
    deductions.sort(key=lambda x: x['employee'].full_name)
    deductions.sort(key=lambda x: x['date'], reverse=True)
    return deductions


def _get_employee_schedule(employee):
    branch = getattr(employee, 'branch', None)
    if not branch:
        return None
    try:
        return branch.schedule
    except OfficeSchedule.DoesNotExist:
        return None


def _get_working_day_set(schedule):
    if schedule is not None:
        return {day.lower() for day in (schedule.working_days or [])}
    
    # Try to load the schedule from the schedule settings page (first office schedule in DB)
    from apps.branches.models import OfficeSchedule
    try:
        first_schedule = OfficeSchedule.objects.first()
        if first_schedule and first_schedule.working_days:
            return {day.lower() for day in first_schedule.working_days}
    except Exception:
        pass
        
    # Default fallback if database query fails or no schedules exist
    return {'saturday', 'sunday', 'monday', 'tuesday', 'wednesday', 'thursday'}


def _is_working_day(day_value, schedule):
    return cal_mod.day_name[day_value.weekday()].lower() in _get_working_day_set(schedule)


def _get_working_days(year, month, schedule=None):
    cal_mod.setfirstweekday(cal_mod.SATURDAY)
    working_day_set = _get_working_day_set(schedule)
    count = 0
    first_wd = cal_mod.firstweekday()
    for week in cal_mod.monthcalendar(year, month):
        for idx, day in enumerate(week):
            if day:
                actual_weekday = (first_wd + idx) % 7
                if cal_mod.day_name[actual_weekday].lower() in working_day_set:
                    count += 1
    return count


def _build_calendar_weeks(year, month, att_by_date, schedule, leave_dates=None):
    """
    Build calendar week rows for a month.
    att_by_date: {date: [Attendance, ...]}
    Returns list of 7-item lists; None = empty padding cell.
    Colors: green=on_time, amber=late, red=absent, purple=field_visit_only, gray=non_working_day
    """
    cal_mod.setfirstweekday(cal_mod.SATURDAY)
    working_day_set = _get_working_day_set(schedule)
    weeks = []
    today = timezone.localdate()
    for week in cal_mod.monthcalendar(year, month):
        row = []
        for idx, day in enumerate(week):
            if day == 0:
                row.append(None)
                continue
            d = _date(year, month, day)
            day_atts = att_by_date.get(d, [])
            ci = next((a for a in day_atts if a.attendance_type == 'check_in'), None)
            has_fv = any(a.attendance_type == 'field_visit' for a in day_atts)
            is_weekend = cal_mod.day_name[d.weekday()].lower() not in working_day_set
            is_late = bool(ci and calculate_attendance_status(ci.check_in_time, schedule) == 'late')
            if is_weekend:
                color = 'gray'
            elif ci:
                color = 'amber' if is_late else 'green'
            elif has_fv:
                color = 'purple'
            elif leave_dates and d in leave_dates:
                color = 'blue'
            elif d > today:
                color = 'future'
            else:
                color = 'red'
            row.append({
                'day': day,
                'date': d,
                'date_str': d.strftime('%Y-%m-%d'),
                'check_in': ci,
                'has_fv': has_fv,
                'is_weekend': is_weekend,
                'is_late': is_late,
                'color': color,
            })
        weeks.append(row)
    return weeks


def _filter_qs_by_request(qs, request):
    """Apply shared GET filters (date/date_from/date_to/employee/branch)."""
    qs = qs.filter(is_expired=False)
    date_single = request.GET.get('date', '')
    date_from   = request.GET.get('date_from', '')
    date_to     = request.GET.get('date_to', '')
    emp         = request.GET.get('employee', '')
    branch      = request.GET.get('branch', '')

    if date_single:
        qs = qs.filter(date=date_single)
    else:
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
    if emp:
        qs = qs.filter(employee_id=emp)
    if branch:
        qs = qs.filter(employee__branch_id=branch)
    return qs


# ─────────────────────────────────────────────────────────────────
# REPORT VIEWS
# ─────────────────────────────────────────────────────────────────

class ReportsMainView(AdminRequiredMixin, TemplateView):
    template_name = 'admin_panel/reports/main.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['employees'] = EmployeeProfile.objects.filter(is_active=True).order_by('full_name')
        return ctx


class DailyReportView(AdminRequiredMixin, View):
    template_name = 'admin_panel/reports/daily.html'

    def get(self, request):
        date_str = request.GET.get('date', '')
        try:
            report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            report_date = timezone.localdate()

        emp_id    = request.GET.get('employee', '')
        branch_id = request.GET.get('branch', '')

        employees = (
            EmployeeProfile.objects.filter(is_active=True)
            .select_related('branch').order_by('full_name')
        )
        if emp_id:
            employees = employees.filter(id=emp_id)
        if branch_id:
            employees = employees.filter(branch_id=branch_id)

        attendances = (
            Attendance.objects
            .filter(date=report_date, is_expired=False)
            .select_related('employee', 'employee__branch')
            .prefetch_related('locations')
        )
        if emp_id:
            attendances = attendances.filter(employee_id=emp_id)
        if branch_id:
            attendances = attendances.filter(employee__branch_id=branch_id)

        att_map = defaultdict(list)
        for a in attendances:
            att_map[a.employee_id].append(a)

        rows = []
        for emp in employees:
            emp_atts = att_map.get(emp.id, [])
            check_in = next((a for a in emp_atts if a.attendance_type == 'check_in'), None)
            fv_list  = [a for a in emp_atts if a.attendance_type == 'field_visit']
            loc = check_in.locations.filter(event='check_in').first() if check_in else None
            
            if check_in:
                status = check_in.status
            elif fv_list:
                status = 'on_time'  # They are working, just off-site
            else:
                status = 'absent'
                
            rows.append({
                'employee':     emp,
                'check_in':     check_in,
                'field_visits': fv_list,
                'location':     loc,
                'status':       status,
            })

        present     = sum(1 for r in rows if r['check_in'])
        late        = sum(1 for r in rows if r['status'] == 'late')
        absent      = sum(1 for r in rows if r['status'] == 'absent')
        field_total = sum(len(r['field_visits']) for r in rows)
        total_hours = round(sum(
            float(r['check_in'].total_hours or 0)
            for r in rows if r['check_in'] and r['check_in'].total_hours
        ), 2)

        return render(request, self.template_name, {
            'report_date':       report_date,
            'rows':              rows,
            'present':           present,
            'absent':            absent,
            'late':              late,
            'field_total':       field_total,
            'total_hours':       total_hours,
            'employees':         EmployeeProfile.objects.filter(is_active=True).order_by('full_name'),
            'branches':          Branch.objects.all(),
            'selected_employee': emp_id,
            'selected_branch':   branch_id,
        })


class MonthlyReportView(AdminRequiredMixin, View):
    template_name = 'admin_panel/reports/monthly.html'

    def get(self, request):
        today = timezone.localdate()
        try:
            year  = int(request.GET.get('year',  today.year))
            month = int(request.GET.get('month', today.month))
        except (ValueError, TypeError):
            year, month = today.year, today.month
        month = max(1, min(12, month))

        emp_id    = request.GET.get('employee', '')
        branch_id = request.GET.get('branch', '')

        summary_schedule = None
        if emp_id:
            try:
                summary_schedule = _get_employee_schedule(
                    EmployeeProfile.objects.select_related('branch').get(id=emp_id)
                )
            except EmployeeProfile.DoesNotExist:
                summary_schedule = None
        elif branch_id:
            try:
                summary_schedule = Branch.objects.get(id=branch_id).schedule
            except (Branch.DoesNotExist, OfficeSchedule.DoesNotExist):
                summary_schedule = None

        working_days         = _get_working_days(year, month, summary_schedule)
        _, last_day          = cal_mod.monthrange(year, month)
        month_start          = _date(year, month, 1)
        month_end            = _date(year, month, last_day)

        employees = (
            EmployeeProfile.objects.filter(is_active=True)
            .select_related('branch').order_by('full_name')
        )
        if emp_id:
            employees = employees.filter(id=emp_id)
        if branch_id:
            employees = employees.filter(branch_id=branch_id)

        attendances = Attendance.objects.filter(
            date__gte=month_start, date__lte=month_end, is_expired=False
        ).select_related('employee')
        if emp_id:
            attendances = attendances.filter(employee_id=emp_id)
        if branch_id:
            attendances = attendances.filter(employee__branch_id=branch_id)

        emp_att_map = defaultdict(list)
        for a in attendances:
            emp_att_map[a.employee_id].append(a)

        # Fetch approved leave requests in the month
        leave_requests = LeaveRequest.objects.filter(
            status='approved',
            start_date__lte=month_end,
            end_date__gte=month_start
        )
        if emp_id:
            leave_requests = leave_requests.filter(employee_id=emp_id)
        if branch_id:
            leave_requests = leave_requests.filter(employee__branch_id=branch_id)
        leave_requests = list(leave_requests.select_related('leave_type'))

        # Fetch leave balances for matching employees in year
        leave_types = list(LeaveType.objects.all())
        employee_ids = [emp.id for emp in employees]
        balances_qs = LeaveBalance.objects.filter(employee_id__in=employee_ids, year=year)
        balances_by_emp = defaultdict(list)
        for bal in balances_qs:
            balances_by_emp[bal.employee_id].append({
                'type': bal.leave_type,
                'remaining': bal.remaining_days,
                'total': bal.total_days
            })
        for e_id in employee_ids:
            emp_bals = balances_by_emp[e_id]
            existing_types = {b['type'].id for b in emp_bals}
            for lt in leave_types:
                if lt.id not in existing_types:
                    emp_bals.append({
                        'type': lt,
                        'remaining': lt.default_days_per_year,
                        'total': lt.default_days_per_year
                    })

        rows = []
        total_present = total_absent = total_on_leave = total_late = total_field = 0
        for emp in employees:
            atts    = emp_att_map.get(emp.id, [])
            cis     = [a for a in atts if a.attendance_type == 'check_in']
            fvs     = [a for a in atts if a.attendance_type == 'field_visit']
            
            check_in_dates = set(a.date for a in cis)
            field_dates = set(a.date for a in fvs)
            
            present_days = len(check_in_dates)
            field_only_days = len(field_dates - check_in_dates)
            
            present = present_days
            late    = sum(1 for a in cis if a.status == 'late')
            
            # Calculate leave days on working days with no attendance
            schedule = _get_employee_schedule(emp)
            emp_working_days = _get_working_days(year, month, schedule)
            emp_leaves = [req for req in leave_requests if req.employee_id == emp.id]
            att_dates = set(a.date for a in atts if a.attendance_type in ['check_in', 'field_visit'])
            
            on_leave_count = 0
            curr_date = month_start
            while curr_date <= month_end:
                if _is_working_day(curr_date, schedule):
                    if curr_date not in att_dates:
                        is_on_leave = False
                        for req in emp_leaves:
                            if req.start_date <= curr_date <= req.end_date:
                                is_on_leave = True
                                break
                        if is_on_leave:
                            on_leave_count += 1
                curr_date += timedelta(days=1)
            
            absent  = max(0, emp_working_days - (present_days + field_only_days) - on_leave_count)
            hours   = round(sum(float(a.total_hours or 0) for a in cis), 2)
            att_pct = round((present / emp_working_days * 100) if emp_working_days else 0, 1)
            
            rows.append({
                'employee':     emp,
                'present':      present,
                'absent':       absent,
                'on_leave':     on_leave_count,
                'late':         late,
                'field_visits': len(fvs),
                'total_hours':  hours,
                'att_pct':      att_pct,
                'leave_balances': balances_by_emp[emp.id]
            })
            total_present += present
            total_absent  += absent
            total_on_leave += on_leave_count
            total_late    += late
            total_field   += len(fvs)

        avg_att_pct = round(
            sum(r['att_pct'] for r in rows) / len(rows) if rows else 0, 1
        )
        prev_y, prev_m = (year, month - 1) if month > 1 else (year - 1, 12)
        next_y, next_m = (year, month + 1) if month < 12 else (year + 1, 1)

        return render(request, self.template_name, {
            'year': year, 'month': month,
            'month_name':        cal_mod.month_name[month],
            'working_days':      working_days,
            'month_start':       month_start,
            'month_end':         month_end,
            'rows':              rows,
            'total_present':     total_present,
            'total_absent':      total_absent,
            'total_on_leave':    total_on_leave,
            'total_late':        total_late,
            'total_field':       total_field,
            'avg_att_pct':       avg_att_pct,
            'employees':         EmployeeProfile.objects.filter(is_active=True).order_by('full_name'),
            'branches':          Branch.objects.all(),
            'selected_employee': emp_id,
            'selected_branch':   branch_id,
            'prev_m': prev_m, 'prev_y': prev_y,
            'next_m': next_m, 'next_y': next_y,
        })


def format_hours_minutes(total_hours_val):
    if not total_hours_val:
        return "0 min"
    total_minutes = int(round(float(total_hours_val) * 60))
    if total_minutes == 0:
        return "0 min"
    hours = total_minutes // 60
    minutes = total_minutes % 60
    
    if hours == 0:
        return f"{minutes} min"
    elif minutes == 0:
        return f"{hours}h"
    else:
        return f"{hours}h {minutes}m"

def format_minutes(total_minutes):
    if not total_minutes or total_minutes == 0:
        return "0 min"
    hours = total_minutes // 60
    minutes = total_minutes % 60
    if hours == 0:
        return f"{minutes} min"
    elif minutes == 0:
        return f"{hours}h"
    else:
        return f"{hours}h {minutes}m"


class EmployeeReportView(AdminRequiredMixin, View):
    template_name = 'admin_panel/reports/employee_report.html'

    def get(self, request, pk, year=None, month=None):
        import datetime as dt_mod
        employee = get_object_or_404(EmployeeProfile, pk=pk)
        today = timezone.localdate()
        if not year or not month:
            try:
                year  = int(request.GET.get('year',  today.year))
                month = int(request.GET.get('month', today.month))
            except (ValueError, TypeError):
                year, month = today.year, today.month
        month = max(1, min(12, month))

        schedule = _get_employee_schedule(employee)
        _, last_day  = cal_mod.monthrange(year, month)
        month_start  = _date(year, month, 1)
        month_end    = _date(year, month, last_day)
        working_days = _get_working_days(year, month, schedule)
        working_day_set = _get_working_day_set(schedule)
        cal_mod.setfirstweekday(cal_mod.SATURDAY)
        first_wd = cal_mod.firstweekday()
        weekday_headers = [
            {
                'label': cal_mod.day_abbr[(first_wd + idx) % 7],
                'is_working_day': cal_mod.day_name[(first_wd + idx) % 7].lower() in working_day_set,
            }
            for idx in range(7)
        ]
        working_day_labels = [
            cal_mod.day_name[idx].title()
            for idx in range(7)
            if cal_mod.day_name[idx].lower() in working_day_set
        ]
        off_day_labels = [
            cal_mod.day_name[idx].title()
            for idx in range(7)
            if cal_mod.day_name[idx].lower() not in working_day_set
        ]

        attendances = list(
            Attendance.objects.filter(
                employee=employee, date__gte=month_start, date__lte=month_end, is_expired=False
            ).prefetch_related('locations').order_by('date', 'check_in_time')
        )

        att_by_date = defaultdict(list)
        for a in attendances:
            att_by_date[a.date].append(a)

        # Fetch approved leave requests in the month
        leave_requests = LeaveRequest.objects.filter(
            employee=employee,
            status='approved',
            start_date__lte=month_end,
            end_date__gte=month_start
        ).select_related('leave_type')

        leave_dates = set()
        for req in leave_requests:
            s_date = max(req.start_date, month_start)
            e_date = min(req.end_date, month_end)
            curr = s_date
            while curr <= e_date:
                leave_dates.add(curr)
                curr += dt_mod.timedelta(days=1)

        calendar_weeks = _build_calendar_weeks(year, month, att_by_date, schedule, leave_dates=leave_dates)

        cis     = [a for a in attendances if a.attendance_type == 'check_in']
        fvs     = [a for a in attendances if a.attendance_type == 'field_visit']
        
        check_in_dates = set(a.date for a in cis)
        field_dates = set(a.date for a in fvs)
        
        present_days = len(check_in_dates)
        field_only_days = len(field_dates - check_in_dates)
        
        present = present_days
        late    = sum(1 for a in cis if calculate_attendance_status(a.check_in_time, schedule) == 'late')
        early_checkouts = sum(
            1 for a in cis
            if a.check_out_time and calculate_early_checkout(a.check_out_time, schedule)
        )
        
        # Calculate stats considering current time limits (avoid marking future as absent)
        if year < today.year or (year == today.year and month < today.month):
            max_date = month_end
        elif year == today.year and month == today.month:
            max_date = today
        else:
            max_date = month_start - dt_mod.timedelta(days=1)

        working_days_so_far = 0
        on_leave_days = 0
        absent_days = 0

        current = month_start
        while current <= month_end:
            if _is_working_day(current, schedule):
                if current <= max_date:
                    working_days_so_far += 1
                    day_atts = att_by_date.get(current, [])
                    has_check_in = any(a.attendance_type == 'check_in' for a in day_atts)
                    has_field_visit = any(a.attendance_type == 'field_visit' for a in day_atts)
                    if not has_check_in and not has_field_visit:
                        if current in leave_dates:
                            on_leave_days += 1
                        else:
                            absent_days += 1
            current += dt_mod.timedelta(days=1)

        absent = absent_days
        total_hours_sum = sum(float(a.total_hours or 0) for a in cis)
        total_hours_str = format_hours_minutes(total_hours_sum)
        
        overtime_minutes = sum(
            calculate_overtime(a.check_out_time, schedule, employee)
            if a.check_out_time else 0
            for a in cis
        )
        overtime_hours_str = format_minutes(overtime_minutes)
        att_pct = round((present / working_days_so_far * 100) if working_days_so_far else 0, 1)

        # Build full-month table
        table_rows = []
        current = month_start
        while current <= month_end:
            day_atts = att_by_date.get(current, [])
            ci = next((a for a in day_atts if a.attendance_type == 'check_in'), None)
            fv = [a for a in day_atts if a.attendance_type == 'field_visit']
            is_weekend = not _is_working_day(current, schedule)
            is_late = bool(ci and calculate_attendance_status(ci.check_in_time, schedule) == 'late')
            is_early_checkout = bool(ci and ci.check_out_time and calculate_early_checkout(ci.check_out_time, schedule))
            overtime_for_day = 0
            overtime_str = ''
            total_hours_for_day_str = ''
            
            if ci:
                if ci.total_hours:
                    total_hours_for_day_str = format_hours_minutes(ci.total_hours)
                if ci.check_out_time:
                    overtime_for_day = calculate_overtime(ci.check_out_time, schedule, employee)
                    if overtime_for_day > 0:
                        overtime_str = format_minutes(overtime_for_day)
            
            if is_weekend and not ci and not fv:
                status_val = 'weekend'
                status_display = 'Weekend'
            elif fv and not ci:
                status_val = 'on_time'
                status_display = 'On Field'
            elif ci:
                status_val = 'late' if is_late else 'on_time'
                status_display = 'Late' if is_late else 'On Time'
            elif current in leave_dates:
                status_val = 'on_leave'
                status_display = 'On Leave'
            elif current > today:
                status_val = 'scheduled'
                status_display = 'Scheduled'
            else:
                status_val = 'absent'
                status_display = 'Absent'

            table_rows.append({
                'date':         current,
                'check_in':     ci,
                'field_visits': fv,
                'is_weekend':   is_weekend,
                'is_late':      is_late,
                'is_early_checkout': is_early_checkout,
                'status':       status_val,
                'status_display': status_display,
                'overtime_minutes': overtime_for_day,
                'overtime_str': overtime_str,
                'total_hours_str': total_hours_for_day_str,
            })
            current += dt_mod.timedelta(days=1)

        prev_y, prev_m = (year, month - 1) if month > 1 else (year - 1, 12)
        next_y, next_m = (year, month + 1) if month < 12 else (year + 1, 1)

        return render(request, self.template_name, {
            'employee':        employee,
            'schedule':        schedule,
            'year': year, 'month': month,
            'month_name':      cal_mod.month_name[month],
            'weekday_headers': weekday_headers,
            'working_day_labels': working_day_labels,
            'off_day_labels':  off_day_labels,
            'calendar_weeks':  calendar_weeks,
            'present':         present,
            'absent':          absent,
            'on_leave':        on_leave_days,
            'late':            late,
            'early_checkouts': early_checkouts,
            'field_visits':    len(fvs),
            'total_hours':     total_hours_str,
            'overtime_hours':  overtime_hours_str,
            'att_pct':         att_pct,
            'table_rows':      table_rows,
            'prev_m': prev_m, 'prev_y': prev_y,
            'next_m': next_m, 'next_y': next_y,
        })


class EmployeeDayDetailView(AdminRequiredMixin, View):
    def get(self, request, pk, date_str):
        employee = get_object_or_404(EmployeeProfile, pk=pk)
        try:
            day = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return HttpResponse('Invalid date', status=400)

        day_atts = list(
            Attendance.objects.filter(employee=employee, date=day, is_expired=False)
            .prefetch_related('locations').order_by('check_in_time')
        )
        ci      = next((a for a in day_atts if a.attendance_type == 'check_in'), None)
        fv_list = [a for a in day_atts if a.attendance_type == 'field_visit']
        ci_loc  = ci.locations.filter(event='check_in').first() if ci else None
        co_loc  = ci.locations.filter(event='check_out').first() if ci else None

        if ci and getattr(ci, 'overtime_minutes', 0) > 0:
            ci.overtime_str = f"{ci.overtime_minutes // 60}h {ci.overtime_minutes % 60}m"
        else:
            if ci: ci.overtime_str = ""

        return render(request, 'admin_panel/reports/partials/day_detail.html', {
            'employee':     employee,
            'day':          day,
            'check_in':     ci,
            'ci_loc':       ci_loc,
            'co_loc':       co_loc,
            'field_visits': fv_list,
        })


class ExportReportCSVView(AdminRequiredMixin, View):
    def get(self, request):
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        emp_id = request.GET.get('employee')
        branch_id = request.GET.get('branch')
        att_type = request.GET.get('type')
        status = request.GET.get('status')
        
        if status == 'absent':
            records = get_absent_records(
                date_from=date_from,
                date_to=date_to,
                employee_id=emp_id,
                branch_id=branch_id
            )
        else:
            qs = (
                Attendance.objects
                .select_related('employee', 'employee__branch')
                .prefetch_related('locations')
            )
            qs = _filter_qs_by_request(qs, request)
            if att_type:
                qs = qs.filter(type=att_type)
            if status:
                if status == 'present':
                    qs = qs.filter(status__in=['on_time', 'late'])
                else:
                    qs = qs.filter(status=status)
            records = list(qs.order_by('date', 'employee__full_name'))

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = (
            f'attachment; filename="attendance_report_{timezone.localdate()}.csv"'
        )
        writer = csv.writer(response)
        writer.writerow([
            'SN', 'Employee', 'Employee ID', 'Date',
            'Check-in', 'Check-out', 'Hours',
            'Type', 'Status', 'Location', 'Notes',
        ])
        for idx, a in enumerate(records, 1):
            loc = None
            if not isinstance(a, SyntheticAttendance):
                loc = a.locations.filter(event='check_in').first()
            writer.writerow([
                idx,
                a.employee.full_name,
                a.employee.employee_id,
                a.date,
                timezone.localtime(a.check_in_time).strftime('%H:%M') if a.check_in_time else '',
                timezone.localtime(a.check_out_time).strftime('%H:%M') if a.check_out_time else '',
                str(a.total_hours) if a.total_hours else '',
                a.get_type_display(),
                a.get_status_display(),
                loc.address if loc else '',
                a.note or '',
            ])
        return response


class ExportReportPDFView(AdminRequiredMixin, View):
    def get(self, request):
        # ── Filter (same logic as CSV export) ──────────────────────────
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        emp_id = request.GET.get('employee')
        branch_id = request.GET.get('branch')
        att_type = request.GET.get('type')
        status = request.GET.get('status')
        
        if status == 'absent':
            attendances = get_absent_records(
                date_from=date_from,
                date_to=date_to,
                employee_id=emp_id,
                branch_id=branch_id
            )
        else:
            qs = (
                Attendance.objects
                .select_related('employee', 'employee__branch')
            )
            qs = _filter_qs_by_request(qs, request)
            if att_type:
                qs = qs.filter(type=att_type)
            if status:
                if status == 'present':
                    qs = qs.filter(status__in=['on_time', 'late'])
                else:
                    qs = qs.filter(status=status)
            attendances = list(qs.order_by('date', 'employee__full_name'))

        # ── HTTP response ───────────────────────────────────────────────
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="attendance_report_{timezone.localdate()}.pdf"'
        )

        # ── ReportLab document ─────────────────────────────────────────
        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30,
        )

        styles  = getSampleStyleSheet()
        elements = []

        # Title & timestamp
        elements.append(Paragraph('Signtech Track — Attendance Report', styles['Title']))
        elements.append(Paragraph(
            f'Generated: {timezone.localtime(timezone.now()).strftime("%d %b %Y, %I:%M %p")}',
            styles['Normal'],
        ))
        elements.append(Spacer(1, 16))

        # ── Table ───────────────────────────────────────────────────────
        header = [
            'SN', 'Employee', 'Emp ID', 'Date',
            'Check-in', 'Check-out', 'Hours',
            'Type', 'Status', 'Notes',
        ]
        data = [header]

        for idx, a in enumerate(attendances, 1):
            branch = a.employee.branch.name if a.employee.branch else '—'
            data.append([
                str(idx),
                a.employee.full_name,
                a.employee.employee_id,
                str(a.date),
                timezone.localtime(a.check_in_time).strftime('%H:%M')  if a.check_in_time  else '—',
                timezone.localtime(a.check_out_time).strftime('%H:%M') if a.check_out_time else '—',
                f'{a.total_hours}h' if a.total_hours else '—',
                a.get_type_display(),
                a.get_status_display(),
                a.note or '—',
            ])

        col_widths = [25, 100, 50, 55, 50, 50, 40, 50, 50, 65]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
            ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, 0), 9),
            ('ALIGN',         (0, 0), (-1, 0), 'CENTER'),
            ('TOPPADDING',    (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            # Body rows
            ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE',      (0, 1), (-1, -1), 8),
            ('ALIGN',         (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 1), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
            # Alternating row backgrounds
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#F9FAFB')]),
            # Grid & border
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#E5E7EB')),
            ('BOX',  (0, 0), (-1, -1), 0.8, colors.HexColor('#D1D5DB')),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(
            'Signtech Track Attendance Management System',
            styles['Normal'],
        ))

        doc.build(elements)
        return response

# ─────────────────────────────────────────────────────────────────
# SCHEDULE SETTINGS
# ─────────────────────────────────────────────────────────────────
class OfficeScheduleView(AdminRequiredMixin, View):
    template_name = 'admin_panel/settings/schedule.html'

    def get(self, request):
        branches = Branch.objects.all()
        schedules = []
        for branch in branches:
            schedule, created = OfficeSchedule.objects.get_or_create(
                branch=branch,
                defaults={
                    'working_days': [
                        'saturday', 'sunday', 'monday',
                        'tuesday', 'wednesday', 'thursday'
                    ]
                }
            )
            schedules.append({
                'branch': branch,
                'schedule': schedule
            })
        TRACKING_CHOICES = [
            (1,   '1 minute'),
            (2,   '2 minutes'),
            (3,   '3 minutes'),
            (5,   '5 minutes'),
            (10,  '10 minutes'),
            (15,  '15 minutes'),
            (20,  '20 minutes'),
            (30,  '30 minutes'),
            (45,  '45 minutes'),
            (60,  '1 hour'),
            (90,  '1.5 hours'),
            (120, '2 hours'),
            (150, '2.5 hours'),
            (180, '3 hours'),
            (210, '3.5 hours'),
            (240, '4 hours'),
            (0,   'Disabled'),
        ]
        from apps.leave.models import LeaveType
        leave_types = LeaveType.objects.all().order_by('category', 'name')
        return render(request, self.template_name, {
            'schedules': schedules,
            'tracking_choices': TRACKING_CHOICES,
            'leave_types': leave_types
        })

    def post(self, request):
        branch_id = request.POST.get('branch_id')
        branch = get_object_or_404(Branch, id=branch_id)
        schedule = get_object_or_404(OfficeSchedule, branch=branch)

        schedule.office_start_time = request.POST.get('office_start_time', schedule.office_start_time)
        schedule.office_end_time = request.POST.get('office_end_time', schedule.office_end_time)
        schedule.late_after_minutes = int(request.POST.get('late_after_minutes', schedule.late_after_minutes))
        schedule.early_checkout_before_minutes = int(request.POST.get('early_checkout_before_minutes', schedule.early_checkout_before_minutes))
        schedule.overtime_after_minutes = int(request.POST.get('overtime_after_minutes', schedule.overtime_after_minutes))
        schedule.tracking_interval_minutes = int(request.POST.get('tracking_interval_minutes', schedule.tracking_interval_minutes))
        
        # Working days
        working_days = request.POST.getlist(f'working_days_{branch_id}')
        if working_days:
            schedule.working_days = working_days
        else:
            schedule.working_days = []

        schedule.save()
        messages.success(request, f"Schedule for {branch.name} updated successfully.")
        return redirect('admin_panel:schedule_settings')

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, 
    Border, Side
)
from openpyxl.utils import get_column_letter

@admin_required
def export_attendance(request):
    format_type = request.GET.get('format', 'xlsx')
    year = request.GET.get('year')
    month = request.GET.get('month')
    
    if year and month:
        if format_type == 'xlsx':
            return export_monthly_xlsx(request)
        elif format_type == 'csv':
            return export_monthly_csv(request)
        elif format_type == 'pdf':
            return export_monthly_pdf(request)
        else:
            return export_monthly_xlsx(request)
    else:
        if format_type == 'xlsx':
            return export_monthly_xlsx(request)
        elif format_type == 'csv':
            return ExportReportCSVView.as_view()(request)
        elif format_type == 'pdf':
            return ExportReportPDFView.as_view()(request)
        else:
            return export_monthly_xlsx(request)

def get_monthly_grid_data(request):
    import calendar
    from datetime import date, timedelta
    from collections import defaultdict
    from django.utils import timezone
    from apps.employees.models import EmployeeProfile
    from apps.branches.models import Branch, OfficeSchedule
    from apps.attendance.models import Attendance
    from apps.leave.models import LeaveRequest
    
    year = int(request.GET.get('year', date.today().year))
    month = int(request.GET.get('month', date.today().month))
    
    branch_id = request.GET.get('branch') or request.GET.get('branch_id')
    employee_id = request.GET.get('employee')
    
    # Get all days in month
    days_in_month = calendar.monthrange(year, month)[1]
    all_days = [date(year, month, d) for d in range(1, days_in_month + 1)]
    
    # Get employees
    employees = EmployeeProfile.objects.filter(is_active=True)
    if branch_id:
        employees = employees.filter(branch_id=branch_id)
    if employee_id:
        employees = employees.filter(id=employee_id)
    employees = employees.select_related('branch', 'branch__schedule').order_by('full_name')
    
    # Get all attendances for this month
    attendances = Attendance.objects.filter(
        date__year=year,
        date__month=month,
        employee__in=employees,
        is_expired=False
    ).select_related('employee')
    
    # Group attendance by employee and date
    att_by_emp_date = defaultdict(list)
    for att in attendances:
        att_by_emp_date[(att.employee_id, att.date)].append(att)
        
    # Get approved leave requests
    leave_requests = LeaveRequest.objects.filter(
        status='approved',
        start_date__lte=date(year, month, days_in_month),
        end_date__gte=date(year, month, 1),
        employee__in=employees
    ).select_related('leave_type')
    
    approved_leaves_map = defaultdict(dict)
    for req in leave_requests:
        s_date = max(req.start_date, date(year, month, 1))
        e_date = min(req.end_date, date(year, month, days_in_month))
        curr = s_date
        while curr <= e_date:
            approved_leaves_map[req.employee_id][curr] = req
            curr += timedelta(days=1)
            
    # Build employee stats and display lookups
    display_att_lookup = defaultdict(dict)  # {emp_id: {date: primary_attendance_for_display}}
    employee_stats = {}
    
    for emp in employees:
        schedule = _get_employee_schedule(emp)
        present_count = 0
        late_count = 0
        total_ot_minutes = 0
        absent_count = 0
        holiday_work_count = 0
        
        for d in all_days:
            day_atts = att_by_emp_date[(emp.id, d)]
            
            # Select the primary attendance record for display in In/Out columns
            # Prefer check_in session
            main_att = next((a for a in day_atts if a.attendance_type == 'check_in'), None)
            if not main_att and day_atts:
                main_att = day_atts[0]
            
            if main_att:
                display_att_lookup[emp.id][d] = main_att
                
            has_check_in = any(a.attendance_type == 'check_in' for a in day_atts)
            if has_check_in:
                present_count += 1
                
            has_field_visit = any(a.attendance_type == 'field_visit' for a in day_atts)
            has_any_attendance = has_check_in or has_field_visit
            
            # Count late and overtime from check_in sessions (using the main check-in for parity)
            if main_att and main_att.attendance_type == 'check_in':
                if main_att.status == 'late':
                    late_count += 1
                if getattr(main_att, 'overtime_minutes', 0) > 0:
                    total_ot_minutes += main_att.overtime_minutes
            
            is_work_day = _is_working_day(d, schedule)
            if is_work_day:
                if not has_any_attendance:
                    # Check if on approved leave
                    is_on_leave = d in approved_leaves_map[emp.id]
                    if not is_on_leave:
                        absent_count += 1
            else:
                # Holiday Work: check_in record on a date outside working days
                if has_check_in:
                    holiday_work_count += 1
                    
        # Overtime display format
        if getattr(emp, 'overtime_enabled', False) and total_ot_minutes > 0:
            ot_hours = total_ot_minutes / 60
            ot_display = f"{int(ot_hours)}h {int(total_ot_minutes % 60)}m"
        else:
            ot_display = '-'
            
        employee_stats[emp.id] = {
            'present_count': present_count,
            'late_count': late_count,
            'total_ot_minutes': total_ot_minutes,
            'absent_count': absent_count,
            'holiday_work_count': holiday_work_count,
            'overtime_display': ot_display,
            'is_overtime_enabled': getattr(emp, 'overtime_enabled', False)
        }
        
    return {
        'year': year,
        'month': month,
        'days_in_month': days_in_month,
        'all_days': all_days,
        'employees': employees,
        'att_lookup': display_att_lookup,
        'employee_stats': employee_stats,
        'approved_leaves': approved_leaves_map
    }
    
@admin_required
def export_monthly_xlsx(request):
    from datetime import date
    data = get_monthly_grid_data(request)
    year = data['year']
    month = data['month']
    all_days = data['all_days']
    employees = data['employees']
    att_lookup = data['att_lookup']
    employee_stats = data['employee_stats']
    approved_leaves = data['approved_leaves']
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    month_name = date(year, month, 1).strftime('%B-%y')
    ws.title = month_name
    
    # ==================
    # STYLES
    # ==================
    company_font = Font(
        name='Calibri', bold=True, size=14)
    header_font = Font(
        name='Calibri', bold=True, size=10,
        color='FFFFFF')
    subheader_font = Font(
        name='Calibri', bold=True, size=9)
    normal_font = Font(name='Calibri', size=9)
    late_font = Font(
        name='Calibri', size=9, color='FF0000')
    
    header_fill = PatternFill(
        'solid', fgColor='1F4E79')  # dark blue
    weekend_fill = PatternFill(
        'solid', fgColor='D9D9D9')  # gray
    late_fill = PatternFill(
        'solid', fgColor='FFE0E0')  # light red
    overtime_fill = PatternFill(
        'solid', fgColor='E0F0FF')  # light blue
    
    center = Alignment(
        horizontal='center', vertical='center',
        wrap_text=True)
    left = Alignment(
        horizontal='left', vertical='center')
    
    thin = Side(style='thin', color='000000')
    border = Border(
        left=thin, right=thin, 
        top=thin, bottom=thin)
    
    # ==================
    # ROW 1: Company name + Dates
    # ==================
    
    # Col A: Company name (merged A1:C1)
    ws.merge_cells('A1:C1')
    ws['A1'] = 'SIGNTECH TECHNOLOGY'  
    # Replace with actual company name from settings
    ws['A1'].font = company_font
    ws['A1'].alignment = center
    
    # Date columns start at col D
    # Each day has 2 columns: In | Out
    # So day 1 = cols D,E | day 2 = cols F,G | etc.
    
    col = 4  # Start at column D
    day_col_map = {}  # {day_number: start_col}
    
    for d in all_days:
        day_col_map[d.day] = col
        # Date in row 1
        ws.cell(row=1, column=col, 
                value=d).number_format = 'D'
        ws.merge_cells(
            start_row=1, start_column=col,
            end_row=1, end_column=col+1)
        cell = ws.cell(row=1, column=col)
        cell.value = d.day  # Just the day number
        cell.font = subheader_font
        cell.alignment = center
        cell.fill = header_fill
        cell.font = Font(
            name='Calibri', bold=True, 
            size=9, color='FFFFFF')
        
        col += 2
    
    # Last cols: Present | Absent | Late | OT | Holiday Work
    summary_start_col = col
    ws.cell(row=1, column=summary_start_col,
            value='Present').font = subheader_font
    ws.cell(row=1, column=summary_start_col+1,
            value='Absent').font = subheader_font
    ws.cell(row=1, column=summary_start_col+2,
            value='Late').font = subheader_font
    ws.cell(row=1, column=summary_start_col+3,
            value='Overtime').font = subheader_font
    ws.cell(row=1, column=summary_start_col+4,
            value='Holiday Work').font = subheader_font
    
    # ==================
    # ROW 2: Day names
    # ==================
    ws.cell(row=2, column=1, value='')
    ws.cell(row=2, column=2, value='')
    ws.cell(row=2, column=3, value='')
    
    day_names = ['Mon','Tue','Wed','Thu',
                 'Fri','Sat','Sun']
    
    col = 4
    for d in all_days:
        day_name = d.strftime('%a')  # Mon, Tue, etc.
        ws.merge_cells(
            start_row=2, start_column=col,
            end_row=2, end_column=col+1)
        cell = ws.cell(row=2, column=col)
        cell.value = day_name
        cell.font = Font(
            name='Calibri', bold=True, size=9)
        cell.alignment = center
        
        # Weekend styling
        if d.weekday() == 4:  # Friday
            cell.fill = PatternFill(
                'solid', fgColor='F2DCDB')
        
        col += 2
        
    ws.cell(row=2, column=summary_start_col, value='')
    ws.cell(row=2, column=summary_start_col+1, value='')
    ws.cell(row=2, column=summary_start_col+2, value='')
    ws.cell(row=2, column=summary_start_col+3, value='')
    ws.cell(row=2, column=summary_start_col+4, value='')
    
    # ==================
    # ROW 3: Headers
    # ==================
    ws.cell(row=3, column=1, 
            value='SN').font = Font(
        bold=True, size=9)
    ws.cell(row=3, column=2, 
            value='Employee name').font = Font(
        bold=True, size=9)
    ws.cell(row=3, column=3, 
            value='Designation').font = Font(
        bold=True, size=9)
    
    col = 4
    for d in all_days:
        ws.cell(row=3, column=col, 
                value='In').font = Font(
            bold=True, size=8)
        ws.cell(row=3, column=col+1, 
                value='Out').font = Font(
            bold=True, size=8)
        ws.cell(row=3, column=col).alignment = center
        ws.cell(row=3, column=col+1).alignment = center
        col += 2
    
    ws.cell(row=3, column=summary_start_col,
            value='Present')
    ws.cell(row=3, column=summary_start_col+1,
            value='Absent')
    ws.cell(row=3, column=summary_start_col+2,
            value='Late Days')
    ws.cell(row=3, column=summary_start_col+3,
            value='OT Hours')
    ws.cell(row=3, column=summary_start_col+4,
            value='Holiday Work')
    
    # Apply header fill to row 3
    for c in range(1, summary_start_col+5):
        cell = ws.cell(row=3, column=c)
        if not cell.value:
            continue
        cell.fill = PatternFill(
            'solid', fgColor='1F4E79')
        cell.font = Font(
            name='Calibri', bold=True,
            size=9, color='FFFFFF')
        cell.alignment = center
        cell.border = border
    
    # ==================
    # ROW 4 onward: Employee data
    # ==================
    
    data_row = 4
    
    for idx, emp in enumerate(employees, 1):
        ws.cell(row=data_row, column=1,
                value=idx).font = normal_font
        ws.cell(row=data_row, column=2,
                value=emp.full_name).font = normal_font
        ws.cell(row=data_row, column=3,
                value=emp.designation).font = normal_font
        
        emp_stat = employee_stats.get(emp.id, {})
        present_count = emp_stat.get('present_count', 0)
        late_count = emp_stat.get('late_count', 0)
        ot_display = emp_stat.get('overtime_display', '-')
        absent_count = emp_stat.get('absent_count', 0)
        holiday_work_count = emp_stat.get('holiday_work_count', 0)
        
        col = 4
        for d in all_days:
            att = att_lookup.get(emp.id, {}).get(d)
            
            in_cell = ws.cell(row=data_row, column=col)
            out_cell = ws.cell(
                row=data_row, column=col+1)
            
            if att:
                # Format time as decimal 
                # (like original: 9.05 = 9:05 AM)
                if att.check_in_time:
                    t = timezone.localtime(att.check_in_time)
                    time_val = (t.hour + 
                                t.minute/100)
                    in_cell.value = time_val
                    
                    # Late highlight
                    if att.status == 'late':
                        in_cell.fill = late_fill
                        in_cell.font = late_font
                
                if att.check_out_time:
                    t = timezone.localtime(att.check_out_time)
                    time_val = (t.hour + 
                                t.minute/100)
                    out_cell.value = time_val
                    
                    # Overtime highlight
                    if getattr(att, 'overtime_minutes', 0) > 0:
                        out_cell.fill = overtime_fill
                
                # Leave text
                if hasattr(att, 'leave_type') and att.leave_type:
                    in_cell.value = 'Leave'
                    in_cell.font = Font(
                        name='Calibri', size=8,
                        italic=True, 
                        color='0070C0')
            else:
                # check if on approved leave
                is_on_leave = d in approved_leaves.get(emp.id, {})
                if is_on_leave:
                    in_cell.value = 'Leave'
                    in_cell.font = Font(
                        name='Calibri', size=8,
                        italic=True, 
                        color='0070C0')
            
            # Weekend styling
            if d.weekday() == 4:  # Friday
                in_cell.fill = PatternFill(
                    'solid', fgColor='F2DCDB')
                out_cell.fill = PatternFill(
                    'solid', fgColor='F2DCDB')
            
            # Apply border and alignment
            for cell in [in_cell, out_cell]:
                cell.alignment = center
                cell.border = border
                cell.number_format = '0.00'
            
            col += 2
        
        # Summary columns: Present, Absent, Late Days, OT Hours, Holiday Work
        # Present
        present_cell = ws.cell(
            row=data_row,
            column=summary_start_col,
            value=present_count)
        present_cell.alignment = center
        present_cell.border = border
        present_cell.font = normal_font
        
        # Absent
        absent_cell = ws.cell(
            row=data_row,
            column=summary_start_col+1,
            value=absent_count)
        if absent_count > 0:
            absent_cell.font = late_font
            absent_cell.fill = late_fill
        else:
            absent_cell.font = normal_font
        absent_cell.alignment = center
        absent_cell.border = border
        
        # Late Days
        late_cell = ws.cell(
            row=data_row,
            column=summary_start_col+2,
            value=late_count)
        if late_count > 0:
            late_cell.font = late_font
            late_cell.fill = late_fill
        else:
            late_cell.font = normal_font
        late_cell.alignment = center
        late_cell.border = border
        
        # Overtime
        ot_cell = ws.cell(
            row=data_row,
            column=summary_start_col+3,
            value=ot_display)
        if ot_display != '-':
            ot_cell.fill = overtime_fill
        ot_cell.alignment = center
        ot_cell.border = border
        ot_cell.font = normal_font
        
        # Holiday Work
        hw_cell = ws.cell(
            row=data_row,
            column=summary_start_col+4,
            value=holiday_work_count)
        hw_cell.alignment = center
        hw_cell.border = border
        hw_cell.font = normal_font
        
        # Apply border to SN, name, and designation
        ws.cell(row=data_row, column=1).border = border
        ws.cell(row=data_row, column=1).alignment = center
        ws.cell(row=data_row, column=2).border = border
        ws.cell(row=data_row, column=3).border = border
        
        data_row += 1
    
    # ==================
    # COLUMN WIDTHS
    # ==================
    ws.column_dimensions['A'].width = 6   # SN
    ws.column_dimensions['B'].width = 22  # Employee name
    ws.column_dimensions['C'].width = 18  # Designation
    
    col = 4
    for d in all_days:
        col_letter = get_column_letter(col)
        col_letter2 = get_column_letter(col+1)
        ws.column_dimensions[col_letter].width = 6
        ws.column_dimensions[col_letter2].width = 6
        col += 2
    
    # Summary columns
    ws.column_dimensions[
        get_column_letter(summary_start_col)].width = 10
    ws.column_dimensions[
        get_column_letter(summary_start_col+1)].width = 10
    ws.column_dimensions[
        get_column_letter(summary_start_col+2)].width = 10
    ws.column_dimensions[
        get_column_letter(summary_start_col+3)].width = 10
    ws.column_dimensions[
        get_column_letter(summary_start_col+4)].width = 12
    
    # Row heights
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    for r in range(4, data_row):
        ws.row_dimensions[r].height = 18
    
    # Freeze panes (keep headers and metadata columns visible)
    ws.freeze_panes = 'D4'
    
    # ==================
    # HTTP RESPONSE
    # ==================
    filename = f"attendance_{month_name}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-'
                     'officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{filename}"')
    wb.save(response)
    return response

@admin_required
def export_monthly_csv(request):
    from datetime import date
    data = get_monthly_grid_data(request)
    year = data['year']
    month = data['month']
    employees = data['employees']
    employee_stats = data['employee_stats']
    
    month_name = date(year, month, 1).strftime('%B-%y')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="attendance_{month_name}.csv"'
    )
    
    writer = csv.writer(response)
    
    # Headers
    headers = ['SN', 'Employee Name', 'Designation', 'Present', 'Absent', 'Late Days', 'OT Hours', 'Holiday Work']
    writer.writerow(headers)
    
    for idx, emp in enumerate(employees, 1):
        emp_stat = employee_stats.get(emp.id, {})
        row = [
            idx,
            emp.full_name,
            emp.designation,
            emp_stat.get('present_count', 0),
            emp_stat.get('absent_count', 0),
            emp_stat.get('late_count', 0),
            emp_stat.get('overtime_display', '-'),
            emp_stat.get('holiday_work_count', 0)
        ]
        writer.writerow(row)
        
    return response

@admin_required
def export_monthly_pdf(request):
    from datetime import date
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    data = get_monthly_grid_data(request)
    year = data['year']
    month = data['month']
    employees = data['employees']
    employee_stats = data['employee_stats']
    
    month_name = date(year, month, 1).strftime('%B-%y')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="attendance_{month_name}.pdf"'
    )
    
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=15,
        leftMargin=15,
        topMargin=20,
        bottomMargin=20,
    )
    
    styles = getSampleStyleSheet()
    
    # Custom cell paragraph styles
    cell_style = ParagraphStyle(
        'GridCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        alignment=1 # Center
    )
    cell_style_left = ParagraphStyle(
        'GridCellLeft',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        alignment=0 # Left
    )
    late_text_style = ParagraphStyle(
        'LateText',
        parent=cell_style,
        textColor=colors.HexColor('#FF0000')
    )
    
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=16,
        textColor=colors.HexColor('#1F4E79'),
        alignment=0
    )
    
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#4B5563')
    )
    
    elements = []
    
    elements.append(Paragraph(f"Monthly Attendance Report — {date(year, month, 1).strftime('%B %Y')}", title_style))
    elements.append(Paragraph(f"Generated: {timezone.localtime().strftime('%d %b %Y, %I:%M %p')}", subtitle_style))
    elements.append(Spacer(1, 10))
    
    # Table headers
    headers = ['SN', 'Employee Name', 'Designation', 'Present', 'Absent', 'Late Days', 'OT Hours', 'Holiday Work']
    table_data = [headers]
    
    t_style = [
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (2, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#D1D5DB')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
    ]
    
    r_idx = 1
    for idx, emp in enumerate(employees, 1):
        emp_stat = employee_stats.get(emp.id, {})
        present_count = emp_stat.get('present_count', 0)
        absent_count = emp_stat.get('absent_count', 0)
        late_count = emp_stat.get('late_count', 0)
        ot_display = emp_stat.get('overtime_display', '-')
        holiday_work_count = emp_stat.get('holiday_work_count', 0)
        
        row = [
            Paragraph(str(idx), cell_style),
            Paragraph(emp.full_name, cell_style_left),
            Paragraph(emp.designation, cell_style_left),
            Paragraph(str(present_count), cell_style),
            Paragraph(str(absent_count), cell_style),
            Paragraph(str(late_count), cell_style),
            Paragraph(ot_display, cell_style),
            Paragraph(str(holiday_work_count), cell_style)
        ]
        
        # Color highlights
        if absent_count > 0:
            t_style.append(('BACKGROUND', (4, r_idx), (4, r_idx), colors.HexColor('#FFE0E0')))
            row[4] = Paragraph(str(absent_count), late_text_style)
            
        if late_count > 0:
            t_style.append(('BACKGROUND', (5, r_idx), (5, r_idx), colors.HexColor('#FFE0E0')))
            row[5] = Paragraph(str(late_count), late_text_style)
            
        if ot_display != '-':
            t_style.append(('BACKGROUND', (6, r_idx), (6, r_idx), colors.HexColor('#E0F0FF')))
            
        table_data.append(row)
        r_idx += 1
        
    col_widths = [30, 160, 140, 70, 70, 70, 80, 80]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle(t_style))
    elements.append(table)
    
    doc.build(elements)
    return response

# ─────────────────────────────────────────────────────────────────
# EXPIRED DATA MANAGEMENT
# ─────────────────────────────────────────────────────────────────
from apps.attendance.retention import get_retention_stats
from apps.attendance.models import AttendanceLocation

class ExpiredDataView(AdminRequiredMixin, ListView):
    model = Attendance
    template_name = 'admin_panel/expired_data.html'
    context_object_name = 'attendances'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = super().get_queryset().filter(is_expired=True).select_related('employee', 'employee__branch')
        
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        emp_id = self.request.GET.get('employee')
        branch_id = self.request.GET.get('branch')
        
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        if emp_id:
            queryset = queryset.filter(employee_id=emp_id)
        if branch_id:
            queryset = queryset.filter(employee__branch_id=branch_id)
            
        return queryset.order_by('-expired_at', '-date')
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        stats = get_retention_stats()
        now = timezone.now()
        auto_delete_threshold = now + timedelta(days=30)

        for att in context['attendances']:
            if att.expired_at:
                auto_delete_on = att.expired_at + relativedelta(months=2)
                att.auto_delete_on = auto_delete_on
                att.auto_delete_soon = now <= auto_delete_on <= auto_delete_threshold
            else:
                att.auto_delete_on = None
                att.auto_delete_soon = False

        context['stats'] = stats
        context['employees'] = EmployeeProfile.objects.all()
        context['branches'] = Branch.objects.all()
        context['selected_employee'] = self.request.GET.get('employee', '')
        context['selected_branch'] = self.request.GET.get('branch', '')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        get_copy = self.request.GET.copy()
        if 'page' in get_copy:
            del get_copy['page']
        context['query_string'] = get_copy.urlencode()
        return context

@admin_required
def delete_expired_selected(request):
    if request.method == 'POST':
        ids = request.POST.getlist('record_ids')
        if ids:
            AttendanceLocation.objects.filter(
                attendance_id__in=ids,
                attendance__is_expired=True
            ).delete()
            Attendance.objects.filter(
                id__in=ids,
                is_expired=True
            ).delete()
            messages.success(request, f"Deleted {len(ids)} selected records.")
        return redirect('admin_panel:expired_data')

@admin_required
def delete_all_expired(request):
    if request.method == 'POST':
        AttendanceLocation.objects.filter(
            attendance__is_expired=True
        ).delete()
        count, _ = Attendance.objects.filter(
            is_expired=True).delete()
        messages.success(request, f"Deleted all {count} expired records.")
        return redirect('admin_panel:expired_data')


class AbsentReportView(AdminRequiredMixin, ListView):
    template_name = 'admin_panel/reports/absent_report.html'
    context_object_name = 'absences'
    paginate_by = 25

    def get_queryset(self):
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        emp_id = self.request.GET.get('employee')
        branch_id = self.request.GET.get('branch')
        lt_id = self.request.GET.get('leave_type')

        return get_unified_deductions(
            date_from=date_from,
            date_to=date_to,
            employee_id=emp_id,
            branch_id=branch_id,
            leave_type_id=lt_id
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from apps.employees.models import EmployeeProfile
        from apps.branches.models import Branch
        from apps.leave.models import LeaveType, LeaveBalance

        # Populate live snapshot of LeaveBalance remaining days
        for absence in context['absences']:
            year = absence['date'].year
            lt = absence['leave_type_deducted']
            if lt:
                bal = LeaveBalance.objects.filter(employee=absence['employee'], leave_type=lt, year=year).first()
                if bal:
                    absence['remaining_days'] = bal.remaining_days
                else:
                    from apps.employees.models import EmployeeLeaveRule
                    rule = EmployeeLeaveRule.objects.filter(employee=absence['employee'], leave_type=lt).first()
                    limit = rule.days_per_year if rule else lt.default_days_per_year
                    absence['remaining_days'] = limit
            else:
                absence['remaining_days'] = None

        context.update({
            'employees': EmployeeProfile.objects.all().order_by('full_name'),
            'branches': Branch.objects.all().order_by('name'),
            'leave_types': LeaveType.objects.all().order_by('name'),
            'selected_employee': self.request.GET.get('employee', ''),
            'selected_branch': self.request.GET.get('branch', ''),
            'selected_leave_type': self.request.GET.get('leave_type', ''),
            'date_from': self.request.GET.get('date_from', ''),
            'date_to': self.request.GET.get('date_to', ''),
        })
        
        # Build query string for pagination links
        get_copy = self.request.GET.copy()
        if 'page' in get_copy:
            del get_copy['page']
        context['query_string'] = get_copy.urlencode()
        
        return context


class ExportAbsentReportExcelView(AdminRequiredMixin, View):
    def get(self, request):
        from apps.leave.models import LeaveBalance
        from django.utils import timezone
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        emp_id = request.GET.get('employee')
        branch_id = request.GET.get('branch')
        lt_id = request.GET.get('leave_type')

        deductions = get_unified_deductions(
            date_from=date_from,
            date_to=date_to,
            employee_id=emp_id,
            branch_id=branch_id,
            leave_type_id=lt_id
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Absence & Leave Deductions"

        company_font = Font(name='Calibri', bold=True, size=14)
        header_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        normal_font = Font(name='Calibri', size=9)
        
        header_fill = PatternFill('solid', fgColor='1F4E79')  # dark blue
        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center')
        
        thin = Side(style='thin', color='E5E7EB')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells('A1:G1')
        ws['A1'] = 'Absence & Leave Deductions Report'
        ws['A1'].font = company_font
        ws['A1'].alignment = center
        ws.row_dimensions[1].height = 30

        headers = ['SN', 'Employee Name', 'Employee ID', 'Date', 'Leave Type Deducted', 'Remaining Balance', 'Branch']
        ws.row_dimensions[2].height = 20
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        data_row = 3
        for idx, item in enumerate(deductions, 1):
            emp = item['employee']
            dt = item['date']
            lt = item['leave_type_deducted']
            
            if lt:
                bal = LeaveBalance.objects.filter(employee=emp, leave_type=lt, year=dt.year).first()
                if bal:
                    remaining = bal.remaining_days
                else:
                    from apps.employees.models import EmployeeLeaveRule
                    rule = EmployeeLeaveRule.objects.filter(employee=emp, leave_type=lt).first()
                    remaining = rule.days_per_year if rule else lt.default_days_per_year
                remaining_str = f"{remaining} days"
                lt_name = lt.name
            else:
                remaining_str = "—"
                lt_name = "—"

            branch_name = emp.branch.name if emp.branch else '—'

            ws.cell(row=data_row, column=1, value=idx).alignment = center
            ws.cell(row=data_row, column=2, value=emp.full_name).alignment = left
            ws.cell(row=data_row, column=3, value=emp.employee_id).alignment = center
            ws.cell(row=data_row, column=4, value=str(dt)).alignment = center
            ws.cell(row=data_row, column=5, value=lt_name).alignment = center
            ws.cell(row=data_row, column=6, value=remaining_str).alignment = center
            ws.cell(row=data_row, column=7, value=branch_name).alignment = left

            for c in range(1, 8):
                cell = ws.cell(row=data_row, column=c)
                cell.font = normal_font
                cell.border = border

            ws.row_dimensions[data_row].height = 18
            data_row += 1

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20

        filename = f"absence_report_{timezone.localdate()}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class ExportAbsentReportPDFView(AdminRequiredMixin, View):
    def get(self, request):
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from apps.leave.models import LeaveBalance
        from django.utils import timezone

        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        emp_id = request.GET.get('employee')
        branch_id = request.GET.get('branch')
        lt_id = request.GET.get('leave_type')

        deductions = get_unified_deductions(
            date_from=date_from,
            date_to=date_to,
            employee_id=emp_id,
            branch_id=branch_id,
            leave_type_id=lt_id
        )

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="absence_report_{timezone.localdate()}.pdf"'

        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30,
        )

        styles = getSampleStyleSheet()
        elements = []

        # Minimal styles
        cell_style = ParagraphStyle(
            'GridCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            leading=10,
            alignment=1  # Center
        )
        cell_style_left = ParagraphStyle(
            'GridCellLeft',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            leading=10,
            alignment=0  # Left
        )
        header_cell_style = ParagraphStyle(
            'HeaderCell',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#374151'),
            alignment=1
        )
        header_cell_style_left = ParagraphStyle(
            'HeaderCellLeft',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#374151'),
            alignment=0
        )
        
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=14,
            leading=16,
            textColor=colors.HexColor('#111827'),
            alignment=0
        )
        
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#6B7280')
        )

        elements.append(Paragraph('Signtech Track — Absence & Leave Deductions Report', title_style))
        elements.append(Paragraph(
            f'Generated: {timezone.localtime(timezone.now()).strftime("%d %b %Y, %I:%M %p")}',
            subtitle_style,
        ))
        elements.append(Spacer(1, 16))

        headers = [
            Paragraph('SN', header_cell_style),
            Paragraph('Employee', header_cell_style_left),
            Paragraph('Emp ID', header_cell_style),
            Paragraph('Date', header_cell_style),
            Paragraph('Leave Type Deducted', header_cell_style),
            Paragraph('Remaining Balance', header_cell_style),
            Paragraph('Branch', header_cell_style_left)
        ]
        data = [headers]

        for idx, item in enumerate(deductions, 1):
            emp = item['employee']
            dt = item['date']
            lt = item['leave_type_deducted']
            
            if lt:
                bal = LeaveBalance.objects.filter(employee=emp, leave_type=lt, year=dt.year).first()
                if bal:
                    remaining = bal.remaining_days
                else:
                    from apps.employees.models import EmployeeLeaveRule
                    rule = EmployeeLeaveRule.objects.filter(employee=emp, leave_type=lt).first()
                    remaining = rule.days_per_year if rule else lt.default_days_per_year
                remaining_str = f"{remaining} days"
                lt_name = lt.name
            else:
                remaining_str = "—"
                lt_name = "—"

            branch_name = emp.branch.name if emp.branch else '—'

            data.append([
                Paragraph(str(idx), cell_style),
                Paragraph(emp.full_name, cell_style_left),
                Paragraph(emp.employee_id, cell_style),
                Paragraph(str(dt), cell_style),
                Paragraph(lt_name, cell_style),
                Paragraph(remaining_str, cell_style),
                Paragraph(branch_name, cell_style_left),
            ])

        col_widths = [30, 130, 60, 70, 95, 75, 75]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#F3F4F6')),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING',    (0, 0), (-1, 0), 6),
            ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING',    (0, 1), (-1, -1), 6),
            ('LINEBELOW',     (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 20))
        
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=7,
            textColor=colors.HexColor('#9CA3AF'),
            alignment=1
        )
        elements.append(Paragraph('Signtech Track Attendance Management System', footer_style))

        doc.build(elements)
        return response


class AdminAddLeaveView(AdminRequiredMixin, CreateView):
    template_name = 'admin_panel/reports/add_leave.html'
    success_url = '/admin-panel/reports/absent/'

    def get_form_class(self):
        from apps.leave.forms import AdminAddLeaveForm
        return AdminAddLeaveForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from apps.employees.models import EmployeeProfile
        from apps.leave.models import LeaveType
        context['employees'] = EmployeeProfile.objects.filter(is_active=True).order_by('full_name')
        context['leave_types'] = LeaveType.objects.all().order_by('name')
        return context

    def form_valid(self, form):
        form.instance.reviewed_by = self.request.user
        form.instance.reviewed_at = timezone.now()
        
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Leave record for {form.instance.employee.full_name} added successfully."
        )
        return response
