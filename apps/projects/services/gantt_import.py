"""
Gantt Excel Import Service

Handles safe, multi-format Excel workbook ingestion for project task scheduling.
Responsibilities:
- Workbook safety validation (archive size, macros, encryption, mime/extension)
- Sheet discovery & metadata inspection
- Format detection (Structured Table, Offset Planner, Zone Matrix, Visual Monthly)
- Header normalization & alias resolution
- Raw row extraction & normalization
- Field validation (date order, duration consistency, percent bounds)
- Duplicate detection against existing project tasks
- Staged batch management
- Idempotent transactional confirmation & execution
"""

import io
import os
import zipfile
import hashlib
from datetime import date, datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple

import openpyxl
from django.core.exceptions import ValidationError, PermissionDenied
from django.db import transaction
from django.utils import timezone
from django.contrib.auth import get_user_model

from apps.projects.models import Project, ProjectTask
from apps.employees.models import EmployeeProfile
from apps.audit.services import AuditService

User = get_user_model()

# Configuration limits (code-level, no env refactor needed)
MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024       # 10 MB
MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024   # 50 MB (zip bomb guard)
MAX_SHEETS = 30
MAX_ROWS_PER_SHEET = 2500
MAX_COLS_PER_SHEET = 120
ALLOWED_EXTENSIONS = ('.xlsx',)


class GanttImportError(Exception):
    """Base domain exception for Gantt import errors."""
    def __init__(self, message: str, code: str = "import_error", field: Optional[str] = None):
        super().__init__(message)
        self.message = message
        self.code = code
        self.field = field


class WorkbookSafetyValidator:
    """Validates archive integrity, size limits, and security properties."""

    @staticmethod
    def validate_file(file_obj, filename: str) -> Tuple[bytes, str]:
        """
        Reads file bytes, validates extension, size, zip structure, macros, and returns
        (content_bytes, sha256_checksum).
        """
        if not filename:
            raise GanttImportError("No filename provided.", code="missing_filename")

        ext = os.path.splitext(filename)[1].lower()
        if ext not in ALLOWED_EXTENSIONS:
            raise GanttImportError(
                f"Unsupported file format '{ext}'. Only .xlsx files are supported.",
                code="invalid_extension",
                field="file"
            )

        # Read content
        if hasattr(file_obj, 'read'):
            content = file_obj.read()
            if hasattr(file_obj, 'seek'):
                file_obj.seek(0)
        elif isinstance(file_obj, bytes):
            content = file_obj
        else:
            raise GanttImportError("Invalid file source provided.", code="invalid_source")

        if len(content) == 0:
            raise GanttImportError("Uploaded file is empty.", code="empty_file", field="file")

        if len(content) > MAX_FILE_SIZE_BYTES:
            raise GanttImportError(
                f"File size ({len(content) / (1024*1024):.1f} MB) exceeds maximum allowed size "
                f"({MAX_FILE_SIZE_BYTES / (1024*1024):.0f} MB).",
                code="file_too_large",
                field="file"
            )

        sha256_hash = hashlib.sha256(content).hexdigest()

        # Validate zip structure and uncompressed ratio
        try:
            with zipfile.ZipFile(io.BytesIO(content), 'r') as z:
                # Check for encryption
                for zinfo in z.infolist():
                    if zinfo.flag_bits & 0x1:
                        raise GanttImportError(
                            "Password-protected or encrypted Excel files are not supported.",
                            code="encrypted_file",
                            field="file"
                        )
                # Check uncompressed size
                total_uncompressed = sum(zinfo.file_size for zinfo in z.infolist())
                if total_uncompressed > MAX_UNCOMPRESSED_BYTES:
                    raise GanttImportError(
                        "File archive expands beyond safe limits (potential zip bomb).",
                        code="zip_bomb_detected",
                        field="file"
                    )

                # Check for macros or VBA
                vba_files = [n for n in z.namelist() if 'vbaProject.bin' in n or 'xl/drawings/vmlDrawing' in n]
                if vba_files:
                    raise GanttImportError(
                        "Macro-enabled workbooks containing executable scripts are strictly prohibited.",
                        code="macro_prohibited",
                        field="file"
                    )
        except zipfile.BadZipFile:
            raise GanttImportError(
                "Uploaded file is corrupted or not a valid XLSX workbook.",
                code="corrupt_archive",
                field="file"
            )

        return content, sha256_hash


class GanttFormatDetector:
    """Detects layout structure of a given worksheet."""

    FORMAT_STRUCTURED = "structured_table"
    FORMAT_OFFSET = "offset_planner"
    FORMAT_ZONE_MATRIX = "zone_matrix"
    FORMAT_VISUAL_MONTHLY = "visual_monthly"
    FORMAT_UNKNOWN = "unknown"

    @classmethod
    def inspect_sheet(cls, rows: List[Tuple]) -> Dict[str, Any]:
        """
        Inspects top rows of worksheet to detect format and metadata.
        Returns:
            {
                'format': format_type,
                'header_row_idx': int,
                'base_date': Optional[date],
                'headers': List[str],
                'columns_meta': Dict,
            }
        """
        if not rows:
            return {'format': cls.FORMAT_UNKNOWN, 'header_row_idx': 0, 'headers': []}

        # Check for Zone Matrix layout
        # Characteristic: Row with WORK / ZONE or ZONE, and subsequent row with START / END column pairs
        for r_idx in range(min(5, len(rows))):
            row = [str(c).strip().upper() if c is not None else '' for c in rows[r_idx]]
            first_col = row[0] if row else ''
            if any(k in first_col for k in ('WORK / ZONE', 'ZONE', 'WORK/ZONE', 'LOCATION')):
                # Check next row for START / END
                if r_idx + 1 < len(rows):
                    next_row = [str(c).strip().upper() if c is not None else '' for c in rows[r_idx + 1]]
                    start_end_count = sum(1 for c in next_row if c in ('START', 'END', 'FINISH'))
                    if start_end_count >= 2:
                        return {
                            'format': cls.FORMAT_ZONE_MATRIX,
                            'header_row_idx': r_idx + 1,
                            'sub_header_idx': r_idx + 2,
                            'activity_headers': row,
                            'sub_headers': next_row,
                            'base_date': None
                        }

        # Check for Visual Monthly layout (e.g., 'DESCRIPTION' with month date columns in header row)
        for r_idx in range(min(5, len(rows))):
            raw_row = rows[r_idx]
            row_strs = [str(c).strip().upper() if c is not None else '' for c in raw_row]
            has_desc = any(k in row_strs for k in ('DESCRIPTION', 'ACTIVITY', 'TASK', 'WORK ITEM'))
            has_plan_start = any(k in row_strs for k in ('PLAN START', 'PLAN_START', 'PLANNED START', 'START DATE', 'DURATION', 'PLAN DURATION', 'PLAN END', 'PLAN FINISH'))
            date_cols = sum(1 for c in raw_row if isinstance(c, (datetime, date)) or (isinstance(c, str) and any(m in c.upper() for m in ('202', 'JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC'))))
            if has_desc and not has_plan_start and date_cols >= 2:
                return {
                    'format': cls.FORMAT_VISUAL_MONTHLY,
                    'header_row_idx': r_idx + 1,
                    'headers': [str(c).strip() if c is not None else '' for c in raw_row],
                    'base_date': None
                }

        # Check for Structured Table vs Offset Planner
        # Both may have ACTIVITY, PLAN START, PLAN DURATION, etc.
        # Check rows 1 to 10 for ACTIVITY header
        for r_idx in range(min(10, len(rows))):
            row = [str(c).strip().upper() if c is not None else '' for c in rows[r_idx]]
            if 'ACTIVITY' in row or 'TASK' in row or 'DESCRIPTION' in row or 'TASK NAME' in row:
                header_row_idx = r_idx + 1
                headers = row

                # Check if this sheet is an offset planner
                # Offset planner signs:
                # 1) Row above or below contains a 'START' cell with a base date (e.g. Row 4 cell with START and date)
                # 2) Data rows have integer numbers in PLAN START (e.g., 1, 4, 7, 16) instead of datetime
                base_date, date_calc_note = cls._find_timeline_base_date(rows, header_row_idx)

                # Sample data rows below header to distinguish offset from date-based
                sample_is_offset = cls._sample_is_offset(rows, header_row_idx, headers)

                if sample_is_offset and base_date:
                    return {
                        'format': cls.FORMAT_OFFSET,
                        'header_row_idx': header_row_idx,
                        'headers': headers,
                        'base_date': base_date,
                        'date_calc_note': date_calc_note
                    }
                elif 'PLAN END' in row or 'PLAN FINISH' in row or not sample_is_offset:
                    return {
                        'format': cls.FORMAT_STRUCTURED,
                        'header_row_idx': header_row_idx,
                        'headers': headers,
                        'base_date': base_date
                    }
                else:
                    return {
                        'format': cls.FORMAT_OFFSET if base_date else cls.FORMAT_STRUCTURED,
                        'header_row_idx': header_row_idx,
                        'headers': headers,
                        'base_date': base_date,
                        'date_calc_note': date_calc_note
                    }

        # Check for Visual Monthly layout (e.g., 'DESCRIPTION' with month date columns in header row 1)
        first_row = [c for c in rows[0] if c is not None]
        first_row_strs = [str(c).strip().upper() for c in first_row]
        has_desc = any(k in first_row_strs for k in ('DESCRIPTION', 'ACTIVITY', 'TASK'))
        # Check if row 1 contains date objects or month strings
        date_cols = sum(1 for c in first_row if isinstance(c, (datetime, date)) or ('202' in str(c)))
        if has_desc and date_cols >= 2:
            return {
                'format': cls.FORMAT_VISUAL_MONTHLY,
                'header_row_idx': 1,
                'headers': [str(c).strip() if c is not None else '' for c in rows[0]],
                'base_date': None
            }

        return {
            'format': cls.FORMAT_UNKNOWN,
            'header_row_idx': 1,
            'headers': [str(c).strip() if c is not None else '' for c in rows[0]] if rows else []
        }

    @classmethod
    def _find_timeline_base_date(cls, rows: List[Tuple], header_row_idx: int) -> Tuple[Optional[date], str]:
        """Searches around header rows for a timeline base date anchor."""
        for r_idx in range(min(12, len(rows))):
            row = rows[r_idx]
            for col_idx, cell in enumerate(row):
                if isinstance(cell, (datetime, date)):
                    d = cell.date() if isinstance(cell, datetime) else cell
                    return d, f"Detected timeline base date {d.isoformat()} from cell (Row {r_idx+1}, Col {col_idx+1})"
                # Also check if adjacent cell says START and next has a date
                val_str = str(cell).strip().upper() if cell is not None else ''
                if val_str == 'START' and col_idx + 1 < len(row):
                    next_val = row[col_idx + 1]
                    if isinstance(next_val, (datetime, date)):
                        d = next_val.date() if isinstance(next_val, datetime) else next_val
                        return d, f"Detected timeline base date {d.isoformat()} from 'START' cell header"
                    # Try parsing date string
                    parsed = cls._parse_date(next_val)
                    if parsed:
                        return parsed, f"Detected timeline base date {parsed.isoformat()} from 'START' header"

        return None, ""

    @classmethod
    def _sample_is_offset(cls, rows: List[Tuple], header_row_idx: int, headers: List[str]) -> bool:
        """Inspects first few data rows to see if PLAN START values are small integers."""
        plan_start_idx = -1
        for idx, h in enumerate(headers):
            if h in ('PLAN START', 'START', 'START DAY', 'OFFSET'):
                plan_start_idx = idx
                break

        if plan_start_idx == -1:
            return False

        numeric_offsets = 0
        date_objects = 0
        for r_idx in range(header_row_idx, min(header_row_idx + 8, len(rows))):
            row = rows[r_idx]
            if len(row) > plan_start_idx:
                val = row[plan_start_idx]
                if isinstance(val, (datetime, date)):
                    date_objects += 1
                elif isinstance(val, (int, float)) and 0 <= val <= 1000:
                    numeric_offsets += 1
                elif isinstance(val, str) and val.strip().isdigit():
                    numeric_offsets += 1

        return numeric_offsets > date_objects

    @staticmethod
    def _parse_date(val: Any) -> Optional[date]:
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        s = str(val).strip()
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%d %b %Y', '%d %B %Y'):
            try:
                return datetime.strptime(s.split(' ')[0], fmt).date()
            except (ValueError, IndexError):
                continue
        return None


class GanttWorkbookParser:
    """Parses worksheets into normalized task candidate records."""

    # Common column aliases for structured sheets
    HEADER_ALIASES = {
        'activity': ['ACTIVITY', 'TASK', 'TASK NAME', 'NAME', 'DESCRIPTION', 'TITLE', 'WORK ITEM'],
        'order': ['ORDER', 'NO', 'NO.', 'SL', 'SL NO', '#', 'TASK ID', 'ID'],
        'planned_start': ['PLAN START', 'PLANNED START', 'START DATE', 'START', 'PLAN_START'],
        'planned_finish': ['PLAN END', 'PLAN FINISH', 'PLANNED FINISH', 'END DATE', 'END', 'FINISH', 'PLAN_FINISH'],
        'duration_days': ['PLAN DURATION', 'DURATION', 'DURATION DAYS', 'DAYS', 'ESTIMATED DURATION'],
        'actual_start': ['ACTUAL START', 'REAL START', 'STARTED AT'],
        'actual_finish': ['ACTUAL FINISH', 'ACTUAL END', 'REAL FINISH', 'COMPLETED AT'],
        'progress_percent': ['PERCENT COMPLETE', '% COMPLETE', 'PROGRESS', 'PROGRESS %', 'PERCENTAGE', 'PROGRESS_PERCENT'],
        'responsible': ['RESPONSIBLE', 'ASSIGNED TO', 'OWNER', 'PERSON', 'ENGINEER', 'RESPONSIBLE PERSON'],
        'status': ['STATUS', 'STATE'],
        'milestone': ['MILESTONE', 'IS MILESTONE', 'MS']
    }

    def __init__(self, file_content: bytes, filename: str):
        self.content = file_content
        self.filename = filename
        self._wb = None
        self._sheet_cache = {}

    def get_workbook(self) -> openpyxl.Workbook:
        if self._wb is None:
            self._wb = openpyxl.load_workbook(
                io.BytesIO(self.content),
                data_only=True,
                read_only=True
            )
        return self._wb

    def close(self):
        if self._wb:
            try:
                self._wb.close()
            except Exception:
                pass
            self._wb = None

    def discover_sheets(self) -> List[Dict[str, Any]]:
        """Returns list of sheet summaries: name, row count, detected format, importable status."""
        wb = self.get_workbook()
        sheets_info = []

        for name in wb.sheetnames[:MAX_SHEETS]:
            ws = wb[name]
            # Read first 25 rows for format detection
            rows = []
            row_count = 0
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if r_idx > MAX_ROWS_PER_SHEET:
                    break
                if any(c is not None for c in row):
                    row_count += 1
                    if len(rows) < 25:
                        rows.append(row[:MAX_COLS_PER_SHEET])

            meta = GanttFormatDetector.inspect_sheet(rows)
            sheets_info.append({
                'name': name,
                'detected_format': meta['format'],
                'total_rows': row_count,
                'header_row': meta.get('header_row_idx', 1),
                'base_date': meta.get('base_date').isoformat() if meta.get('base_date') else None,
                'calc_note': meta.get('date_calc_note', '')
            })

        return sheets_info

    def parse_sheet(self, sheet_name: str, default_employee_id: Optional[int] = None) -> Dict[str, Any]:
        """
        Parses specified sheet into normalized raw rows and validation results.
        Returns:
            {
                'sheet_name': str,
                'detected_format': str,
                'base_date': Optional[str],
                'rows': List[Dict],
                'stats': {
                    'total': int,
                    'valid': int,
                    'warning': int,
                    'invalid': int,
                    'importable': int
                }
            }
        """
        wb = self.get_workbook()
        if sheet_name not in wb.sheetnames:
            raise GanttImportError(f"Sheet '{sheet_name}' does not exist in workbook.", code="sheet_not_found")

        ws = wb[sheet_name]
        raw_rows = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if r_idx > MAX_ROWS_PER_SHEET:
                break
            raw_rows.append((r_idx, row[:MAX_COLS_PER_SHEET]))

        meta = GanttFormatDetector.inspect_sheet([r[1] for r in raw_rows[:25]])
        meta['sheet_name'] = sheet_name
        detected_format = meta['format']
        base_date = meta.get('base_date')

        if detected_format == GanttFormatDetector.FORMAT_ZONE_MATRIX:
            normalized_rows = self._parse_zone_matrix(raw_rows, meta, default_employee_id)
        elif detected_format == GanttFormatDetector.FORMAT_OFFSET:
            normalized_rows = self._parse_offset_sheet(raw_rows, meta, default_employee_id)
        elif detected_format == GanttFormatDetector.FORMAT_VISUAL_MONTHLY:
            normalized_rows = self._parse_visual_monthly(raw_rows, meta, default_employee_id, sheet_name=sheet_name)
        else:
            # Standard structured table or unknown
            normalized_rows = self._parse_structured_table(raw_rows, meta, default_employee_id)

        # Aggregate summary statistics
        total = len(normalized_rows)
        valid = sum(1 for r in normalized_rows if not r['errors'] and not r['warnings'])
        warning = sum(1 for r in normalized_rows if r['warnings'] and not r['errors'])
        invalid = sum(1 for r in normalized_rows if r['errors'])
        importable = sum(1 for r in normalized_rows if not r['errors'] and not r.get('excluded', False))

        return {
            'sheet_name': sheet_name,
            'detected_format': detected_format,
            'base_date': base_date.isoformat() if base_date else None,
            'calc_note': meta.get('date_calc_note', ''),
            'rows': normalized_rows,
            'stats': {
                'total': total,
                'valid': valid,
                'warning': warning,
                'invalid': invalid,
                'importable': importable
            }
        }

    # ── Specialized Parsers ──────────────────────────────────────────────────

    def _parse_structured_table(self, raw_rows: List[Tuple], meta: Dict, default_emp_id: Optional[int]) -> List[Dict]:
        header_idx = meta.get('header_row_idx', 1)
        headers = meta.get('headers', [])
        col_map = self._map_columns(headers)

        parsed_rows = []
        for r_num, row in raw_rows[header_idx:]:
            if not any(c is not None for c in row):
                continue

            activity_val = self._get_col_val(row, col_map.get('activity'))
            if not activity_val or not str(activity_val).strip():
                # Skip empty activity rows
                continue

            activity = str(activity_val).strip()
            raw_start = self._get_col_val(row, col_map.get('planned_start'))
            raw_finish = self._get_col_val(row, col_map.get('planned_finish'))
            raw_duration = self._get_col_val(row, col_map.get('duration_days'))
            raw_act_start = self._get_col_val(row, col_map.get('actual_start'))
            raw_act_finish = self._get_col_val(row, col_map.get('actual_finish'))
            raw_progress = self._get_col_val(row, col_map.get('progress_percent'))
            raw_status = self._get_col_val(row, col_map.get('status'))
            raw_order = self._get_col_val(row, col_map.get('order'))
            raw_resp = self._get_col_val(row, col_map.get('responsible'))

            item = self._build_normalized_item(
                source_sheet=meta.get('sheet_name', ''),
                source_row=r_num,
                activity=activity,
                raw_start=raw_start,
                raw_finish=raw_finish,
                raw_duration=raw_duration,
                raw_act_start=raw_act_start,
                raw_act_finish=raw_act_finish,
                raw_progress=raw_progress,
                raw_status=raw_status,
                raw_order=raw_order,
                raw_resp=raw_resp,
                default_emp_id=default_emp_id
            )
            parsed_rows.append(item)

        return parsed_rows

    def _parse_offset_sheet(self, raw_rows: List[Tuple], meta: Dict, default_emp_id: Optional[int]) -> List[Dict]:
        header_idx = meta.get('header_row_idx', 1)
        headers = meta.get('headers', [])
        base_date = meta.get('base_date')
        col_map = self._map_columns(headers)

        parsed_rows = []
        for r_num, row in raw_rows[header_idx:]:
            if not any(c is not None for c in row):
                continue

            activity_val = self._get_col_val(row, col_map.get('activity'))
            if not activity_val or not str(activity_val).strip():
                continue

            activity = str(activity_val).strip()
            raw_start = self._get_col_val(row, col_map.get('planned_start'))
            raw_duration = self._get_col_val(row, col_map.get('duration_days'))
            raw_finish = self._get_col_val(row, col_map.get('planned_finish'))
            raw_act_start = self._get_col_val(row, col_map.get('actual_start'))
            raw_act_finish = self._get_col_val(row, col_map.get('actual_finish'))
            raw_progress = self._get_col_val(row, col_map.get('progress_percent'))
            raw_status = self._get_col_val(row, col_map.get('status'))
            raw_order = self._get_col_val(row, col_map.get('order'))
            raw_resp = self._get_col_val(row, col_map.get('responsible'))

            calc_notes = []
            planned_start = None

            # Calculate planned_start from base date and numeric offset
            if base_date and raw_start is not None:
                offset_val = self._to_int(raw_start)
                if offset_val is not None:
                    if offset_val < 0:
                        # Negative offset - invalid
                        planned_start = None
                    elif offset_val == 0:
                        planned_start = base_date
                        calc_notes.append(f"Day offset 0 -> {base_date.isoformat()}")
                    else:
                        # Day 1 is base_date, Day 2 is base_date + 1 day
                        planned_start = base_date + timedelta(days=offset_val - 1)
                        calc_notes.append(f"Day offset {offset_val} from {base_date.isoformat()} -> {planned_start.isoformat()}")
                else:
                    # Might already be a date object
                    planned_start = self._to_date(raw_start)
            elif not base_date and raw_start is not None:
                planned_start = self._to_date(raw_start)

            item = self._build_normalized_item(
                source_sheet=meta.get('sheet_name', ''),
                source_row=r_num,
                activity=activity,
                raw_start=planned_start or raw_start,
                raw_finish=raw_finish,
                raw_duration=raw_duration,
                raw_act_start=raw_act_start,
                raw_act_finish=raw_act_finish,
                raw_progress=raw_progress,
                raw_status=raw_status,
                raw_order=raw_order,
                raw_resp=raw_resp,
                default_emp_id=default_emp_id
            )

            if not base_date and raw_start is not None and not self._to_date(raw_start):
                item['errors'].append("Timeline base date could not be proven; numeric offset cannot be converted.")
                item['errors_by_field']['planned_start'] = "Base date missing for offset calculation"
            elif calc_notes:
                item['calc_notes'].extend(calc_notes)

            parsed_rows.append(item)

        return parsed_rows

    def _parse_zone_matrix(self, raw_rows: List[Tuple], meta: Dict, default_emp_id: Optional[int]) -> List[Dict]:
        """
        Parses zone-based matrix where Row 1 has activities and Row 2 has repeated START / END pairs.
        Each data row represents a Zone / Location.
        Normalized activity: f"{Zone} — {Activity}"
        """
        activity_headers = meta.get('activity_headers', [])
        sub_headers = meta.get('sub_headers', [])
        header_idx = meta.get('header_row_idx', 1)
        sub_idx = meta.get('sub_header_idx', 2)

        # Build column pairs: (activity_name, start_col_idx, end_col_idx)
        col_pairs = []
        current_activity = ""
        col = 1
        while col < len(activity_headers) and col < len(sub_headers):
            act = activity_headers[col]
            if act and act.strip():
                current_activity = act.strip()

            sub = sub_headers[col] if col < len(sub_headers) else ''
            sub_next = sub_headers[col + 1] if (col + 1) < len(sub_headers) else ''

            if sub.upper() == 'START' and sub_next.upper() in ('END', 'FINISH'):
                if current_activity:
                    col_pairs.append({
                        'activity': current_activity,
                        'start_col': col,
                        'end_col': col + 1
                    })
                col += 2
            else:
                col += 1

        parsed_rows = []
        order_counter = 1
        for r_num, row in raw_rows[sub_idx:]:
            if not any(c is not None for c in row):
                continue

            zone_name = str(row[0]).strip() if row and row[0] is not None else ''
            if not zone_name:
                continue

            for pair in col_pairs:
                act_name = pair['activity']
                start_val = row[pair['start_col']] if pair['start_col'] < len(row) else None
                end_val = row[pair['end_col']] if pair['end_col'] < len(row) else None

                # Skip completely blank pairs
                if start_val is None and end_val is None:
                    continue

                full_activity_name = f"{zone_name} — {act_name}"
                item = self._build_normalized_item(
                    source_sheet=meta.get('sheet_name', ''),
                    source_row=r_num,
                    activity=full_activity_name,
                    raw_start=start_val,
                    raw_finish=end_val,
                    raw_duration=None,
                    raw_act_start=None,
                    raw_act_finish=None,
                    raw_progress=0,
                    raw_status='Not Started',
                    raw_order=order_counter,
                    raw_resp=None,
                    default_emp_id=default_emp_id
                )
                item['zone'] = zone_name
                item['activity_component'] = act_name

                # Both dates must be resolved and valid for zone matrix import
                if not item['planned_start']:
                    item['errors'].append("Start date unresolved or invalid for zone task.")
                    item['errors_by_field']['planned_start'] = "Start date required"
                if not item['planned_finish']:
                    item['errors'].append("End date unresolved or invalid for zone task.")
                    item['errors_by_field']['planned_finish'] = "End date required"

                parsed_rows.append(item)
                order_counter += 1

        return parsed_rows

    def _parse_visual_monthly(self, raw_rows: List[Tuple], meta: Dict, default_emp_id: Optional[int], sheet_name: str = "") -> List[Dict]:
        """
        Parses visual monthly Gantt charts (e.g. Chillers, Pumps with bi-weekly slots).
        Accurately maps half-month columns to authentic calendar dates from the header row.
        """
        import calendar

        sheet_name = meta.get('sheet_name', '')
        wb = self.get_workbook()
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else None

        # Map header columns to half-month date intervals
        col_dates = {}
        if ws is not None:
            cur_m = None
            for c in range(2, min(ws.max_column + 1, MAX_COLS_PER_SHEET)):
                v = ws.cell(1, c).value
                if v and isinstance(v, (datetime, date)):
                    cur_m = v
                elif v and isinstance(v, str):
                    try:
                        cur_m = datetime.strptime(v.strip()[:10], '%Y-%m-%d')
                    except Exception:
                        pass

                if cur_m:
                    yr = cur_m.year
                    mo = cur_m.month
                    last_day = calendar.monthrange(yr, mo)[1]
                    is_2nd = (c % 2 == 1)
                    if not is_2nd:
                        st = date(yr, mo, 1)
                        en = date(yr, mo, 15)
                    else:
                        st = date(yr, mo, 16)
                        en = date(yr, mo, last_day)
                    col_dates[c] = (st, en)

        parsed_rows = []
        for r_num, row in raw_rows[1:]:
            if not any(c is not None for c in row):
                continue

            activity_val = row[0] if row else None
            if not activity_val or not str(activity_val).strip():
                continue

            activity = str(activity_val).strip()

            # Find filled columns in this row
            filled_cols = []
            if ws is not None:
                for c in range(2, min(ws.max_column + 1, MAX_COLS_PER_SHEET)):
                    cell = ws.cell(r_num, c)
                    is_solid = bool(cell.fill and cell.fill.fill_type == 'solid')
                    if is_solid or cell.value is not None:
                        filled_cols.append(c)

            calc_start = None
            calc_finish = None
            calc_notes = []
            if filled_cols and all(c in col_dates for c in filled_cols):
                calc_start = col_dates[filled_cols[0]][0]
                calc_finish = col_dates[filled_cols[-1]][1]
                calc_notes.append(f"Calculated from half-month timeline columns {filled_cols[0]}..{filled_cols[-1]}")

            item = self._build_normalized_item(
                source_sheet=sheet_name,
                source_row=r_num,
                activity=activity,
                raw_start=calc_start,
                raw_finish=calc_finish,
                raw_duration=((calc_finish - calc_start).days + 1) if (calc_start and calc_finish) else None,
                raw_act_start=None,
                raw_act_finish=None,
                raw_progress=0,
                raw_status='Not Started',
                raw_order=len(parsed_rows) + 1,
                raw_resp=None,
                default_emp_id=default_emp_id
            )

            if not calc_start or not calc_finish:
                item['warnings'].append("Visual monthly chart layout: exact planned dates are ambiguous.")
                item['errors'].append("Planned start and finish dates require manual entry before confirmation.")
                item['errors_by_field']['planned_start'] = "Enter planned start date"
                item['errors_by_field']['planned_finish'] = "Enter planned finish date"
            elif calc_notes:
                item['calc_notes'].extend(calc_notes)

            parsed_rows.append(item)

        return parsed_rows

    # ── Normalization Helpers ────────────────────────────────────────────────

    def _build_normalized_item(
        self,
        source_sheet: str,
        source_row: int,
        activity: str,
        raw_start: Any,
        raw_finish: Any,
        raw_duration: Any,
        raw_act_start: Any,
        raw_act_finish: Any,
        raw_progress: Any,
        raw_status: Any,
        raw_order: Any,
        raw_resp: Any,
        default_emp_id: Optional[int]
    ) -> Dict[str, Any]:
        warnings = []
        errors = []
        errors_by_field = {}
        calc_notes = []

        # Parse Activity
        clean_activity = str(activity).strip()[:255]
        if not clean_activity:
            errors.append("Task activity name cannot be empty.")
            errors_by_field['activity'] = "Activity name required"

        # Parse Order
        order = self._to_int(raw_order)
        if order is None or order < 1:
            order = source_row

        # Parse Duration
        duration_days = self._to_int(raw_duration)
        if duration_days is not None and duration_days <= 0:
            warnings.append(f"Duration ({duration_days} days) is zero or negative; adjusted to 1.")
            duration_days = 1

        # Parse Planned Dates
        p_start = self._to_date(raw_start)
        p_finish = self._to_date(raw_finish)

        # Derive missing finish if duration and start are present
        # Formula: planned_finish = planned_start + duration_days - 1
        if p_start and duration_days and not p_finish:
            p_finish = p_start + timedelta(days=duration_days - 1)
            calc_notes.append(f"Derived finish {p_finish.isoformat()} from start + duration ({duration_days}d - 1)")
        elif p_start and p_finish and not duration_days:
            # Derive duration from start and finish inclusive
            if p_finish >= p_start:
                duration_days = (p_finish - p_start).days + 1
                calc_notes.append(f"Derived duration {duration_days}d from date span")
            else:
                errors.append("Planned finish date cannot be earlier than planned start date.")
                errors_by_field['planned_finish'] = "Finish must be on or after start"
        elif p_start and p_finish and duration_days:
            expected_days = (p_finish - p_start).days + 1
            if p_finish < p_start:
                errors.append("Planned finish date cannot be earlier than planned start date.")
                errors_by_field['planned_finish'] = "Finish must be on or after start"
            elif expected_days != duration_days:
                warnings.append(
                    f"Duration {duration_days}d differs from calendar span ({expected_days}d). Calendar dates take precedence."
                )
                duration_days = expected_days

        # Parse Actual Dates
        a_start = self._to_date(raw_act_start)
        a_finish = self._to_date(raw_act_finish)
        if a_start and a_finish and a_finish < a_start:
            errors.append("Actual finish date cannot be earlier than actual start date.")
            errors_by_field['actual_finish'] = "Actual finish must be on or after actual start"

        # Parse Progress Percent
        progress = self._to_int(raw_progress)
        if progress is None:
            progress = 0
        elif progress < 0 or progress > 100:
            errors.append(f"Progress percent ({progress}%) must be between 0 and 100.")
            errors_by_field['progress_percent'] = "Must be between 0 and 100"

        # Parse Status
        status_str = str(raw_status).strip() if raw_status is not None else ''
        status = self._normalize_status(status_str, progress, p_finish)

        # Responsible Person Mapping
        responsible_id, resp_name = self._resolve_responsible(raw_resp, default_emp_id)

        # Milestone detection (e.g. duration == 0 or 1 day with milestone keyword)
        is_milestone = False
        if duration_days == 1 and any(k in clean_activity.upper() for k in ('MILESTONE', 'AWARD', 'SIGNING', 'KICK OFF', 'COMPLETION')):
            is_milestone = True

        return {
            'source_sheet': source_sheet,
            'source_row': source_row,
            'activity': clean_activity,
            'order': order,
            'planned_start': p_start.isoformat() if p_start else None,
            'planned_finish': p_finish.isoformat() if p_finish else None,
            'duration_days': duration_days,
            'actual_start': a_start.isoformat() if a_start else None,
            'actual_finish': a_finish.isoformat() if a_finish else None,
            'progress_percent': progress,
            'status': status,
            'is_milestone': is_milestone,
            'responsible_id': responsible_id,
            'responsible_name': resp_name,
            'warnings': warnings,
            'errors': errors,
            'errors_by_field': errors_by_field,
            'calc_notes': calc_notes,
            'excluded': False,
            'is_duplicate': False,
            'duplicate_reason': ''
        }

    def _map_columns(self, headers: List[str]) -> Dict[str, int]:
        mapping = {}
        for idx, h in enumerate(headers):
            if not h:
                continue
            h_upper = h.strip().upper()
            for key, aliases in self.HEADER_ALIASES.items():
                if key not in mapping:
                    if h_upper in aliases or any(a == h_upper for a in aliases):
                        mapping[key] = idx
        return mapping

    @staticmethod
    def _get_col_val(row: Tuple, col_idx: Optional[int]) -> Any:
        if col_idx is not None and 0 <= col_idx < len(row):
            return row[col_idx]
        return None

    @staticmethod
    def _to_int(val: Any) -> Optional[int]:
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return int(val)
        try:
            # Handle strings like '100%' or '10.0'
            s = str(val).strip().replace('%', '')
            return int(float(s))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _to_date(val: Any) -> Optional[date]:
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        s = str(val).strip()
        if not s or s.lower() in ('none', 'nan', 'null', '-', '--'):
            return None
        # Clean timestamp strings
        clean_s = s.split(' ')[0]
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%m/%d/%Y'):
            try:
                return datetime.strptime(clean_s, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _normalize_status(raw_status: str, progress: int, planned_finish: Optional[date]) -> str:
        s = raw_status.upper()
        if 'COMPLET' in s or progress == 100:
            return 'Completed'
        if 'DELAY' in s or (planned_finish and planned_finish < date.today() and progress < 100):
            return 'Delayed'
        if 'REVIEW' in s:
            return 'Under Review'
        if 'PROGRESS' in s or 'ONGOING' in s or progress > 0:
            return 'In Progress'
        return 'Not Started'

    @staticmethod
    def _resolve_responsible(raw_resp: Any, default_emp_id: Optional[int]) -> Tuple[Optional[int], str]:
        if default_emp_id:
            emp = EmployeeProfile.objects.filter(pk=default_emp_id, is_active=True).first()
            if emp:
                return emp.pk, emp.full_name

        if raw_resp:
            resp_str = str(raw_resp).strip()
            # Try matching by name or employee ID
            emp = EmployeeProfile.objects.filter(is_active=True).filter(
                full_name__iexact=resp_str
            ).first()
            if emp:
                return emp.pk, emp.full_name
            emp = EmployeeProfile.objects.filter(is_active=True).filter(
                user__email__iexact=resp_str
            ).first()
            if emp:
                return emp.pk, emp.full_name
            return None, resp_str

        return None, "Unassigned"


class GanttDuplicateDetector:
    """Detects probable and exact duplicate tasks against existing project tasks."""

    @staticmethod
    def annotate_duplicates(project: Project, normalized_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        existing_tasks = list(project.tasks.values('id', 'activity', 'planned_start', 'planned_finish', 'order'))
        existing_activity_map = {t['activity'].strip().lower(): t for t in existing_tasks}

        for row in normalized_rows:
            act_clean = row['activity'].strip().lower()
            if act_clean in existing_activity_map:
                existing = existing_activity_map[act_clean]
                row['is_duplicate'] = True
                row['duplicate_reason'] = (
                    f"Matches existing task #{existing['id']} ('{existing['activity']}')"
                )
                row['warnings'].append(f"Probable duplicate: task with same name already exists (#{existing['id']}).")

        return normalized_rows


class GanttImportStagingManager:
    """Manages staged import state with atomic locking, checksum binding, and expiration."""

    EXPIRY_HOURS = 4

    @classmethod
    def create_batch(
        cls,
        project: Project,
        user: Any,
        filename: str,
        file_sha256: str,
        detected_format: str,
        selected_sheet: str,
        staged_rows: List[Dict[str, Any]],
        stats: Dict[str, Any]
    ) -> Any:
        from apps.projects.models import GanttImportBatch
        expires_at = timezone.now() + timedelta(hours=cls.EXPIRY_HOURS)

        batch = GanttImportBatch.objects.create(
            project=project,
            user=user if getattr(user, 'is_authenticated', False) else None,
            filename=filename,
            file_sha256=file_sha256,
            detected_format=detected_format,
            selected_sheet=selected_sheet,
            staged_data={'rows': staged_rows, 'stats': stats},
            task_count=len(staged_rows),
            status='staged',
            expires_at=expires_at
        )
        return batch

    @classmethod
    def get_staged_batch(cls, batch_id: str, project: Project, user: Any) -> Any:
        from apps.projects.models import GanttImportBatch
        batch = GanttImportBatch.objects.filter(
            uuid=batch_id,
            project=project
        ).first()

        if not batch:
            raise GanttImportError("Staged import batch not found or project mismatch.", code="batch_not_found")

        # Authorization / Tenant scope check
        if batch.user_id and user.is_authenticated and batch.user_id != user.id and not user.is_superuser:
            raise PermissionDenied("You do not have access to this staged import batch.")

        if batch.is_expired:
            raise GanttImportError("This staged import batch has expired. Please re-upload.", code="batch_expired")

        return batch

    @classmethod
    def update_staged_row(cls, batch: Any, row_idx: int, field: str, value: Any) -> Dict[str, Any]:
        """Applies an in-place user correction to a staged preview row and revalidates."""
        staged_data = batch.staged_data or {}
        rows = staged_data.get('rows', [])

        if not (0 <= row_idx < len(rows)):
            raise GanttImportError(f"Row index {row_idx} is out of bounds.", code="invalid_row_index")

        row = rows[row_idx]

        # Update specified field
        if field == 'activity':
            row['activity'] = str(value).strip()[:255]
        elif field == 'planned_start':
            d = GanttWorkbookParser._to_date(value)
            row['planned_start'] = d.isoformat() if d else None
        elif field == 'planned_finish':
            d = GanttWorkbookParser._to_date(value)
            row['planned_finish'] = d.isoformat() if d else None
        elif field == 'duration_days':
            dur = GanttWorkbookParser._to_int(value)
            row['duration_days'] = dur
        elif field == 'progress_percent':
            p = GanttWorkbookParser._to_int(value)
            row['progress_percent'] = max(0, min(100, p if p is not None else 0))
        elif field == 'status':
            row['status'] = value
        elif field == 'responsible_id':
            emp_id = GanttWorkbookParser._to_int(value)
            emp = EmployeeProfile.objects.filter(pk=emp_id, is_active=True).first()
            if emp:
                row['responsible_id'] = emp.pk
                row['responsible_name'] = emp.full_name
            else:
                row['responsible_id'] = None
                row['responsible_name'] = "Unassigned"
        elif field == 'excluded':
            row['excluded'] = bool(value)

        # Re-validate row
        cls._revalidate_row(row)

        # Recalculate stats
        total = len(rows)
        valid = sum(1 for r in rows if not r['errors'] and not r['warnings'])
        warning = sum(1 for r in rows if r['warnings'] and not r['errors'])
        invalid = sum(1 for r in rows if r['errors'])
        importable = sum(1 for r in rows if not r['errors'] and not r.get('excluded', False))

        batch.staged_data['rows'] = rows
        batch.staged_data['stats'] = {
            'total': total,
            'valid': valid,
            'warning': warning,
            'invalid': invalid,
            'importable': importable
        }
        batch.save(update_fields=['staged_data', 'updated_at'])
        return row

    @classmethod
    def _revalidate_row(cls, row: Dict[str, Any]):
        errors = []
        errors_by_field = {}
        warnings = [w for w in row.get('warnings', []) if 'duplicate' in w.lower()]

        act = row.get('activity', '').strip()
        if not act:
            errors.append("Task activity name cannot be empty.")
            errors_by_field['activity'] = "Activity name required"

        p_start = GanttWorkbookParser._to_date(row.get('planned_start'))
        p_finish = GanttWorkbookParser._to_date(row.get('planned_finish'))
        dur = row.get('duration_days')

        if not p_start:
            errors.append("Planned start date is required.")
            errors_by_field['planned_start'] = "Start date required"

        if not p_finish:
            errors.append("Planned finish date is required.")
            errors_by_field['planned_finish'] = "Finish date required"

        if p_start and p_finish:
            if p_finish < p_start:
                errors.append("Planned finish date cannot be earlier than planned start date.")
                errors_by_field['planned_finish'] = "Finish must be on or after start"
            else:
                expected_days = (p_finish - p_start).days + 1
                row['duration_days'] = expected_days

        p_prog = row.get('progress_percent', 0)
        if p_prog is not None and (p_prog < 0 or p_prog > 100):
            errors.append("Progress percent must be between 0 and 100.")
            errors_by_field['progress_percent'] = "Must be between 0 and 100"

        row['errors'] = errors
        row['errors_by_field'] = errors_by_field
        row['warnings'] = warnings


class GanttImportExecutor:
    """Executes validated staged batch import atomically with audit tracking and idempotency."""

    @classmethod
    def confirm_import(
        cls,
        batch: Any,
        project: Project,
        user: Any,
        request: Any = None
    ) -> Dict[str, Any]:
        """
        Atomically commits the import batch:
        - Locks batch record to prevent duplicate concurrent writes
        - Returns existing result if already completed (idempotency)
        - Calculates non-colliding order numbers
        - Creates ProjectTask records
        - Recalculates project progress
        - Emits a single audit log event
        - Rolls back completely if any row fails
        """
        from apps.projects.models import GanttImportBatch
        from django.db.models import Max

        # Check if already completed (Idempotency check before lock)
        if batch.status == 'completed':
            return {
                'status': 'already_completed',
                'imported_count': len(batch.imported_task_ids),
                'batch_id': str(batch.uuid),
                'task_ids': batch.imported_task_ids
            }

        with transaction.atomic():
            # Lock the batch record for update
            locked_batch = GanttImportBatch.objects.select_for_update().get(pk=batch.pk)
            if locked_batch.status == 'completed':
                return {
                    'status': 'already_completed',
                    'imported_count': len(locked_batch.imported_task_ids),
                    'batch_id': str(locked_batch.uuid),
                    'task_ids': locked_batch.imported_task_ids
                }

            staged_data = locked_batch.staged_data or {}
            rows = staged_data.get('rows', [])

            # Filter rows to import (exclude excluded and error rows)
            importable_rows = [r for r in rows if not r.get('excluded', False) and not r.get('errors')]

            if not importable_rows:
                raise GanttImportError(
                    "No valid, non-excluded rows available to import.",
                    code="no_importable_rows"
                )

            # Determine starting order: max(existing_order) + 1
            max_order = project.tasks.aggregate(m=Max('order'))['m'] or 0
            current_order = max_order + 1

            created_tasks = []
            created_task_ids = []

            for r_idx, r in enumerate(importable_rows):
                p_start = GanttWorkbookParser._to_date(r.get('planned_start'))
                p_finish = GanttWorkbookParser._to_date(r.get('planned_finish'))
                a_start = GanttWorkbookParser._to_date(r.get('actual_start'))
                a_finish = GanttWorkbookParser._to_date(r.get('actual_finish'))
                duration = r.get('duration_days') or 1
                progress = r.get('progress_percent') or 0
                status = r.get('status') or 'Not Started'
                resp_id = r.get('responsible_id')

                # Verification on authoritative server data
                if not r.get('activity') or not p_start or not p_finish or p_finish < p_start:
                    raise GanttImportError(
                        f"Row {r.get('source_row', r_idx+1)} failed validation during confirmation: "
                        f"{r.get('activity', 'Unnamed task')}",
                        code="row_validation_failed"
                    )

                remarks_note = (
                    f"Imported from {locked_batch.filename} "
                    f"[Batch: {str(locked_batch.uuid)[:8]} | Sheet: {locked_batch.selected_sheet} | Row: {r.get('source_row')}]"
                )

                task = ProjectTask(
                    project=project,
                    order=current_order,
                    activity=r['activity'][:255],
                    responsible_person_id=resp_id,
                    planned_start=p_start,
                    planned_finish=p_finish,
                    duration_days=duration,
                    baseline_start=p_start,
                    baseline_finish=p_finish,
                    actual_start=a_start,
                    actual_finish=a_finish,
                    is_milestone=bool(r.get('is_milestone', False)),
                    status=status,
                    progress_percent=progress,
                    remarks=remarks_note,
                    points=10
                )
                task.save()
                created_tasks.append(task)
                created_task_ids.append(task.pk)
                current_order += 1

            # Recalculate project progress
            project.recalculate_progress()

            # Mark batch as completed
            locked_batch.status = 'completed'
            locked_batch.imported_task_ids = created_task_ids
            locked_batch.completed_at = timezone.now()
            locked_batch.save(update_fields=['status', 'imported_task_ids', 'completed_at', 'updated_at'])

            # Log single audit event
            AuditService.log_event(
                actor=user,
                action='gantt_imported',
                instance=project,
                module='projects',
                object_type='ProjectTaskImport',
                object_id=str(locked_batch.uuid),
                object_label=f"Imported {len(created_tasks)} tasks into {project.name}",
                request=request,
                reason=f"Excel import from file '{locked_batch.filename}' ({len(created_tasks)} tasks created)."
            )

        return {
            'status': 'success',
            'imported_count': len(created_tasks),
            'batch_id': str(locked_batch.uuid),
            'task_ids': created_task_ids
        }


def check_gantt_import_permission(user: Any, project: Project) -> bool:
    """
    Authoritative permission check for Gantt import operations.
    Enforces project-scoped access and task assignment/creation permissions.
    """
    if not user or not user.is_authenticated:
        return False

    # Superuser has global access
    if user.is_superuser:
        return True

    # User role check
    user_role = getattr(user, 'role', '')
    if user_role in ('admin', 'system_owner'):
        return True

    # Check permission engine
    from apps.accounts.engine import PermissionEngine
    edit_perm = PermissionEngine.evaluate(user, 'projects.edit').allowed
    assign_perm = user.has_perm('projects.assign_projecttask') or user.has_perm('projects.add_projecttask')

    # Project relationship
    emp = getattr(user, 'employee_profile', None)
    is_pm = emp and project.project_managers.filter(id=emp.id).exists()
    is_site_eng = emp and project.site_engineers.filter(id=emp.id).exists()
    is_member = emp and project.project_members.filter(id=emp.id).exists()

    # Rule: Project members without assignment permission must receive 403
    if is_member and not (is_pm or is_site_eng or assign_perm or edit_perm):
        return False

    if is_pm or is_site_eng:
        return True

    if (assign_perm or edit_perm) and (is_member or project.created_by == user):
        return True

    return False
