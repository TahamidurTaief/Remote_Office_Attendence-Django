"""
Gantt Excel Export Service

Generates pixel-perfect, professionally formatted Excel workbooks (.xlsx)
matching the exact layout and structure of the reference project schedule
and planner spreadsheets:
- Sheet 1: 'Project Planner' (Daily tracking matrix, Period Highlight, Plan vs Actual)
- Sheet 2: 'Schedule' (Multi-Zone / Stage Work Breakdown Schedule)
- Sheet 3: 'Monthly Summary' (High-level monthly milestone Gantt overview)
"""

import io
from datetime import date, datetime, timedelta
from typing import Optional, List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.projects.models import Project, ProjectTask


class GanttExcelExportService:
    """Generates dynamic Excel Gantt Chart workbooks for projects."""

    # Color Palette matching reference workbooks
    COLOR_PRIMARY_NAVY = "1F497D"      # Dark navy headers
    COLOR_PLAN_DURATION = "B8CCE4"     # Light steel blue (planned bar)
    COLOR_PERCENT_COMPLETE = "366092"  # Deep blue (% complete bar)
    COLOR_ACTUAL_START = "548DD4"      # Medium blue
    COLOR_ACTUAL_BEYOND = "C00000"     # Red / beyond plan
    COLOR_HEADER_BG = "DCE6F1"         # Subtle blue accent header
    COLOR_BORDER_LIGHT = "D9D9D9"      # Grid line border
    COLOR_BORDER_DARK = "808080"       # Medium border
    COLOR_HIGHLIGHT_BG = "FFF2CC"      # Period highlight soft yellow

    @classmethod
    def export_project_workbook(cls, project: Project) -> bytes:
        """
        Creates a complete multi-sheet Gantt workbook for the project and returns
        the raw .xlsx bytes.
        """
        wb = openpyxl.Workbook()

        # Sheet 1: Project Planner
        ws_planner = wb.active
        ws_planner.title = "Project Planner"
        cls._build_project_planner_sheet(ws_planner, project)

        # Sheet 2: Multi-Zone / Activity Schedule
        ws_schedule = wb.create_sheet(title="Schedule")
        cls._build_zone_schedule_sheet(ws_schedule, project)

        # Sheet 3: Monthly Overview
        ws_monthly = wb.create_sheet(title="Monthly Overview")
        cls._build_monthly_summary_sheet(ws_monthly, project)

        # Output to bytes
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()

    @classmethod
    def _build_project_planner_sheet(cls, ws, project: Project):
        """Builds the exact Project Planner sheet matching the reference Excel."""
        tasks = list(project.tasks.select_related('responsible_person').order_by('order'))
        today = date.today()

        # Determine date range
        all_dates = []
        for t in tasks:
            if t.planned_start:
                all_dates.append(t.planned_start)
            if t.planned_finish:
                all_dates.append(t.planned_finish)
            if t.actual_start:
                all_dates.append(t.actual_start)
            if t.actual_finish:
                all_dates.append(t.actual_finish)

        if all_dates:
            chart_start = min(all_dates)
            chart_end = max(all_dates)
        else:
            chart_start = project.start_date or today
            chart_end = project.completion_date or (chart_start + timedelta(days=30))

        if chart_end < chart_start:
            chart_end = chart_start + timedelta(days=1)

        # Ensure reasonable day span (minimum 14 days, max 90 days in daily view)
        span_days = (chart_end - chart_start).days + 1
        display_days = max(span_days, 14)
        if display_days > 90:
            display_days = 90  # Cap daily view width for readability

        day_dates = [chart_start + timedelta(days=i) for i in range(display_days)]

        # --- Fonts & Styles ---
        font_title = Font(name="Calibri", size=14, bold=True, color=cls.COLOR_PRIMARY_NAVY)
        font_subtitle = Font(name="Calibri", size=9, italic=True, color="595959")
        font_legend_label = Font(name="Calibri", size=9, bold=True)
        font_header = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=9)
        font_data_bold = Font(name="Calibri", size=9, bold=True)
        font_date_header = Font(name="Calibri", size=8, color="595959")
        font_day_num = Font(name="Calibri", size=9, bold=True, color=cls.COLOR_PRIMARY_NAVY)

        fill_header = PatternFill(start_color=cls.COLOR_PRIMARY_NAVY, end_color=cls.COLOR_PRIMARY_NAVY, fill_type="solid")
        fill_plan = PatternFill(start_color=cls.COLOR_PLAN_DURATION, end_color=cls.COLOR_PLAN_DURATION, fill_type="solid")
        fill_progress = PatternFill(start_color=cls.COLOR_PERCENT_COMPLETE, end_color=cls.COLOR_PERCENT_COMPLETE, fill_type="solid")
        fill_actual = PatternFill(start_color=cls.COLOR_ACTUAL_START, end_color=cls.COLOR_ACTUAL_START, fill_type="solid")
        fill_beyond = PatternFill(start_color=cls.COLOR_ACTUAL_BEYOND, end_color=cls.COLOR_ACTUAL_BEYOND, fill_type="solid")
        fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        border_thin = Border(
            left=Side(style="thin", color=cls.COLOR_BORDER_LIGHT),
            right=Side(style="thin", color=cls.COLOR_BORDER_LIGHT),
            top=Side(style="thin", color=cls.COLOR_BORDER_LIGHT),
            bottom=Side(style="thin", color=cls.COLOR_BORDER_LIGHT),
        )
        border_box = Border(
            left=Side(style="thin", color=cls.COLOR_BORDER_DARK),
            right=Side(style="thin", color=cls.COLOR_BORDER_DARK),
            top=Side(style="thin", color=cls.COLOR_BORDER_DARK),
            bottom=Side(style="thin", color=cls.COLOR_BORDER_DARK),
        )

        # Row 1: Title
        system_type = f" ({project.system_type})" if project.system_type else ""
        ws["B1"] = f"{project.name.upper()} HVAC WORK{system_type}"
        ws["B1"].font = font_title

        # Row 2: Subtitle & Legend
        ws["B2"] = "Select a period to highlight at right. A legend describing the charting follows."
        ws["B2"].font = font_subtitle

        ws["G2"] = "Period Highlight:"
        ws["G2"].font = font_legend_label
        ws["G2"].alignment = align_right

        ws["I2"] = 1
        ws["I2"].font = font_legend_label
        ws["I2"].alignment = align_center
        ws["I2"].border = border_box

        # Legend boxes
        ws["L2"] = "Plan Duration"
        ws["L2"].font = font_legend_label
        ws["L2"].fill = fill_plan
        ws["L2"].alignment = align_center
        ws["L2"].border = border_box

        ws["R2"] = "Actual Start"
        ws["R2"].font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        ws["R2"].fill = fill_actual
        ws["R2"].alignment = align_center
        ws["R2"].border = border_box

        ws["W2"] = "% Complete"
        ws["W2"].font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        ws["W2"].fill = fill_progress
        ws["W2"].alignment = align_center
        ws["W2"].border = border_box

        ws["AB2"] = "Actual (beyond plan)"
        ws["AB2"].font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        ws["AB2"].fill = fill_beyond
        ws["AB2"].alignment = align_center
        ws["AB2"].border = border_box

        # Row 3: Table Headers
        headers = [
            ("B3", "ACTIVITY", 30),
            ("C3", "PLAN START", 13),
            ("D3", "PLAN DURATION", 14),
            ("E3", "PLAN END", 13),
            ("F3", "ACTUAL START", 13),
            ("G3", "ACTUAL DURATION", 15),
            ("H3", "PERCENT COMPLETE", 16),
        ]
        for coord, label, width in headers:
            cell = ws[coord]
            cell.value = label
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            col_letter = coord[0]
            ws.column_dimensions[col_letter].width = width

        # Row 4, 5, 6: Timeline Headers (starting at column I / col 9)
        start_col = 9
        for idx, dt in enumerate(day_dates):
            col_idx = start_col + idx
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 4.2

            # Row 4: Day number (1, 2, 3...)
            c4 = ws.cell(row=4, column=col_idx, value=idx + 1)
            c4.font = font_day_num
            c4.alignment = align_center
            c4.border = border_thin

            # Row 5: Day of week abbreviation
            c5 = ws.cell(row=5, column=col_idx, value=dt.strftime("%a").upper())
            c5.font = font_date_header
            c5.alignment = align_center
            c5.border = border_thin

            # Row 6: Date (e.g. 27 Aug)
            c6 = ws.cell(row=6, column=col_idx, value=dt.strftime("%d-%b"))
            c6.font = font_date_header
            c6.alignment = align_center
            c6.border = border_thin

        # Rows 7+: Task Data Rows
        current_row = 7
        for t in tasks:
            is_zebra = (current_row % 2 == 0)
            row_fill = fill_zebra if is_zebra else None

            # B: Activity Name
            cB = ws.cell(row=current_row, column=2, value=t.activity)
            cB.font = font_data_bold if t.is_milestone else font_data
            cB.alignment = align_left
            cB.border = border_thin
            if row_fill:
                cB.fill = row_fill

            # C: Planned Start
            cC = ws.cell(row=current_row, column=3, value=t.planned_start.strftime("%Y-%m-%d") if t.planned_start else None)
            cC.font = font_data
            cC.alignment = align_center
            cC.border = border_thin
            if row_fill:
                cC.fill = row_fill

            # D: Planned Duration
            dur = t.duration_days
            if not dur and t.planned_start and t.planned_finish:
                dur = max(1, (t.planned_finish - t.planned_start).days + 1)
            cD = ws.cell(row=current_row, column=4, value=dur or 1)
            cD.font = font_data
            cD.alignment = align_center
            cD.border = border_thin
            if row_fill:
                cD.fill = row_fill

            # E: Planned End (Excel formula =C{row}+D{row}-1 if dates exist)
            if t.planned_start:
                cE = ws.cell(row=current_row, column=5, value=f"=C{current_row}+D{current_row}-1")
            elif t.planned_finish:
                cE = ws.cell(row=current_row, column=5, value=t.planned_finish.strftime("%Y-%m-%d"))
            else:
                cE = ws.cell(row=current_row, column=5, value=None)
            cE.font = font_data
            cE.alignment = align_center
            cE.border = border_thin
            if row_fill:
                cE.fill = row_fill

            # F: Actual Start
            cF = ws.cell(row=current_row, column=6, value=t.actual_start.strftime("%Y-%m-%d") if t.actual_start else None)
            cF.font = font_data
            cF.alignment = align_center
            cF.border = border_thin
            if row_fill:
                cF.fill = row_fill

            # G: Actual Duration
            act_dur = None
            if t.actual_start and t.actual_finish:
                act_dur = max(1, (t.actual_finish - t.actual_start).days + 1)
            elif t.actual_start:
                act_dur = max(1, (today - t.actual_start).days + 1)
            cG = ws.cell(row=current_row, column=7, value=act_dur)
            cG.font = font_data
            cG.alignment = align_center
            cG.border = border_thin
            if row_fill:
                cG.fill = row_fill

            # H: Percent Complete
            cH = ws.cell(row=current_row, column=8, value=f"{t.progress_percent}%")
            cH.font = font_data_bold if t.progress_percent > 0 else font_data
            cH.alignment = align_center
            cH.border = border_thin
            if row_fill:
                cH.fill = row_fill

            # Timeline daily cells (columns I onwards)
            p_start = t.planned_start
            p_dur = dur or 1
            p_finish = t.planned_finish or (p_start + timedelta(days=p_dur - 1) if p_start else None)
            pct = t.progress_percent or 0
            completed_days = round((pct / 100.0) * p_dur) if p_dur else 0

            for idx, dt in enumerate(day_dates):
                col_idx = start_col + idx
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = border_thin

                # Determine cell shading
                in_plan = (p_start and p_finish and p_start <= dt <= p_finish)
                in_progress = in_plan and (p_start and (dt - p_start).days < completed_days)
                in_actual_beyond = (t.actual_finish and p_finish and p_finish < dt <= t.actual_finish)

                if in_progress:
                    cell.fill = fill_progress
                elif in_plan:
                    cell.fill = fill_plan
                elif in_actual_beyond:
                    cell.fill = fill_beyond
                elif row_fill:
                    cell.fill = row_fill

            current_row += 1

        # Set row heights
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 26
        ws.row_dimensions[4].height = 16
        ws.row_dimensions[5].height = 15
        ws.row_dimensions[6].height = 15
        for r in range(7, current_row):
            ws.row_dimensions[r].height = 20

    @classmethod
    def _build_zone_schedule_sheet(cls, ws, project: Project):
        """
        Builds the Work / Zone Matrix schedule sheet matching
        'PROJECT SCHEDULE GANTT CHART.xlsx' Schedule (2) / DATE.
        """
        tasks = list(project.tasks.all().order_by('order'))

        font_header_zone = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        font_sub = Font(name="Calibri", size=8, bold=True, color="595959")
        font_data = Font(name="Calibri", size=9)
        font_bold = Font(name="Calibri", size=9, bold=True)

        fill_header = PatternFill(start_color=cls.COLOR_PRIMARY_NAVY, end_color=cls.COLOR_PRIMARY_NAVY, fill_type="solid")
        fill_sub = PatternFill(start_color=cls.COLOR_HEADER_BG, end_color=cls.COLOR_HEADER_BG, fill_type="solid")
        border_thin = Border(
            left=Side(style="thin", color=cls.COLOR_BORDER_LIGHT),
            right=Side(style="thin", color=cls.COLOR_BORDER_LIGHT),
            top=Side(style="thin", color=cls.COLOR_BORDER_LIGHT),
            bottom=Side(style="thin", color=cls.COLOR_BORDER_LIGHT),
        )

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        stages = [
            "CEILING/ ACP OPENING",
            "MARKING",
            "SUPPORT WORK",
            "COPPER PIPING",
            "PRESSURE TESTING-1",
            "DRAIN PIPE",
            "NETWORKING",
            "AIR DUCT",
            "INDOOR INSTALLATION",
            "OUTDOOR INSTALLATION",
            "COMMISSIONING",
        ]

        # Row 1: Zone header + Merged Stage Headers
        ws["A1"] = "WORK / ZONE"
        ws["A1"].font = font_header_zone
        ws["A1"].fill = fill_header
        ws["A1"].alignment = align_center
        ws.column_dimensions["A"].width = 30

        col_ptr = 2
        for stage in stages:
            start_letter = get_column_letter(col_ptr)
            end_letter = get_column_letter(col_ptr + 1)
            ws.merge_cells(f"{start_letter}1:{end_letter}1")

            c = ws[f"{start_letter}1"]
            c.value = stage
            c.font = font_header_zone
            c.fill = fill_header
            c.alignment = align_center

            # Row 2: START / END sub-headers
            c_start = ws.cell(row=2, column=col_ptr, value="START")
            c_start.font = font_sub
            c_start.fill = fill_sub
            c_start.alignment = align_center
            c_start.border = border_thin
            ws.column_dimensions[start_letter].width = 11

            c_end = ws.cell(row=2, column=col_ptr + 1, value="END")
            c_end.font = font_sub
            c_end.fill = fill_sub
            c_end.alignment = align_center
            c_end.border = border_thin
            ws.column_dimensions[end_letter].width = 11

            col_ptr += 2

        # Standard HVAC project zones / departments
        standard_zones = [
            "KNITTING & MAINT WORKSHOP",
            "CUTTING & MAINTENANCE WORKSHOP",
            "ENGINEERING WORKSHOP GF",
            "PROCUREMENT & HR SECTION GF",
            "CANTEEN & SUBSTATION",
            "PROCUREMENT & HR SECTION MF",
            "SEWING CANTEEN & OFFICE MZ",
            "COLOR SERVICE ZONE",
            "SPARE PARTS WAREHOUSE",
            "CANTEEN & OFFICE",
            "SEWING FLOOR",
        ]

        curr_row = 3
        for zone in standard_zones:
            c_zone = ws.cell(row=curr_row, column=1, value=zone)
            c_zone.font = font_bold
            c_zone.alignment = align_left
            c_zone.border = border_thin

            # Look up matching tasks for this zone if tasks follow "Zone - Stage" naming
            for s_idx, stage in enumerate(stages):
                matched = [t for t in tasks if stage.lower() in t.activity.lower() and (zone.lower() in t.activity.lower() or len(standard_zones) <= len(tasks))]
                t = matched[0] if matched else None

                scol = 2 + (s_idx * 2)
                ecol = scol + 1

                c_s = ws.cell(row=curr_row, column=scol)
                c_s.border = border_thin
                c_s.alignment = align_center
                c_s.font = font_data

                c_e = ws.cell(row=curr_row, column=ecol)
                c_e.border = border_thin
                c_e.alignment = align_center
                c_e.font = font_data

                if t and t.planned_start:
                    c_s.value = t.planned_start.strftime("%Y-%m-%d")
                if t and t.planned_finish:
                    c_e.value = t.planned_finish.strftime("%Y-%m-%d")

            curr_row += 1

        # Blank row separator
        curr_row += 1

        # Project Milestones section
        ws.cell(row=curr_row, column=1, value="KEY PROJECT MILESTONES").font = Font(name="Calibri", size=10, bold=True, color=cls.COLOR_PRIMARY_NAVY)
        curr_row += 1

        milestones = [
            ("Site Assessment", project.start_date),
            ("Technical Discussion", project.start_date + timedelta(days=3) if project.start_date else None),
            ("Drawing Approval", project.start_date + timedelta(days=5) if project.start_date else None),
            ("Local Material Requisition", project.start_date + timedelta(days=7) if project.start_date else None),
            ("Local Material Delivery", project.start_date + timedelta(days=12) if project.start_date else None),
            ("Ready Stock Machine Delivery", project.start_date + timedelta(days=20) if project.start_date else None),
            ("Site Mobilization", project.start_date + timedelta(days=10) if project.start_date else None),
            ("Induction", project.start_date + timedelta(days=11) if project.start_date else None),
            ("Installation Start", project.start_date + timedelta(days=13) if project.start_date else None),
            ("Commissioning & Handover", project.completion_date),
        ]

        for name, mdate in milestones:
            cm = ws.cell(row=curr_row, column=1, value=name)
            cm.font = font_data
            cm.alignment = align_left
            cm.border = border_thin

            cd = ws.cell(row=curr_row, column=2, value=mdate.strftime("%Y-%m-%d") if mdate else "TBD")
            cd.font = font_data
            cd.alignment = align_center
            cd.border = border_thin
            curr_row += 1

    @classmethod
    def _build_monthly_summary_sheet(cls, ws, project: Project):
        """
        Builds the Monthly Summary sheet matching 'GANTT CHART (2).xlsx' Schedule.
        """
        tasks = list(project.tasks.all().order_by('order'))
        today = date.today()

        font_header = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=9)
        fill_header = PatternFill(start_color=cls.COLOR_PRIMARY_NAVY, end_color=cls.COLOR_PRIMARY_NAVY, fill_type="solid")
        fill_bar = PatternFill(start_color=cls.COLOR_PERCENT_COMPLETE, end_color=cls.COLOR_PERCENT_COMPLETE, fill_type="solid")
        border_thin = Border(
            left=Side(style="thin", color=cls.COLOR_BORDER_LIGHT),
            right=Side(style="thin", color=cls.COLOR_BORDER_LIGHT),
            top=Side(style="thin", color=cls.COLOR_BORDER_LIGHT),
            bottom=Side(style="thin", color=cls.COLOR_BORDER_LIGHT),
        )

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        ws["A1"] = "DESCRIPTION"
        ws["A1"].font = font_header
        ws["A1"].fill = fill_header
        ws["A1"].alignment = align_left
        ws.column_dimensions["A"].width = 32

        # Months from project start to completion (or next 6 months)
        start_d = project.start_date or today
        end_d = project.completion_date or (start_d + timedelta(days=180))

        months = []
        cur = start_d.replace(day=1)
        while cur <= end_d:
            months.append(cur)
            next_m = (cur.replace(day=28) + timedelta(days=4)).replace(day=1)
            cur = next_m

        # Each month gets 2 columns (1st half, 2nd half)
        col_ptr = 2
        month_col_map = {}
        for m in months:
            m_label = m.strftime("%b %Y")
            c1_letter = get_column_letter(col_ptr)
            c2_letter = get_column_letter(col_ptr + 1)
            ws.merge_cells(f"{c1_letter}1:{c2_letter}1")

            c = ws[f"{c1_letter}1"]
            c.value = m_label
            c.font = font_header
            c.fill = fill_header
            c.alignment = align_center

            ws.column_dimensions[c1_letter].width = 6
            ws.column_dimensions[c2_letter].width = 6

            month_col_map[m.strftime("%Y-%m")] = (col_ptr, col_ptr + 1)
            col_ptr += 2

        # Task rows
        curr_row = 2
        for t in tasks:
            cA = ws.cell(row=curr_row, column=1, value=t.activity)
            cA.font = font_data
            cA.alignment = align_left
            cA.border = border_thin

            # Shade columns that overlap with task planned dates
            for c in range(2, col_ptr):
                cell = ws.cell(row=curr_row, column=c)
                cell.border = border_thin

            if t.planned_start and t.planned_finish:
                p_cur = t.planned_start.replace(day=1)
                while p_cur <= t.planned_finish:
                    k = p_cur.strftime("%Y-%m")
                    if k in month_col_map:
                        c1, c2 = month_col_map[k]
                        ws.cell(row=curr_row, column=c1).fill = fill_bar
                        ws.cell(row=curr_row, column=c2).fill = fill_bar
                    p_cur = (p_cur.replace(day=28) + timedelta(days=4)).replace(day=1)

            curr_row += 1
