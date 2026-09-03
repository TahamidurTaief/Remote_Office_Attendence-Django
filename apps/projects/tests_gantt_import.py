import io
import zipfile
from datetime import date, timedelta
import openpyxl

from django.test import TestCase, Client
from django.contrib.auth import get_user_model
from django.urls import reverse
from django.core.exceptions import PermissionDenied

from apps.branches.models import Branch
from apps.employees.models import EmployeeProfile
from apps.projects.models import Project, ProjectType, ProjectTask, GanttImportBatch
from apps.projects.services.gantt_import import (
    WorkbookSafetyValidator,
    GanttWorkbookParser,
    GanttFormatDetector,
    GanttDuplicateDetector,
    GanttImportStagingManager,
    GanttImportExecutor,
    GanttImportError,
    check_gantt_import_permission
)
from apps.audit.models import AuditEvent

User = get_user_model()


def build_structured_workbook() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.append(["ACTIVITY", "PLAN START", "PLAN DURATION", "PLAN END", "PERCENT COMPLETE"])
    ws.append(["Kickoff Meeting", date(2024, 6, 1), 1, date(2024, 6, 1), 100])
    ws.append(["Site Assessment", date(2024, 6, 5), 3, date(2024, 6, 7), 50])
    ws.append(["Material Delivery", date(2024, 6, 10), 2, date(2024, 6, 11), 0])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_offset_workbook() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Offset Planner"
    ws.append(["Project Timeline", "", "", ""])
    ws.append(["", "", "START", date(2024, 6, 1)])
    ws.append(["ACTIVITY", "PLAN START", "PLAN DURATION", "PERCENT COMPLETE"])
    ws.append(["PO Receive", 1, 1, 100])
    ws.append(["OEM Meeting", 4, 1, 50])
    ws.append(["Site Assessment", 10, 2, 0])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_zone_matrix_workbook() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zone Schedule"
    ws.append(["WORK / ZONE", "COPPER PIPING", "", "DRAIN PIPE", ""])
    ws.append(["", "START", "END", "START", "END"])
    ws.append(["Zone A", date(2024, 9, 1), date(2024, 9, 3), date(2024, 9, 4), date(2024, 9, 6)])
    ws.append(["Zone B", date(2024, 9, 5), date(2024, 9, 8), None, None])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_visual_monthly_workbook() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visual Timeline"
    ws.append(["DESCRIPTION", date(2024, 5, 1), date(2024, 6, 1), date(2024, 7, 1)])
    ws.append(["Chillers", "", "", ""])
    ws.append(["Pumps", "", "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_macro_file() -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w') as z:
        z.writestr('[Content_Types].xml', '<Types/>')
        z.writestr('xl/vbaProject.bin', b'MOCK_VBA_PAYLOAD')
    return buf.getvalue()


class GanttImportServiceTests(TestCase):
    """Unit tests for the underlying Gantt import services."""
    def setUp(self):
        self.branch = Branch.objects.create(
            name="Central",
            address="Dhaka",
            latitude=23.8103,
            longitude=90.4125,
            radius_meters=100
        )
        self.project_type, _ = ProjectType.objects.get_or_create(name="HVAC Commercial")
        self.project = Project.objects.create(
            name="Alpha VRF Facility",
            client_name="Alpha Corp",
            project_type=self.project_type,
            start_date=date(2024, 6, 1),
            branch=self.branch
        )
        self.admin_user = User.objects.create_user(
            email="admin_gantt@example.com",
            phone="+8801700999001",
            password="testpassword123",
            role="admin"
        )
        self.emp_profile = EmployeeProfile.objects.create(
            user=self.admin_user,
            full_name="Chief Engineer",
            branch=self.branch,
            employee_id="EMP_G1",
            phone="+8801700999001",
            joined_date=date.today(),
            is_active=True
        )

    def test_safety_validator_rejects_corrupt_file(self):
        with self.assertRaises(GanttImportError) as ctx:
            WorkbookSafetyValidator.validate_file(b"NOT_A_VALID_ZIP_ARCHIVE", "corrupt.xlsx")
        self.assertEqual(ctx.exception.code, "corrupt_archive")

    def test_safety_validator_rejects_macro_file(self):
        macro_bytes = build_macro_file()
        with self.assertRaises(GanttImportError) as ctx:
            WorkbookSafetyValidator.validate_file(macro_bytes, "schedule.xlsx")
        self.assertEqual(ctx.exception.code, "macro_prohibited")

    def test_safety_validator_rejects_bad_extension(self):
        with self.assertRaises(GanttImportError) as ctx:
            WorkbookSafetyValidator.validate_file(b"dummy", "schedule.xls")
        self.assertEqual(ctx.exception.code, "invalid_extension")

    def test_safety_validator_accepts_valid_xlsx(self):
        content = build_structured_workbook()
        bytes_out, sha256_hash = WorkbookSafetyValidator.validate_file(content, "schedule.xlsx")
        self.assertEqual(bytes_out, content)
        self.assertEqual(len(sha256_hash), 64)

    def test_parse_structured_table(self):
        content = build_structured_workbook()
        parser = GanttWorkbookParser(content, "schedule.xlsx")
        res = parser.parse_sheet("Schedule")
        parser.close()

        self.assertEqual(res['detected_format'], GanttFormatDetector.FORMAT_STRUCTURED)
        self.assertEqual(len(res['rows']), 3)

        row1 = res['rows'][0]
        self.assertEqual(row1['activity'], "Kickoff Meeting")
        self.assertEqual(row1['planned_start'], "2024-06-01")
        self.assertEqual(row1['planned_finish'], "2024-06-01")
        self.assertEqual(row1['duration_days'], 1)
        self.assertEqual(row1['progress_percent'], 100)
        self.assertEqual(row1['status'], "Completed")
        self.assertEqual(len(row1['errors']), 0)

    def test_parse_offset_planner_with_base_date(self):
        content = build_offset_workbook()
        parser = GanttWorkbookParser(content, "offset.xlsx")
        res = parser.parse_sheet("Offset Planner")
        parser.close()

        self.assertEqual(res['detected_format'], GanttFormatDetector.FORMAT_OFFSET)
        self.assertEqual(res['base_date'], "2024-06-01")
        self.assertEqual(len(res['rows']), 3)

        # Row 1: offset 1 -> 2024-06-01, duration 1 -> finish 2024-06-01
        r1 = res['rows'][0]
        self.assertEqual(r1['activity'], "PO Receive")
        self.assertEqual(r1['planned_start'], "2024-06-01")
        self.assertEqual(r1['planned_finish'], "2024-06-01")
        self.assertEqual(len(r1['errors']), 0)

        # Row 2: offset 4 -> 2024-06-04, duration 1 -> finish 2024-06-04
        r2 = res['rows'][1]
        self.assertEqual(r2['activity'], "OEM Meeting")
        self.assertEqual(r2['planned_start'], "2024-06-04")
        self.assertEqual(r2['planned_finish'], "2024-06-04")
        self.assertEqual(len(r2['errors']), 0)

    def test_parse_zone_matrix(self):
        content = build_zone_matrix_workbook()
        parser = GanttWorkbookParser(content, "zone.xlsx")
        res = parser.parse_sheet("Zone Schedule")
        parser.close()

        self.assertEqual(res['detected_format'], GanttFormatDetector.FORMAT_ZONE_MATRIX)
        # Zone A has 2 valid tasks; Zone B has 1 valid and 1 missing date
        activities = [r['activity'] for r in res['rows']]
        self.assertIn("Zone A — COPPER PIPING", activities)
        self.assertIn("Zone A — DRAIN PIPE", activities)
        self.assertIn("Zone B — COPPER PIPING", activities)

        # Verified task has start and finish
        za_copper = next(r for r in res['rows'] if r['activity'] == "Zone A — COPPER PIPING")
        self.assertEqual(za_copper['planned_start'], "2024-09-01")
        self.assertEqual(za_copper['planned_finish'], "2024-09-03")
        self.assertEqual(len(za_copper['errors']), 0)

    def test_parse_visual_monthly_chart_leaves_dates_unresolved(self):
        content = build_visual_monthly_workbook()
        parser = GanttWorkbookParser(content, "visual.xlsx")
        res = parser.parse_sheet("Visual Timeline")
        parser.close()

        self.assertEqual(res['detected_format'], GanttFormatDetector.FORMAT_VISUAL_MONTHLY)
        self.assertEqual(len(res['rows']), 2)
        chillers = res['rows'][0]
        self.assertEqual(chillers['activity'], "Chillers")
        self.assertIsNone(chillers['planned_start'])
        self.assertIsNone(chillers['planned_finish'])
        # Must require manual correction
        self.assertTrue(any("require manual entry" in e for e in chillers['errors']))

    def test_finish_date_before_start_triggers_validation_error(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ACTIVITY", "PLAN START", "PLAN END"])
        ws.append(["Inverted Task", date(2024, 6, 10), date(2024, 6, 5)])
        buf = io.BytesIO()
        wb.save(buf)

        parser = GanttWorkbookParser(buf.getvalue(), "inv.xlsx")
        res = parser.parse_sheet("Sheet")
        parser.close()

        row = res['rows'][0]
        self.assertIn("planned_finish", row['errors_by_field'])
        self.assertTrue(any("earlier than" in e for e in row['errors']))

    def test_progress_percent_out_of_bounds(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ACTIVITY", "PLAN START", "PLAN END", "PERCENT COMPLETE"])
        ws.append(["Task A", date(2024, 6, 1), date(2024, 6, 2), 150])
        buf = io.BytesIO()
        wb.save(buf)

        parser = GanttWorkbookParser(buf.getvalue(), "prog.xlsx")
        res = parser.parse_sheet("Sheet")
        parser.close()

        row = res['rows'][0]
        self.assertIn("progress_percent", row['errors_by_field'])

    def test_duplicate_detection(self):
        # Create an existing task
        ProjectTask.objects.create(
            project=self.project,
            order=1,
            activity="Kickoff Meeting",
            planned_start=date(2024, 6, 1),
            planned_finish=date(2024, 6, 1)
        )

        content = build_structured_workbook()
        parser = GanttWorkbookParser(content, "schedule.xlsx")
        res = parser.parse_sheet("Schedule")
        parser.close()

        annotated = GanttDuplicateDetector.annotate_duplicates(self.project, res['rows'])
        kickoff_row = next(r for r in annotated if r['activity'] == "Kickoff Meeting")
        self.assertTrue(kickoff_row['is_duplicate'])
        self.assertIn("Probable duplicate", kickoff_row['warnings'][0])

    def test_zero_writes_during_preview(self):
        content = build_structured_workbook()
        initial_task_count = ProjectTask.objects.filter(project=self.project).count()

        parser = GanttWorkbookParser(content, "schedule.xlsx")
        res = parser.parse_sheet("Schedule")
        parser.close()

        # Staging batch creation must NOT create tasks
        batch = GanttImportStagingManager.create_batch(
            project=self.project,
            user=self.admin_user,
            filename="schedule.xlsx",
            file_sha256="mock_hash_1234567890",
            detected_format=res['detected_format'],
            selected_sheet="Schedule",
            staged_rows=res['rows'],
            stats=res['stats']
        )

        post_staging_count = ProjectTask.objects.filter(project=self.project).count()
        self.assertEqual(initial_task_count, post_staging_count)
        self.assertEqual(batch.status, "staged")

    def test_confirmed_import_executes_atomically_and_idempotently(self):
        content = build_structured_workbook()
        parser = GanttWorkbookParser(content, "schedule.xlsx")
        res = parser.parse_sheet("Schedule")
        parser.close()

        batch = GanttImportStagingManager.create_batch(
            project=self.project,
            user=self.admin_user,
            filename="schedule.xlsx",
            file_sha256="sha256_mock_123",
            detected_format=res['detected_format'],
            selected_sheet="Schedule",
            staged_rows=res['rows'],
            stats=res['stats']
        )

        # Confirm import
        result = GanttImportExecutor.confirm_import(batch, self.project, self.admin_user)
        self.assertEqual(result['status'], 'success')
        self.assertEqual(result['imported_count'], 3)

        tasks = ProjectTask.objects.filter(project=self.project).order_by('order')
        self.assertEqual(tasks.count(), 3)
        self.assertEqual(tasks[0].activity, "Kickoff Meeting")
        self.assertEqual(tasks[0].order, 1)
        self.assertEqual(tasks[1].order, 2)
        self.assertEqual(tasks[2].order, 3)

        # Check remarks contain trace metadata
        self.assertIn(str(batch.uuid)[:8], tasks[0].remarks)
        self.assertIn("Schedule", tasks[0].remarks)

        # Check single audit event created
        audit_events = AuditEvent.objects.filter(
            object_id=str(batch.uuid),
            action='gantt_imported'
        )
        self.assertEqual(audit_events.count(), 1)

        # Idempotency check: Confirming again returns existing result with 0 duplicate tasks
        repeat_result = GanttImportExecutor.confirm_import(batch, self.project, self.admin_user)
        self.assertEqual(repeat_result['status'], 'already_completed')
        self.assertEqual(ProjectTask.objects.filter(project=self.project).count(), 3)
        self.assertEqual(AuditEvent.objects.filter(object_id=str(batch.uuid), action='gantt_imported').count(), 1)

    def test_full_rollback_when_one_row_is_invalid(self):
        content = build_structured_workbook()
        parser = GanttWorkbookParser(content, "schedule.xlsx")
        res = parser.parse_sheet("Schedule")
        parser.close()

        # Inject an uncorrected invalid row marked without errors in stage data but broken data
        res['rows'].append({
            'activity': '',  # Empty activity violates validation
            'source_row': 99,
            'planned_start': '2024-06-01',
            'planned_finish': '2024-06-01',
            'duration_days': 1,
            'progress_percent': 0,
            'status': 'Not Started',
            'errors': []
        })

        batch = GanttImportStagingManager.create_batch(
            project=self.project,
            user=self.admin_user,
            filename="broken.xlsx",
            file_sha256="broken_sha_123",
            detected_format="structured_table",
            selected_sheet="Schedule",
            staged_rows=res['rows'],
            stats=res['stats']
        )

        with self.assertRaises(GanttImportError):
            GanttImportExecutor.confirm_import(batch, self.project, self.admin_user)

        # Total tasks must remain 0 (complete rollback)
        self.assertEqual(ProjectTask.objects.filter(project=self.project).count(), 0)


class GanttImportViewTests(TestCase):
    """Integration tests for HTMX endpoints, permissions, mobile UI, and accessibility."""

    def setUp(self):
        self.branch = Branch.objects.create(
            name="Mirpur",
            address="Dhaka",
            latitude=23.8103,
            longitude=90.4125,
            radius_meters=100
        )
        self.project_type, _ = ProjectType.objects.get_or_create(name="HVAC Installation")
        self.project = Project.objects.create(
            name="Beta Towers",
            client_name="Beta Group",
            project_type=self.project_type,
            start_date=date(2024, 7, 1),
            branch=self.branch
        )

        # Admin user (permitted)
        self.admin_user = User.objects.create_user(
            email="admin_view@example.com",
            phone="+8801700888001",
            password="testpassword123",
            role="admin"
        )

        # Regular staff user without task assign permission (restricted)
        self.staff_user = User.objects.create_user(
            email="staff_view@example.com",
            phone="+8801700888002",
            password="testpassword123",
            role="staff"
        )
        self.staff_profile = EmployeeProfile.objects.create(
            user=self.staff_user,
            full_name="Junior Tech",
            branch=self.branch,
            employee_id="EMP_G2",
            phone="+8801700888002",
            joined_date=date.today(),
            is_active=True
        )
        # Add staff to project members (member without assignment permission -> 403)
        self.project.project_members.add(self.staff_profile)

        # Another admin user for cross-user isolation test
        self.other_admin = User.objects.create_user(
            email="other_admin@example.com",
            phone="+8801700888003",
            password="testpassword123",
            role="admin"
        )

        self.client = Client()

    def test_anonymous_access_redirects(self):
        url = reverse('projects:project_gantt_import', kwargs={'pk': self.project.pk})
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 302)

    def test_project_member_without_assignment_permission_receives_403(self):
        self.client.login(username='+8801700888002', password='testpassword123')
        url = reverse('projects:project_gantt_import', kwargs={'pk': self.project.pk})
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 403)

    def test_admin_can_access_upload_page(self):
        self.client.login(username='+8801700888001', password='testpassword123')
        url = reverse('projects:project_gantt_import', kwargs={'pk': self.project.pk})
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Gantt Excel Import")
        self.assertContains(resp, "Upload Gantt Schedule")

    def test_upload_valid_file_returns_preview(self):
        self.client.login(username='+8801700888001', password='testpassword123')
        content = build_structured_workbook()
        file_obj = io.BytesIO(content)
        file_obj.name = "valid_schedule.xlsx"

        url = reverse('projects:project_gantt_import', kwargs={'pk': self.project.pk})
        resp = self.client.post(url, {'file': file_obj})
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Kickoff Meeting")
        self.assertContains(resp, "Site Assessment")
        self.assertContains(resp, "Ready to import")

    def test_upload_via_htmx_returns_partial(self):
        self.client.login(username='+8801700888001', password='testpassword123')
        content = build_structured_workbook()
        file_obj = io.BytesIO(content)
        file_obj.name = "htmx_schedule.xlsx"

        url = reverse('projects:project_gantt_import', kwargs={'pk': self.project.pk})
        resp = self.client.post(url, {'file': file_obj}, HTTP_HX_REQUEST='true')
        self.assertEqual(resp.status_code, 200)
        self.assertTemplateUsed(resp, 'projects/partials/gantt_import_preview.html')
        self.assertContains(resp, "gantt-preview-wrapper")

    def test_upload_corrupt_file_returns_accessible_field_error(self):
        self.client.login(username='+8801700888001', password='testpassword123')
        corrupt_obj = io.BytesIO(b"NOT_A_VALID_FILE")
        corrupt_obj.name = "corrupt.xlsx"

        url = reverse('projects:project_gantt_import', kwargs={'pk': self.project.pk})
        resp = self.client.post(url, {'file': corrupt_obj})
        self.assertEqual(resp.status_code, 400)
        self.assertContains(resp, "Workbook Upload Failed", status_code=400)

    def test_preview_in_place_edit_via_htmx(self):
        self.client.login(username='+8801700888001', password='testpassword123')
        content = build_structured_workbook()
        parser = GanttWorkbookParser(content, "sched.xlsx")
        res = parser.parse_sheet("Schedule")
        parser.close()

        batch = GanttImportStagingManager.create_batch(
            project=self.project,
            user=self.admin_user,
            filename="sched.xlsx",
            file_sha256="sha_preview_test",
            detected_format=res['detected_format'],
            selected_sheet="Schedule",
            staged_rows=res['rows'],
            stats=res['stats']
        )

        preview_url = reverse('projects:project_gantt_import_preview', kwargs={'pk': self.project.pk, 'batch_id': batch.uuid})
        resp = self.client.post(
            preview_url,
            {
                'action': 'update_field',
                'row_idx': 0,
                'field': 'activity',
                'value': 'Updated Kickoff Activity'
            },
            HTTP_HX_REQUEST='true'
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Updated Kickoff Activity")

    def test_preview_row_exclusion_toggle(self):
        self.client.login(username='+8801700888001', password='testpassword123')
        content = build_structured_workbook()
        parser = GanttWorkbookParser(content, "sched.xlsx")
        res = parser.parse_sheet("Schedule")
        parser.close()

        batch = GanttImportStagingManager.create_batch(
            project=self.project,
            user=self.admin_user,
            filename="sched.xlsx",
            file_sha256="sha_exclude_test",
            detected_format=res['detected_format'],
            selected_sheet="Schedule",
            staged_rows=res['rows'],
            stats=res['stats']
        )

        preview_url = reverse('projects:project_gantt_import_preview', kwargs={'pk': self.project.pk, 'batch_id': batch.uuid})
        resp = self.client.post(
            preview_url,
            {
                'action': 'toggle_exclude',
                'row_idx': 0
            },
            HTTP_HX_REQUEST='true'
        )
        self.assertEqual(resp.status_code, 200)
        batch.refresh_from_db()
        self.assertTrue(batch.staged_data['rows'][0]['excluded'])

    def test_cross_user_staged_batch_access_blocked(self):
        self.client.login(username='+8801700888001', password='testpassword123')
        content = build_structured_workbook()
        parser = GanttWorkbookParser(content, "sched.xlsx")
        res = parser.parse_sheet("Schedule")
        parser.close()

        batch = GanttImportStagingManager.create_batch(
            project=self.project,
            user=self.admin_user,
            filename="sched.xlsx",
            file_sha256="sha_cross_user",
            detected_format=res['detected_format'],
            selected_sheet="Schedule",
            staged_rows=res['rows'],
            stats=res['stats']
        )

        # Login as regular staff user and try to access admin's staged batch -> 403
        self.client.login(username='+8801700888002', password='testpassword123')
        url = reverse('projects:project_gantt_import_preview', kwargs={'pk': self.project.pk, 'batch_id': batch.uuid})
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 403)

    def test_full_confirm_workflow_via_htmx(self):
        self.client.login(username='+8801700888001', password='testpassword123')
        content = build_structured_workbook()
        parser = GanttWorkbookParser(content, "sched.xlsx")
        res = parser.parse_sheet("Schedule")
        parser.close()

        batch = GanttImportStagingManager.create_batch(
            project=self.project,
            user=self.admin_user,
            filename="sched.xlsx",
            file_sha256="sha_confirm_test",
            detected_format=res['detected_format'],
            selected_sheet="Schedule",
            staged_rows=res['rows'],
            stats=res['stats']
        )

        confirm_url = reverse('projects:project_gantt_import_confirm', kwargs={'pk': self.project.pk, 'batch_id': batch.uuid})
        resp = self.client.post(confirm_url, HTTP_HX_REQUEST='true')
        self.assertEqual(resp.status_code, 200)
        self.assertTemplateUsed(resp, 'projects/partials/gantt_import_success.html')
        self.assertContains(resp, "Gantt Import Successful")
        self.assertContains(resp, "Open Gantt Chart")

        # Confirm 3 tasks created
        self.assertEqual(ProjectTask.objects.filter(project=self.project).count(), 3)
